"""Normalize formatting for the nonregistered menu Excel export."""

from __future__ import annotations

from functools import wraps
from io import BytesIO

from flask import Flask
from openpyxl import load_workbook

import blueprints.order_tool as order_tool_module


def install_nonregistered_menu_export_format_fix(app: Flask):
    """Force every exported data row to use the template's first-row style."""

    endpoint = "order_tool.school_menus_export"
    original_view = app.view_functions.get(endpoint)
    if original_view is None:
        return

    @wraps(original_view)
    def normalized_export(*args, **kwargs):
        response = app.make_response(original_view(*args, **kwargs))
        if response.status_code != 200:
            return response

        # send_file() responses use direct passthrough; disable it so the small
        # workbook can be normalized before being returned to the browser.
        response.direct_passthrough = False
        raw = response.get_data()
        if not raw:
            return response

        workbook = load_workbook(BytesIO(raw))
        sheet = workbook.active
        first_data_row = 2
        for row_number in range(first_data_row + 1, sheet.max_row + 1):
            order_tool_module._copy_nonregistered_template_row(
                sheet, first_data_row, row_number
            )

        output = BytesIO()
        workbook.save(output)
        response.set_data(output.getvalue())
        return response

    app.view_functions[endpoint] = normalized_export
