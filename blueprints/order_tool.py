"""團膳菜單 / 配方 / 採購叫貨正式模組。

核心流程：
Recipe BOM（每人 AP 數量，可用 g 或 個）→ 中央菜單 → 學校人數
→ 食材需求彙總 → 供應商採購草稿 → 人工調整 → Confirm snapshot。
"""

from __future__ import annotations

from copy import copy
import secrets
import re
from collections import defaultdict
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation, ROUND_CEILING
from io import BytesIO
from pathlib import Path

from flask import (
    Blueprint,
    abort,
    current_app,
    flash,
    redirect,
    render_template,
    request,
    send_file,
    session,
    url_for,
)
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import selectinload
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from itsdangerous import BadSignature, URLSafeSerializer

from extensions import db
from models import (
    KitchenDailyDishNote,
    KitchenIngredient,
    KitchenMenuAssignment,
    KitchenMenuPlan,
    KitchenMenuPlanItem,
    KitchenPurchaseOrder,
    KitchenPurchaseOrderItem,
    KitchenRecipe,
    KitchenRecipeIngredient,
    KitchenSchool,
    KitchenSupplier,
    KitchenSupplierItem,
)

order_bp = Blueprint("order_tool", __name__)

CATEGORIES = ("主食", "主菜", "副菜", "青菜", "湯品", "點心", "其他")
SCHOOL_SERVICE_STATUSES = ("serving", "no_service")
MEAL_TYPES = ("早餐", "午餐", "晚餐", "點心")
PURCHASE_UNITS = ("kg", "箱", "包", "斤", "瓶", "個", "袋", "桶")
PACKAGE_UNITS = ("箱", "包", "袋", "桶", "瓶", "罐", "盒", "籃", "籠", "件", "組", "個", "支", "kg", "斤")
BASE_UNITS = ("g", "個")
WEEKDAY_LABELS = ("週一", "週二", "週三", "週四", "週五", "週六", "週日")
MENU_HEADER_CATEGORIES = {
    "主食": "主食",
    "主菜": "主菜",
    "副菜": "副菜",
    "配菜": "副菜",
    "蔬菜": "青菜",
    "青菜": "青菜",
    "湯品": "湯品",
    "湯": "湯品",
    "點心": "點心",
}
NONREGISTERED_MENU_TEMPLATE = Path("static/非登合菜名範本.xlsx")
NONREGISTERED_MENU_DEFAULTS = (4, 2, 1.7, 0, 0, 2, 563)
NONREGISTERED_MENU_COLUMNS = {
    "主食": (11, 12),
    "主菜": (13, 14, 15, 16),
    "副菜": (17, 18, 19, 20, 21, 22),
    "青菜": (23,),
    "湯品": (24,),
    "點心": (25, 26, 27),
    "其他": (25, 26, 27),
}
CATEGORY_ORDER = {category: index for index, category in enumerate(CATEGORIES)}


def _menu_item_category_sort_key(item: KitchenMenuPlanItem):
    category = item.recipe.category or "其他"
    return (
        CATEGORY_ORDER.get(category, len(CATEGORIES)),
        item.sort_order,
        item.recipe.name.casefold(),
    )


# ─────────────────────────────────────────────
# Security / shared helpers
# ─────────────────────────────────────────────


def _csrf_token() -> str:
    token = session.get("_kitchen_csrf")
    if not token:
        token = secrets.token_urlsafe(32)
        session["_kitchen_csrf"] = token
    return token


@order_bp.before_request
def _protect_kitchen():
    if request.method == "POST" and current_app.config.get("KITCHEN_CSRF_ENABLED", True):
        expected = session.get("_kitchen_csrf", "")
        received = request.form.get("_csrf_token", "")
        if not expected or not received or not secrets.compare_digest(expected, received):
            abort(400, description="安全驗證已失效，請重新整理頁面再操作。")
    return None


@order_bp.errorhandler(400)
def _bad_request(error):
    flash(getattr(error, "description", None) or "輸入資料不正確。", "error")
    return redirect(request.referrer or url_for("order_tool.index"))


@order_bp.context_processor
def _template_helpers():
    production_nav_date, production_nav_available = _production_nav_state()
    return {
        "csrf_token": _csrf_token,
        "status_label": _status_label,
        "order_total": _order_total,
        "recipe_total_g": _recipe_total_g,
        "component_cost": _component_cost,
        "trim_decimal": _trim_decimal,
        "base_unit_label": _base_unit_label,
        "required_display": _required_display,
        "production_nav_date": production_nav_date,
        "production_nav_available": production_nav_available,
    }


def _status_label(status: str) -> str:
    return {"draft": "草稿", "confirmed": "已確認", "cancelled": "已取消"}.get(status, status or "-")


def _base_unit_label(ingredient: KitchenIngredient) -> str:
    return ingredient.base_unit or "g"


def _decimal(raw, *, default=None) -> Decimal | None:
    if raw is None or str(raw).strip() == "":
        return default
    try:
        value = Decimal(str(raw).strip())
        return value if value.is_finite() else default
    except (InvalidOperation, ValueError):
        return default


def _int(raw, *, default=None) -> int | None:
    try:
        return int(str(raw).strip())
    except (TypeError, ValueError):
        return default


def _date(raw, *, default=None) -> date | None:
    if not raw:
        return default
    try:
        return date.fromisoformat(str(raw))
    except ValueError:
        return default


def _recipe_total_g(recipe: KitchenRecipe) -> Decimal:
    """只加總 g 型食材，避免把 1 個直接誤當 1g 加進總生料重量。"""
    return sum(
        (x.grams_per_person or Decimal("0") for x in recipe.ingredients if (x.ingredient.base_unit or "g") == "g"),
        Decimal("0"),
    )


def _component_cost(component: KitchenRecipeIngredient) -> Decimal:
    ing = component.ingredient
    units_per_purchase = ing.grams_per_purchase_unit or Decimal("0")
    if units_per_purchase <= 0:
        return Decimal("0")
    return (component.grams_per_person or Decimal("0")) / units_per_purchase * (ing.unit_price or Decimal("0"))


def _recipe_cost(recipe: KitchenRecipe) -> Decimal:
    return sum((_component_cost(x) for x in recipe.ingredients), Decimal("0"))


def _order_total(order: KitchenPurchaseOrder) -> Decimal:
    return sum((x.amount or Decimal("0") for x in order.items), Decimal("0"))


def _trim_decimal(value) -> str:
    d = _decimal(value, default=Decimal("0")) or Decimal("0")
    text = format(d, "f")
    if "." in text:
        text = text.rstrip("0").rstrip(".")
    return text or "0"


_WEIGHT_UNIT_KG = {
    "kg": Decimal("1"),
    "公斤": Decimal("1"),
    "g": Decimal("0.001"),
    "公克": Decimal("0.001"),
    "克": Decimal("0.001"),
    "斤": Decimal("0.6"),
    "台斤": Decimal("0.6"),
}


def _clean_conversion_unit(raw: str) -> str:
    unit = (raw or "").strip().lower().replace("臺", "台")
    return {"公斤": "kg", "公克": "g", "克": "g", "台斤": "斤"}.get(unit, unit)


def _convert_unit_amount(amount: Decimal, source_unit: str, target_unit: str) -> Decimal | None:
    source = _clean_conversion_unit(source_unit)
    target = _clean_conversion_unit(target_unit)
    if source == target:
        return amount
    if source in _WEIGHT_UNIT_KG and target in _WEIGHT_UNIT_KG:
        return amount * _WEIGHT_UNIT_KG[source] / _WEIGHT_UNIT_KG[target]
    return None


def _package_conversion_rule(raw: str | None, purchase_unit: str) -> dict | None:
    """Turn e.g. 1箱＝12kg or 2箱＝24kg into purchase units per package."""
    if not raw:
        return None
    parts = re.split(r"[=＝]", raw, maxsplit=1)
    if len(parts) != 2:
        return None

    def parse_side(value):
        match = re.fullmatch(r"\s*(\d+(?:\.\d+)?)\s*([^\d\s]+)\s*", value)
        if not match:
            return None
        qty = _decimal(match.group(1), default=None)
        return (qty, match.group(2).strip()) if qty and qty > 0 else None

    left, right = parse_side(parts[0]), parse_side(parts[1])
    if not left or not right:
        return None
    left_as_purchase = _convert_unit_amount(left[0], left[1], purchase_unit)
    right_as_purchase = _convert_unit_amount(right[0], right[1], purchase_unit)
    if right_as_purchase is not None:
        return {
            "package_unit": left[1],
            "purchase_per_package": right_as_purchase / left[0],
        }
    if left_as_purchase is not None:
        return {
            "package_unit": right[1],
            "purchase_per_package": left_as_purchase / right[0],
        }
    return None


def _supplier_item_match(supplier_id: int | None, ingredient_id: int | None, ingredient_name: str):
    if not supplier_id:
        return None
    rows = KitchenSupplierItem.query.filter_by(supplier_id=supplier_id, active=True).all()
    if ingredient_id:
        exact = next((row for row in rows if row.ingredient_id == ingredient_id), None)
        if exact:
            return exact
    normalized = (ingredient_name or "").strip().casefold()
    return next((row for row in rows if row.name.strip().casefold() == normalized), None)


def _package_qty_from_rule(actual_qty, rule):
    actual = _decimal(actual_qty, default=None)
    factor = rule.get("purchase_per_package") if rule else None
    if actual is None or not factor or factor <= 0:
        return None
    return actual / factor


def _remember_supplier_conversion(
    supplier: KitchenSupplier,
    item: KitchenPurchaseOrderItem,
    actual_qty: Decimal,
    package_qty: Decimal,
    package_unit: str,
):
    """Persist an edited procurement equation back to the supplier catalog."""
    supplier_item = _supplier_item_match(
        supplier.id,
        item.ingredient_id,
        item.ingredient_name_snapshot,
    )
    if supplier_item is None:
        supplier_item = KitchenSupplierItem(
            supplier_id=supplier.id,
            ingredient_id=item.ingredient_id,
            name=item.ingredient_name_snapshot[:120],
            unit=item.purchase_unit_snapshot[:20],
            last_unit_price=item.unit_price_snapshot or Decimal("0"),
            manual_override=True,
            active=True,
        )
        db.session.add(supplier_item)
        db.session.flush()
    elif supplier_item.ingredient_id is None:
        supplier_item.ingredient_id = item.ingredient_id

    purchase_per_package = actual_qty / package_qty
    old_rule = _package_conversion_rule(supplier_item.package_conversion, item.purchase_unit_snapshot)
    same_rule = bool(
        old_rule
        and _clean_conversion_unit(old_rule["package_unit"]) == _clean_conversion_unit(package_unit)
        and abs(old_rule["purchase_per_package"] - purchase_per_package)
        <= max(Decimal("0.0001"), old_rule["purchase_per_package"] * Decimal("0.0001"))
    )
    changed = not same_rule
    if changed:
        display_factor = purchase_per_package.quantize(Decimal("0.0001"))
        supplier_item.package_conversion = (
            f"1{package_unit}＝{_trim_decimal(display_factor)}{item.purchase_unit_snapshot}"
        )[:120]
    supplier_item.manual_override = True
    supplier_item.active = True
    return supplier_item, changed


def _required_display(item: KitchenPurchaseOrderItem) -> str:
    amount = item.required_grams or Decimal("0")
    base = item.base_unit_snapshot or "g"
    if base == "g":
        return f"{_trim_decimal(amount / Decimal('1000'))} kg（{_trim_decimal(amount)} g）"
    return f"{_trim_decimal(amount)} {base}"


def _round_up_increment(value: Decimal, increment: Decimal) -> Decimal:
    if value <= 0:
        return Decimal("0")
    if increment <= 0:
        increment = Decimal("1")
    return (value / increment).to_integral_value(rounding=ROUND_CEILING) * increment


def _commit(success: str, redirect_to: str, **values):
    try:
        db.session.commit()
        flash(success, "success")
    except IntegrityError:
        db.session.rollback()
        flash("資料重複或違反資料關聯，請檢查名稱與設定。", "error")
    return redirect(url_for(redirect_to, **values))


def _active_confirmed_orders(service_date: date) -> bool:
    return KitchenPurchaseOrder.query.filter_by(service_date=service_date, status="confirmed").first() is not None


def _require_draft_plan(plan: KitchenMenuPlan) -> bool:
    if plan.status != "draft":
        flash("這張菜單已確認，請先重開草稿才能修改。", "warning")
        return False
    return True


def _school_meal_conflict(plan: KitchenMenuPlan, school_id: int) -> KitchenMenuAssignment | None:
    """同校、同日、同餐別只能出現在一張菜單，避免採購重複計算。"""
    return (
        KitchenMenuAssignment.query.join(KitchenMenuPlan)
        .filter(
            KitchenMenuAssignment.school_id == school_id,
            KitchenMenuAssignment.plan_id != plan.id,
            KitchenMenuPlan.service_date == plan.service_date,
            KitchenMenuPlan.meal_type == plan.meal_type,
        )
        .first()
    )


def _plan_has_assignment_conflict(plan: KitchenMenuPlan, service_date: date, meal_type: str) -> bool:
    school_ids = [x.school_id for x in plan.assignments]
    if not school_ids:
        return False
    return (
        KitchenMenuAssignment.query.join(KitchenMenuPlan)
        .filter(
            KitchenMenuAssignment.school_id.in_(school_ids),
            KitchenMenuAssignment.plan_id != plan.id,
            KitchenMenuPlan.service_date == service_date,
            KitchenMenuPlan.meal_type == meal_type,
        )
        .first()
        is not None
    )


# ─────────────────────────────────────────────
# Dashboard
# ─────────────────────────────────────────────


@order_bp.get("/")
def index():
    recent_orders = KitchenPurchaseOrder.query.order_by(
        KitchenPurchaseOrder.service_date.desc(), KitchenPurchaseOrder.id.desc()
    ).limit(10).all()
    return render_template(
        "kitchen/dashboard.html",
        recent_orders=recent_orders,
        recipe_count=KitchenRecipe.query.filter_by(active=True).count(),
        ingredient_count=KitchenIngredient.query.filter_by(active=True).count(),
        school_count=KitchenSchool.query.filter_by(active=True).count(),
        supplier_count=KitchenSupplier.query.filter_by(active=True).count(),
        pending_component_count=KitchenRecipeIngredient.query.filter(
            KitchenRecipeIngredient.quantity_status == "pending"
        ).count(),
    )


# ─────────────────────────────────────────────
# School / Supplier CRUD
# ─────────────────────────────────────────────


@order_bp.route("/schools", methods=["GET", "POST"])
def schools():
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        if not name:
            flash("學校名稱不可空白。", "error")
            return redirect(url_for("order_tool.schools"))
        default_headcount = _int(request.form.get("default_headcount"), default=0)
        if default_headcount is None or default_headcount < 0:
            flash("平常人數不可為負數。", "error")
            return redirect(url_for("order_tool.schools"))
        default_vegetarian_headcount = _int(request.form.get("default_vegetarian_headcount"), default=0)
        if default_vegetarian_headcount is None or default_vegetarian_headcount < 0:
            flash("平常素食人數不可為負數。", "error")
            return redirect(url_for("order_tool.schools"))
        db.session.add(KitchenSchool(
            name=name,
            code=request.form.get("code", "").strip() or None,
            default_headcount=default_headcount,
            default_vegetarian_headcount=default_vegetarian_headcount,
        ))
        return _commit("學校已新增。", "order_tool.schools")
    q = request.args.get("q", "").strip()
    query = KitchenSchool.query
    if q:
        query = query.filter(KitchenSchool.name.ilike(f"%{q}%"))
    rows = query.order_by(KitchenSchool.active.desc(), KitchenSchool.name).all()
    edit_row = db.session.get(KitchenSchool, _int(request.args.get("edit"), default=0)) if request.args.get("edit") else None
    return render_template("kitchen/schools.html", rows=rows, edit_row=edit_row, q=q)


@order_bp.post("/schools/<int:school_id>/update")
def school_update(school_id: int):
    row = db.session.get(KitchenSchool, school_id)
    if not row:
        abort(404)
    name = request.form.get("name", "").strip()
    if not name:
        flash("學校名稱不可空白。", "error")
        return redirect(url_for("order_tool.schools", edit=school_id))
    row.name = name
    row.code = request.form.get("code", "").strip() or None
    default_headcount = _int(request.form.get("default_headcount"), default=0)
    if default_headcount is None or default_headcount < 0:
        flash("平常人數不可為負數。", "error")
        return redirect(url_for("order_tool.schools", edit=school_id))
    row.default_headcount = default_headcount
    default_vegetarian_headcount = _int(request.form.get("default_vegetarian_headcount"), default=0)
    if default_vegetarian_headcount is None or default_vegetarian_headcount < 0:
        flash("平常素食人數不可為負數。", "error")
        return redirect(url_for("order_tool.schools", edit=school_id))
    row.default_vegetarian_headcount = default_vegetarian_headcount
    return _commit("學校資料已更新。", "order_tool.schools")


@order_bp.post("/schools/<int:school_id>/toggle")
def school_toggle(school_id: int):
    row = db.session.get(KitchenSchool, school_id)
    if not row:
        abort(404)
    row.active = not row.active
    return _commit("學校狀態已更新。", "order_tool.schools")


@order_bp.route("/suppliers", methods=["GET", "POST"])
def suppliers():
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        if not name:
            flash("廠商名稱不可空白。", "error")
            return redirect(url_for("order_tool.suppliers"))
        db.session.add(KitchenSupplier(
            name=name,
            phone=request.form.get("phone", "").strip() or None,
            mobile=request.form.get("mobile", "").strip() or None,
            fax=request.form.get("fax", "").strip() or None,
            contact=request.form.get("contact", "").strip() or None,
            address=request.form.get("address", "").strip() or None,
            note=request.form.get("note", "").strip() or None,
        ))
        return _commit("廠商已新增。", "order_tool.suppliers")
    q = request.args.get("q", "").strip()
    query = KitchenSupplier.query
    if q:
        query = query.filter(KitchenSupplier.name.ilike(f"%{q}%"))
    rows = query.options(selectinload(KitchenSupplier.items)).order_by(
        KitchenSupplier.active.desc(), KitchenSupplier.name
    ).all()
    edit_row = db.session.get(KitchenSupplier, _int(request.args.get("edit"), default=0)) if request.args.get("edit") else None
    return render_template("kitchen/suppliers.html", rows=rows, edit_row=edit_row, q=q)


@order_bp.post("/suppliers/<int:supplier_id>/update")
def supplier_update(supplier_id: int):
    row = db.session.get(KitchenSupplier, supplier_id)
    if not row:
        abort(404)
    name = request.form.get("name", "").strip()
    if not name:
        flash("廠商名稱不可空白。", "error")
        return redirect(url_for("order_tool.suppliers", edit=supplier_id))
    row.name = name
    row.phone = request.form.get("phone", "").strip() or None
    row.mobile = request.form.get("mobile", "").strip() or None
    row.fax = request.form.get("fax", "").strip() or None
    row.contact = request.form.get("contact", "").strip() or None
    row.address = request.form.get("address", "").strip() or None
    row.note = request.form.get("note", "").strip() or None
    return _commit("廠商資料已更新。", "order_tool.suppliers")


@order_bp.post("/suppliers/<int:supplier_id>/toggle")
def supplier_toggle(supplier_id: int):
    row = db.session.get(KitchenSupplier, supplier_id)
    if not row:
        abort(404)
    row.active = not row.active
    return _commit("廠商狀態已更新。", "order_tool.suppliers")


@order_bp.get("/suppliers/<int:supplier_id>")
def supplier_detail(supplier_id: int):
    row = db.session.get(KitchenSupplier, supplier_id)
    if not row:
        abort(404)
    q = request.args.get("q", "").strip()
    item_query = KitchenSupplierItem.query.filter_by(supplier_id=supplier_id, active=True)
    if q:
        item_query = item_query.filter(KitchenSupplierItem.name.ilike(f"%{q}%"))
    supplier_items = item_query.order_by(
        KitchenSupplierItem.last_purchase_date.desc(), KitchenSupplierItem.name
    ).all()
    return render_template(
        "kitchen/supplier_detail.html",
        supplier=row,
        supplier_items=supplier_items,
        q=q,
    )


@order_bp.post("/suppliers/<int:supplier_id>/items")
def supplier_item_create(supplier_id: int):
    supplier = db.session.get(KitchenSupplier, supplier_id)
    if not supplier:
        abort(404)
    name = request.form.get("name", "").strip()
    unit = request.form.get("unit", "").strip()
    package_conversion = request.form.get("package_conversion", "").strip()
    price = _decimal(request.form.get("last_unit_price"), default=None)
    if not name or not unit or price is None or price < 0:
        flash("食材、單位或最近單價不正確。", "error")
        return redirect(url_for("order_tool.supplier_detail", supplier_id=supplier_id))

    row = KitchenSupplierItem.query.filter_by(supplier_id=supplier_id, name=name[:120]).one_or_none()
    if row is None:
        row = KitchenSupplierItem(supplier_id=supplier_id, name=name[:120], unit=unit[:20])
        db.session.add(row)
    row.unit = unit[:20]
    row.package_conversion = package_conversion[:120] or None
    row.last_unit_price = price
    row.manual_override = True
    row.active = True
    return _commit("廠商品項已新增。", "order_tool.supplier_detail", supplier_id=supplier_id)


@order_bp.post("/supplier-items/<int:item_id>/update")
def supplier_item_update(item_id: int):
    row = db.session.get(KitchenSupplierItem, item_id)
    if not row:
        abort(404)
    name = request.form.get("name", "").strip()
    unit = request.form.get("unit", "").strip()
    package_conversion = request.form.get("package_conversion", "").strip()
    price = _decimal(request.form.get("last_unit_price"), default=None)
    if not name or not unit or price is None or price < 0:
        flash("食材、單位或最近單價不正確。", "error")
        return redirect(url_for("order_tool.supplier_detail", supplier_id=row.supplier_id))
    row.name = name[:120]
    row.unit = unit[:20]
    row.package_conversion = package_conversion[:120] or None
    row.last_unit_price = price
    row.manual_override = True
    return _commit("廠商品項已更新。", "order_tool.supplier_detail", supplier_id=row.supplier_id)


@order_bp.post("/supplier-items/<int:item_id>/delete")
def supplier_item_delete(item_id: int):
    row = db.session.get(KitchenSupplierItem, item_id)
    if not row:
        abort(404)
    supplier_id = row.supplier_id
    row.active = False
    row.manual_override = True
    db.session.commit()
    flash("品項已從這家廠商移除；食材主檔不受影響。", "success")
    return redirect(url_for("order_tool.supplier_detail", supplier_id=supplier_id))


# ─────────────────────────────────────────────
# Ingredient CRUD
# ─────────────────────────────────────────────


def _ingredient_form_values():
    name = request.form.get("name", "").strip()
    base_unit = request.form.get("base_unit", "g").strip()
    purchase_unit = request.form.get("purchase_unit", "kg").strip()
    units_per_purchase = _decimal(request.form.get("grams_per_purchase_unit"))
    unit_price = _decimal(request.form.get("unit_price"))
    increment = _decimal(request.form.get("order_increment"))
    note = request.form.get("note", "").strip() or None

    if not name:
        return None, "食材名稱不可空白。"
    if base_unit not in BASE_UNITS:
        return None, "配方基本單位不正確。"
    if purchase_unit not in PURCHASE_UNITS:
        return None, "採購單位不正確。"
    if units_per_purchase is None or units_per_purchase <= 0:
        return None, f"每採購單位包含的 {base_unit} 數量必須大於 0。"
    if unit_price is None or unit_price < 0:
        return None, "單價不可為負數。"
    if increment is None or increment <= 0:
        return None, "最小叫貨增量必須大於 0。"

    supplier_name_raw = request.form.get("supplier_name")
    if supplier_name_raw is None:
        supplier_id = _int(request.form.get("supplier_id"), default=None)
        if supplier_id is not None and not db.session.get(KitchenSupplier, supplier_id):
            return None, "找不到指定廠商。"
    else:
        supplier_name = supplier_name_raw.strip()
        if len(supplier_name) > 100:
            return None, "廠商名稱不可超過 100 個字。"
        supplier = None
        if supplier_name:
            supplier = KitchenSupplier.query.filter(
                db.func.lower(KitchenSupplier.name) == supplier_name.lower()
            ).first()
            if supplier is None:
                supplier = KitchenSupplier(
                    name=supplier_name,
                    note="由食材主檔新增",
                    active=True,
                )
                db.session.add(supplier)
                db.session.flush()
        supplier_id = supplier.id if supplier else None
    return {
        "name": name,
        "supplier_id": supplier_id,
        "base_unit": base_unit,
        "purchase_unit": purchase_unit,
        "grams_per_purchase_unit": units_per_purchase,
        "unit_price": unit_price,
        "order_increment": increment,
        "note": note,
    }, None


@order_bp.route("/ingredients", methods=["GET", "POST"])
def ingredients():
    if request.method == "POST":
        values, error = _ingredient_form_values()
        if error:
            flash(error, "error")
            return redirect(url_for("order_tool.ingredients"))
        db.session.add(KitchenIngredient(**values))
        return _commit("食材已新增。", "order_tool.ingredients")
    q = request.args.get("q", "").strip()
    query = KitchenIngredient.query
    if q:
        query = query.filter(KitchenIngredient.name.ilike(f"%{q}%"))
    rows = query.order_by(KitchenIngredient.active.desc(), KitchenIngredient.name).all()
    suppliers_all = KitchenSupplier.query.order_by(KitchenSupplier.active.desc(), KitchenSupplier.name).all()
    edit_row = db.session.get(KitchenIngredient, _int(request.args.get("edit"), default=0)) if request.args.get("edit") else None
    return render_template(
        "kitchen/ingredients.html",
        rows=rows,
        suppliers=suppliers_all,
        edit_row=edit_row,
        units=PURCHASE_UNITS,
        base_units=BASE_UNITS,
        q=q,
    )


@order_bp.post("/ingredients/<int:ingredient_id>/update")
def ingredient_update(ingredient_id: int):
    row = db.session.get(KitchenIngredient, ingredient_id)
    if not row:
        abort(404)
    values, error = _ingredient_form_values()
    if error:
        flash(error, "error")
        return redirect(url_for("order_tool.ingredients", edit=ingredient_id))
    for key, value in values.items():
        setattr(row, key, value)
    return _commit("食材資料已更新。", "order_tool.ingredients")


@order_bp.post("/ingredients/<int:ingredient_id>/toggle")
def ingredient_toggle(ingredient_id: int):
    row = db.session.get(KitchenIngredient, ingredient_id)
    if not row:
        abort(404)
    row.active = not row.active
    return _commit("食材狀態已更新。", "order_tool.ingredients")


# ─────────────────────────────────────────────
# Recipe CRUD / BOM
# ─────────────────────────────────────────────


def _recipe_form_values():
    name = request.form.get("name", "").strip()
    category = request.form.get("category", "主菜").strip()
    output_raw = request.form.get("serving_output_g", "").strip()
    serving_output = _decimal(output_raw, default=None) if output_raw else None
    if not name:
        return None, "菜色名稱不可空白。"
    if category not in CATEGORIES:
        return None, "菜色分類不正確。"
    if serving_output is not None and serving_output < 0:
        return None, "打菜量不可為負數。"
    return {
        "name": name,
        "category": category,
        "serving_output_g": serving_output,
        "note": request.form.get("note", "").strip() or None,
    }, None


@order_bp.route("/recipes", methods=["GET", "POST"])
def recipes():
    if request.method == "POST":
        values, error = _recipe_form_values()
        if error:
            flash(error, "error")
            return redirect(url_for("order_tool.recipes"))
        recipe = KitchenRecipe(**values)
        db.session.add(recipe)
        try:
            db.session.commit()
            flash("菜色已建立，請設定每人配方。", "success")
            return redirect(url_for("order_tool.recipe_detail", recipe_id=recipe.id))
        except IntegrityError:
            db.session.rollback()
            flash("菜色名稱已存在。", "error")
            return redirect(url_for("order_tool.recipes"))
    q = request.args.get("q", "").strip()
    query = KitchenRecipe.query
    if q:
        query = query.filter(KitchenRecipe.name.ilike(f"%{q}%"))
    rows = query.order_by(KitchenRecipe.active.desc(), KitchenRecipe.category, KitchenRecipe.name).all()
    edit_row = db.session.get(KitchenRecipe, _int(request.args.get("edit"), default=0)) if request.args.get("edit") else None
    return render_template("kitchen/recipes.html", rows=rows, edit_row=edit_row, categories=CATEGORIES, q=q)


@order_bp.post("/recipes/<int:recipe_id>/update")
def recipe_update(recipe_id: int):
    row = db.session.get(KitchenRecipe, recipe_id)
    if not row:
        abort(404)
    values, error = _recipe_form_values()
    if error:
        flash(error, "error")
        return redirect(url_for("order_tool.recipes", edit=recipe_id))
    for key, value in values.items():
        setattr(row, key, value)
    return _commit("菜色資料已更新。", "order_tool.recipes")


@order_bp.post("/recipes/<int:recipe_id>/toggle")
def recipe_toggle(recipe_id: int):
    row = db.session.get(KitchenRecipe, recipe_id)
    if not row:
        abort(404)
    row.active = not row.active
    return _commit("菜色狀態已更新。", "order_tool.recipes")


@order_bp.get("/recipes/<int:recipe_id>")
def recipe_detail(recipe_id: int):
    recipe = db.session.get(KitchenRecipe, recipe_id)
    if not recipe:
        abort(404)
    ingredients_all = KitchenIngredient.query.filter_by(active=True).order_by(KitchenIngredient.name).all()
    return render_template(
        "kitchen/recipe_detail.html",
        recipe=recipe,
        ingredients=ingredients_all,
        ingredient_options=[
            {
                "id": ingredient.id,
                "name": ingredient.name,
                "base_unit": ingredient.base_unit or "g",
                "purchase_unit": ingredient.purchase_unit,
            }
            for ingredient in ingredients_all
        ],
        categories=CATEGORIES,
        total_g=_recipe_total_g(recipe),
        total_cost=_recipe_cost(recipe),
    )


@order_bp.post("/recipes/<int:recipe_id>/category")
def recipe_category_update(recipe_id: int):
    recipe = db.session.get(KitchenRecipe, recipe_id)
    if not recipe:
        abort(404)
    category = request.form.get("category", "").strip()
    if category not in CATEGORIES:
        flash("菜色分類不正確。", "error")
        return redirect(url_for("order_tool.recipe_detail", recipe_id=recipe_id))
    recipe.category = category
    return _commit("菜色分類已更新。", "order_tool.recipe_detail", recipe_id=recipe_id)


@order_bp.post("/recipes/<int:recipe_id>/copy")
def recipe_copy(recipe_id: int):
    source = db.session.get(KitchenRecipe, recipe_id)
    if not source:
        abort(404)
    base = f"{source.name} 複製"
    name = base
    index = 2
    while KitchenRecipe.query.filter_by(name=name).first():
        name = f"{base} {index}"
        index += 1
    copy = KitchenRecipe(name=name, category=source.category, serving_output_g=source.serving_output_g, note=source.note)
    db.session.add(copy)
    db.session.flush()
    for x in source.ingredients:
        db.session.add(KitchenRecipeIngredient(
            recipe_id=copy.id,
            ingredient_id=x.ingredient_id,
            grams_per_person=x.grams_per_person,
        ))
    db.session.commit()
    flash("已複製菜色，可直接修改食材或每人用量。", "success")
    return redirect(url_for("order_tool.recipe_detail", recipe_id=copy.id))


@order_bp.post("/recipes/<int:recipe_id>/ingredients")
def recipe_ingredient_add(recipe_id: int):
    recipe = db.session.get(KitchenRecipe, recipe_id)
    if not recipe:
        abort(404)
    ingredient_id = _int(request.form.get("ingredient_id"), default=0) or 0
    amount = _decimal(request.form.get("grams_per_person"))
    ingredient = db.session.get(KitchenIngredient, ingredient_id)
    if not ingredient or not ingredient.active or amount is None or amount <= 0:
        flash("食材或每人用量不正確。", "error")
        return redirect(url_for("order_tool.recipe_detail", recipe_id=recipe_id))
    existing = KitchenRecipeIngredient.query.filter_by(recipe_id=recipe_id, ingredient_id=ingredient_id).first()
    if existing:
        existing.grams_per_person = amount
        existing.quantity_status = "manual"
        existing.source_note = "人工確認"
        message = "配方用量已更新。"
    else:
        db.session.add(KitchenRecipeIngredient(
            recipe_id=recipe_id,
            ingredient_id=ingredient_id,
            grams_per_person=amount,
            quantity_status="manual",
            source_note="人工確認",
        ))
        message = "食材已加入配方。"
    db.session.commit()
    flash(message, "success")
    return redirect(url_for("order_tool.recipe_detail", recipe_id=recipe_id))


@order_bp.post("/recipe-ingredients/<int:row_id>/update")
def recipe_ingredient_update(row_id: int):
    row = db.session.get(KitchenRecipeIngredient, row_id)
    if not row:
        abort(404)
    amount = _decimal(request.form.get("grams_per_person"))
    if amount is None or amount <= 0:
        flash("每人用量必須大於 0。", "error")
        return redirect(url_for("order_tool.recipe_detail", recipe_id=row.recipe_id))
    row.grams_per_person = amount
    row.quantity_status = "manual"
    row.source_note = "人工確認"
    db.session.commit()
    flash("每人用量已更新。", "success")
    return redirect(url_for("order_tool.recipe_detail", recipe_id=row.recipe_id))


@order_bp.post("/recipe-ingredients/<int:row_id>/delete")
def recipe_ingredient_delete(row_id: int):
    row = db.session.get(KitchenRecipeIngredient, row_id)
    if not row:
        abort(404)
    recipe_id = row.recipe_id
    db.session.delete(row)
    db.session.commit()
    flash("食材已從配方移除。", "success")
    return redirect(url_for("order_tool.recipe_detail", recipe_id=recipe_id))


# ─────────────────────────────────────────────
# Menu plan CRUD
# ─────────────────────────────────────────────


@order_bp.route("/plans", methods=["GET", "POST"])
def plans():
    if request.method == "GET":
        selected = _date(request.args.get("start"), default=date.today()) or date.today()
        week_start = selected - timedelta(days=selected.weekday())
        return redirect(url_for("order_tool.summary", week=week_start.isoformat()))
    if request.method == "POST":
        service_date = _date(request.form.get("service_date"))
        meal_type = request.form.get("meal_type", "午餐").strip()
        name = request.form.get("name", "").strip()
        if not service_date or meal_type not in MEAL_TYPES or not name:
            flash("日期、餐別或菜單名稱不正確。", "error")
            return redirect(url_for("order_tool.plans"))
        plan = KitchenMenuPlan(
            service_date=service_date,
            meal_type=meal_type,
            name=name,
            note=request.form.get("note", "").strip() or None,
        )
        db.session.add(plan)
        try:
            db.session.commit()
            flash("菜單已建立。", "success")
            return redirect(url_for("order_tool.plan_detail", plan_id=plan.id))
        except IntegrityError:
            db.session.rollback()
            flash("同日、同餐別、同名稱的菜單已存在。", "error")
            return redirect(url_for("order_tool.plans"))

    start = _date(request.args.get("start"), default=date.today()) or date.today()
    end = start + timedelta(days=13)
    rows = KitchenMenuPlan.query.filter(KitchenMenuPlan.service_date.between(start, end)).order_by(
        KitchenMenuPlan.service_date, KitchenMenuPlan.meal_type, KitchenMenuPlan.name
    ).all()
    days = []
    for offset in range(7):
        day = start + timedelta(days=offset)
        days.append({"date": day, "weekday": WEEKDAY_LABELS[day.weekday()], "plans": [p for p in rows if p.service_date == day]})
    return render_template(
        "kitchen/plans.html",
        rows=rows,
        days=days,
        start=start,
        today=date.today().isoformat(),
        meal_types=MEAL_TYPES,
    )


@order_bp.get("/plans/<int:plan_id>")
def plan_detail(plan_id: int):
    plan = db.session.get(KitchenMenuPlan, plan_id)
    if not plan:
        abort(404)
    recipes = KitchenRecipe.query.filter_by(active=True).order_by(
        KitchenRecipe.category, KitchenRecipe.name
    ).all()
    return render_template(
        "kitchen/plan_detail.html",
        plan=plan,
        recipes=recipes,
        recipe_options=[
            {"id": recipe.id, "name": recipe.name, "category": recipe.category or "其他"}
            for recipe in recipes
        ],
        schools=KitchenSchool.query.filter_by(active=True).order_by(KitchenSchool.name).all(),
        total_people=sum(max(x.headcount, 0) for x in plan.assignments if x.service_status == "serving"),
        has_confirmed_orders=_active_confirmed_orders(plan.service_date),
        meal_types=MEAL_TYPES,
    )


@order_bp.post("/plans/<int:plan_id>/update")
def plan_update(plan_id: int):
    plan = db.session.get(KitchenMenuPlan, plan_id)
    if not plan:
        abort(404)
    if not _require_draft_plan(plan):
        return redirect(url_for("order_tool.plan_detail", plan_id=plan_id))
    service_date = _date(request.form.get("service_date"))
    meal_type = request.form.get("meal_type", "").strip()
    name = request.form.get("name", "").strip()
    if not service_date or meal_type not in MEAL_TYPES or not name:
        flash("日期、餐別或名稱不正確。", "error")
        return redirect(url_for("order_tool.plan_detail", plan_id=plan_id))
    if _plan_has_assignment_conflict(plan, service_date, meal_type):
        flash("這張菜單裡有學校已被安排在目標日期的同一餐別，無法變更日期/餐別。", "error")
        return redirect(url_for("order_tool.plan_detail", plan_id=plan_id))
    plan.service_date = service_date
    plan.meal_type = meal_type
    plan.name = name
    plan.note = request.form.get("note", "").strip() or None
    return _commit("菜單基本資料已更新。", "order_tool.plan_detail", plan_id=plan_id)


@order_bp.post("/plans/<int:plan_id>/items")
def plan_item_add(plan_id: int):
    plan = db.session.get(KitchenMenuPlan, plan_id)
    if not plan:
        abort(404)
    if not _require_draft_plan(plan):
        return redirect(url_for("order_tool.plan_detail", plan_id=plan_id))
    recipe_id = _int(request.form.get("recipe_id"), default=0) or 0
    recipe = db.session.get(KitchenRecipe, recipe_id)
    if not recipe or not recipe.active:
        flash("找不到可用菜色。", "error")
        return redirect(url_for("order_tool.plan_detail", plan_id=plan_id))
    if not KitchenMenuPlanItem.query.filter_by(plan_id=plan_id, recipe_id=recipe_id).first():
        max_order = max((x.sort_order for x in plan.items), default=-1)
        db.session.add(KitchenMenuPlanItem(plan_id=plan_id, recipe_id=recipe_id, sort_order=max_order + 1))
        db.session.commit()
        flash("菜色已加入。", "success")
    else:
        flash("這道菜已在菜單中。", "warning")
    return redirect(url_for("order_tool.plan_detail", plan_id=plan_id))


@order_bp.post("/plan-items/<int:row_id>/delete")
def plan_item_delete(row_id: int):
    row = db.session.get(KitchenMenuPlanItem, row_id)
    if not row:
        abort(404)
    plan_id = row.plan_id
    if not _require_draft_plan(row.plan):
        return redirect(url_for("order_tool.plan_detail", plan_id=plan_id))
    db.session.delete(row)
    db.session.commit()
    flash("菜色已移除。", "success")
    return redirect(url_for("order_tool.plan_detail", plan_id=plan_id))


@order_bp.post("/plans/<int:plan_id>/assignments")
def assignment_add(plan_id: int):
    plan = db.session.get(KitchenMenuPlan, plan_id)
    if not plan:
        abort(404)
    if not _require_draft_plan(plan):
        return redirect(url_for("order_tool.plan_detail", plan_id=plan_id))
    school_id = _int(request.form.get("school_id"), default=0) or 0
    headcount = _int(request.form.get("headcount"), default=None)
    school = db.session.get(KitchenSchool, school_id)
    if not school or not school.active or headcount is None or headcount < 0:
        flash("學校或人數不正確。", "error")
        return redirect(url_for("order_tool.plan_detail", plan_id=plan_id))
    row = KitchenMenuAssignment.query.filter_by(plan_id=plan_id, school_id=school_id).first()
    if row:
        row.headcount = headcount
        message = "學校人數已更新。"
    else:
        conflict = _school_meal_conflict(plan, school_id)
        if conflict:
            flash(
                f"{school.name} 已經在 {plan.service_date} 的 {plan.meal_type} 另一張菜單中，不能重複加入，否則採購會重複計算。",
                "error",
            )
            return redirect(url_for("order_tool.plan_detail", plan_id=plan_id))
        db.session.add(KitchenMenuAssignment(plan_id=plan_id, school_id=school_id, headcount=headcount))
        message = "學校已加入菜單。"
    db.session.commit()
    flash(message, "success")
    return redirect(url_for("order_tool.plan_detail", plan_id=plan_id))


@order_bp.post("/assignments/<int:row_id>/update")
def assignment_update(row_id: int):
    row = db.session.get(KitchenMenuAssignment, row_id)
    if not row:
        abort(404)
    if not _require_draft_plan(row.plan):
        return redirect(url_for("order_tool.plan_detail", plan_id=row.plan_id))
    headcount = _int(request.form.get("headcount"), default=None)
    if headcount is None or headcount < 0:
        flash("人數不可為負數。", "error")
        return redirect(url_for("order_tool.plan_detail", plan_id=row.plan_id))
    row.headcount = headcount
    db.session.commit()
    flash("人數已更新。", "success")
    return redirect(url_for("order_tool.plan_detail", plan_id=row.plan_id))


@order_bp.post("/assignments/<int:row_id>/delete")
def assignment_delete(row_id: int):
    row = db.session.get(KitchenMenuAssignment, row_id)
    if not row:
        abort(404)
    plan_id = row.plan_id
    if not _require_draft_plan(row.plan):
        return redirect(url_for("order_tool.plan_detail", plan_id=plan_id))
    db.session.delete(row)
    db.session.commit()
    flash("學校已從菜單移除。", "success")
    return redirect(url_for("order_tool.plan_detail", plan_id=plan_id))


@order_bp.post("/plans/<int:plan_id>/copy")
def plan_copy(plan_id: int):
    source = db.session.get(KitchenMenuPlan, plan_id)
    if not source:
        abort(404)
    target_date = _date(request.form.get("target_date"))
    if not target_date:
        flash("請選擇正確的複製日期。", "error")
        return redirect(url_for("order_tool.plan_detail", plan_id=plan_id))
    if KitchenMenuPlan.query.filter_by(service_date=target_date, meal_type=source.meal_type, name=source.name).first():
        flash("目標日期已有相同餐別與名稱的菜單。", "error")
        return redirect(url_for("order_tool.plan_detail", plan_id=plan_id))
    conflicts = []
    for x in source.assignments:
        probe = KitchenMenuPlan(service_date=target_date, meal_type=source.meal_type)
        probe.id = -1
        existing = (
            KitchenMenuAssignment.query.join(KitchenMenuPlan)
            .filter(
                KitchenMenuAssignment.school_id == x.school_id,
                KitchenMenuPlan.service_date == target_date,
                KitchenMenuPlan.meal_type == source.meal_type,
            )
            .first()
        )
        if existing:
            conflicts.append(x.school.name)
    if conflicts:
        flash("以下學校在目標日期同餐別已有菜單，請先處理：" + "、".join(conflicts), "error")
        return redirect(url_for("order_tool.plan_detail", plan_id=plan_id))
    copy = KitchenMenuPlan(
        service_date=target_date,
        meal_type=source.meal_type,
        name=source.name,
        note=source.note,
        status="draft",
    )
    db.session.add(copy)
    db.session.flush()
    for x in source.items:
        db.session.add(KitchenMenuPlanItem(plan_id=copy.id, recipe_id=x.recipe_id, sort_order=x.sort_order))
    for x in source.assignments:
        db.session.add(KitchenMenuAssignment(plan_id=copy.id, school_id=x.school_id, headcount=x.headcount))
    db.session.commit()
    flash("菜單已複製，請確認新日期的人數。", "success")
    return redirect(url_for("order_tool.plan_detail", plan_id=copy.id))


@order_bp.post("/plans/<int:plan_id>/confirm")
def plan_confirm(plan_id: int):
    plan = db.session.get(KitchenMenuPlan, plan_id)
    if not plan:
        abort(404)
    if not plan.items or sum(x.headcount for x in plan.assignments if x.service_status == "serving") <= 0:
        flash("至少要有菜色與大於 0 的供餐人數才能確認。", "error")
        return redirect(url_for("order_tool.plan_detail", plan_id=plan_id))
    for x in plan.assignments:
        if _school_meal_conflict(plan, x.school_id):
            flash(f"{x.school.name} 在同一天同餐別還出現在另一張菜單，請先修正再確認。", "error")
            return redirect(url_for("order_tool.plan_detail", plan_id=plan_id))
    plan.status = "confirmed"
    db.session.commit()
    flash("菜單已確認。", "success")
    return redirect(url_for("order_tool.plan_detail", plan_id=plan_id))


@order_bp.post("/plans/<int:plan_id>/reopen")
def plan_reopen(plan_id: int):
    plan = db.session.get(KitchenMenuPlan, plan_id)
    if not plan:
        abort(404)
    if _active_confirmed_orders(plan.service_date):
        flash("這一天已有已確認採購單，請先到採購頁重開相關採購單。", "error")
        return redirect(url_for("order_tool.plan_detail", plan_id=plan_id))
    plan.status = "draft"
    db.session.commit()
    flash("菜單已重開為草稿。", "success")
    return redirect(url_for("order_tool.plan_detail", plan_id=plan_id))


@order_bp.post("/plans/<int:plan_id>/delete")
def plan_delete(plan_id: int):
    plan = db.session.get(KitchenMenuPlan, plan_id)
    if not plan:
        abort(404)
    if plan.status != "draft" or _active_confirmed_orders(plan.service_date):
        flash("只能刪除沒有正式採購紀錄的草稿菜單。", "error")
        return redirect(url_for("order_tool.plan_detail", plan_id=plan_id))
    db.session.delete(plan)
    db.session.commit()
    flash("草稿菜單已刪除。", "success")
    return redirect(url_for("order_tool.summary", week=plan.service_date.isoformat()))


# ─────────────────────────────────────────────
# Purchase calculation / persisted orders
# ─────────────────────────────────────────────


def _supplier_identity(ingredient: KitchenIngredient):
    if ingredient.supplier_id and ingredient.supplier:
        return f"supplier:{ingredient.supplier_id}", ingredient.supplier_id, ingredient.supplier.name
    # 未指定供應商要每個食材各自一張草稿，否則不同食材會被硬塞進同一張「未指定」單。
    return f"unassigned:{ingredient.id}", None, "⚠ 未指定供應商"


def _requirements_for_date(service_date: date):
    plans_on_day = KitchenMenuPlan.query.filter_by(service_date=service_date).all()
    grouped: dict[str, dict[int, dict]] = defaultdict(dict)
    for plan in plans_on_day:
        serving_assignments = [
            assignment
            for assignment in plan.assignments
            if assignment.service_status == "serving" and assignment.headcount > 0
        ]
        people = sum(assignment.headcount for assignment in serving_assignments)
        if people <= 0:
            continue
        for menu_item in plan.items:
            for component in menu_item.recipe.ingredients:
                ing = component.ingredient
                if (component.grams_per_person or Decimal("0")) <= 0:
                    continue
                supplier_key, supplier_id, supplier_name = _supplier_identity(ing)
                base_amount = (component.grams_per_person or Decimal("0")) * people
                current = grouped[supplier_key].get(ing.id)
                if current is None:
                    current = {
                        "supplier_id": supplier_id,
                        "supplier_name": supplier_name,
                        "ingredient": ing,
                        "required_amount": Decimal("0"),
                        "total_people": 0,
                        "school_names": set(),
                    }
                    grouped[supplier_key][ing.id] = current
                current["required_amount"] += base_amount
                current["total_people"] += people
                current["school_names"].update(
                    assignment.school.name for assignment in serving_assignments
                )
    return grouped


def _summary_data(service_date: date):
    plans = KitchenMenuPlan.query.filter_by(service_date=service_date).order_by(
        KitchenMenuPlan.meal_type, KitchenMenuPlan.name
    ).all()
    school_rows = []
    pending_rows = []
    seen_pending = set()
    for plan in plans:
        dishes = "、".join(item.recipe.name for item in plan.items)
        for assignment in plan.assignments:
            if assignment.service_status != "serving":
                continue
            school_rows.append({
                "plan": plan,
                "school": assignment.school,
                "headcount": assignment.headcount,
                "dishes": dishes,
            })
        if not any(x.service_status == "serving" and x.headcount > 0 for x in plan.assignments):
            continue
        for item in plan.items:
            for component in item.recipe.ingredients:
                if (component.grams_per_person or Decimal("0")) > 0:
                    continue
                key = (item.recipe_id, component.ingredient_id)
                if key in seen_pending:
                    continue
                seen_pending.add(key)
                pending_rows.append({"recipe": item.recipe, "component": component})

    material_rows = []
    for ingredient_rows in _requirements_for_date(service_date).values():
        for data in ingredient_rows.values():
            ingredient = data["ingredient"]
            amount = data["required_amount"]
            divisor = ingredient.grams_per_purchase_unit or Decimal("0")
            material_rows.append({
                "supplier_name": data["supplier_name"],
                "ingredient": ingredient,
                "required_amount": amount,
                "required_qty": amount / divisor if divisor > 0 else Decimal("0"),
            })
    material_rows.sort(key=lambda row: (row["supplier_name"], row["ingredient"].name))
    return {
        "plans": plans,
        "school_rows": school_rows,
        "material_rows": material_rows,
        "pending_rows": pending_rows,
        "total_people": sum(max(row["headcount"], 0) for row in school_rows),
    }


@order_bp.get("/summary")
def summary():
    selected_date = _date(request.args.get("week") or request.args.get("date"), default=date.today()) or date.today()
    week_start = selected_date - timedelta(days=selected_date.weekday())
    week_end = week_start + timedelta(days=6)
    plans = (
        KitchenMenuPlan.query.filter(
            KitchenMenuPlan.service_date.between(week_start, week_end),
            db.or_(KitchenMenuPlan.name == "中央菜單", ~KitchenMenuPlan.assignments.any()),
        )
        .options(selectinload(KitchenMenuPlan.items).selectinload(KitchenMenuPlanItem.recipe))
        .order_by(KitchenMenuPlan.service_date, KitchenMenuPlan.meal_type, KitchenMenuPlan.name)
        .all()
    )
    days = []
    for offset, weekday in enumerate(WEEKDAY_LABELS):
        day_date = week_start + timedelta(days=offset)
        day_plans = [plan for plan in plans if plan.service_date == day_date]
        days.append({
            "date": day_date,
            "weekday": weekday,
            "plans": day_plans,
            "draft_plans": [plan for plan in day_plans if plan.status == "draft"],
        })
    recipes = KitchenRecipe.query.filter_by(active=True).order_by(
        KitchenRecipe.category, KitchenRecipe.name
    ).all()
    return render_template(
        "kitchen/summary.html",
        days=days,
        recipes=recipes,
        recipe_options=[
            {"id": recipe.id, "name": recipe.name, "category": recipe.category or "其他"}
            for recipe in recipes
        ],
        categories=CATEGORIES,
        week_start=week_start,
        week_end=week_end,
        previous_week=week_start - timedelta(days=7),
        next_week=week_start + timedelta(days=7),
        today=date.today(),
    )


def _vegetarian_menu_name(school: KitchenSchool) -> str:
    return f"{school.name}素食菜單"


def _school_assignment_for_day(
    school_id: int, service_date: date, variant: str = "regular"
) -> KitchenMenuAssignment | None:
    query = (
        KitchenMenuAssignment.query.join(KitchenMenuPlan)
        .filter(
            KitchenMenuAssignment.school_id == school_id,
            KitchenMenuPlan.service_date == service_date,
            KitchenMenuPlan.meal_type == "午餐",
        )
    )
    if variant == "vegetarian":
        query = query.filter(KitchenMenuPlan.name.like("%素食菜單"))
    else:
        query = query.filter(~KitchenMenuPlan.name.like("%素食菜單"))
    return query.first()


def _editable_school_plan(
    school: KitchenSchool, service_date: date, variant: str = "regular"
) -> KitchenMenuPlan:
    """取得學校專屬菜單；舊資料若把學校掛在中央菜單，先安全拆開。"""

    assignment = _school_assignment_for_day(school.id, service_date, variant)
    if assignment and assignment.plan.name != "中央菜單":
        return assignment.plan
    if assignment:
        db.session.delete(assignment)
        db.session.flush()

    menu_name = _vegetarian_menu_name(school) if variant == "vegetarian" else f"{school.name}菜單"
    plan = KitchenMenuPlan.query.filter_by(
        service_date=service_date,
        meal_type="午餐",
        name=menu_name,
    ).first()
    if plan is None:
        plan = KitchenMenuPlan(
            service_date=service_date,
            meal_type="午餐",
            name=menu_name,
            status="draft",
        )
        db.session.add(plan)
        db.session.flush()
    if not KitchenMenuAssignment.query.filter_by(plan_id=plan.id, school_id=school.id).first():
        db.session.add(KitchenMenuAssignment(
            plan_id=plan.id,
            school_id=school.id,
            headcount=max(
                school.default_vegetarian_headcount if variant == "vegetarian" else school.default_headcount,
                0,
            ),
        ))
        db.session.flush()
    return plan


def _school_week_completion(week_start: date) -> tuple[list[KitchenSchool], set[int], set[date]]:
    week_end = week_start + timedelta(days=6)
    schools = KitchenSchool.query.filter_by(active=True).order_by(KitchenSchool.name).all()
    central_plans = KitchenMenuPlan.query.filter(
        KitchenMenuPlan.service_date.between(week_start, week_end),
        KitchenMenuPlan.meal_type == "午餐",
        db.or_(KitchenMenuPlan.name == "中央菜單", ~KitchenMenuPlan.assignments.any()),
    ).all()
    required_dates = {plan.service_date for plan in central_plans if plan.items}
    assignments = (
        KitchenMenuAssignment.query.join(KitchenMenuPlan)
        .filter(
            KitchenMenuPlan.service_date.between(week_start, week_end),
            KitchenMenuPlan.meal_type == "午餐",
            KitchenMenuAssignment.school_id.in_([school.id for school in schools] or [-1]),
        )
        .options(selectinload(KitchenMenuAssignment.plan).selectinload(KitchenMenuPlan.items))
        .all()
    )
    assignments_by_school_date: dict[tuple[int, date], list[KitchenMenuAssignment]] = defaultdict(list)
    for assignment in assignments:
        assignments_by_school_date[(assignment.school_id, assignment.plan.service_date)].append(assignment)
    dates_by_school: dict[int, set[date]] = defaultdict(set)
    for school in schools:
        for service_date in required_dates:
            rows = assignments_by_school_date[(school.id, service_date)]
            if any(row.service_status == "no_service" for row in rows):
                dates_by_school[school.id].add(service_date)
                continue
            regular = next((row for row in rows if not row.plan.name.endswith("素食菜單")), None)
            vegetarian = next((row for row in rows if row.plan.name.endswith("素食菜單")), None)
            regular_ok = bool(regular and regular.plan.items and regular.headcount > 0)
            vegetarian_ok = (
                bool(vegetarian and (
                    vegetarian.headcount == 0
                    or (vegetarian.plan.items and vegetarian.headcount > 0)
                ))
                or (vegetarian is None and school.default_vegetarian_headcount <= 0)
            )
            if regular_ok and vegetarian_ok:
                dates_by_school[school.id].add(service_date)
    complete_ids = {
        school.id for school in schools
        if required_dates and required_dates.issubset(dates_by_school[school.id])
    }
    return schools, complete_ids, required_dates


def _missing_school_names_for_date(service_date: date) -> list[str]:
    schools = KitchenSchool.query.filter_by(active=True).order_by(KitchenSchool.name).all()
    assignments = (
        KitchenMenuAssignment.query.join(KitchenMenuPlan)
        .filter(
            KitchenMenuPlan.service_date == service_date,
            KitchenMenuPlan.meal_type == "午餐",
            KitchenMenuAssignment.school_id.in_([school.id for school in schools] or [-1]),
        )
        .options(selectinload(KitchenMenuAssignment.plan).selectinload(KitchenMenuPlan.items))
        .all()
    )
    rows_by_school: dict[int, list[KitchenMenuAssignment]] = defaultdict(list)
    for row in assignments:
        rows_by_school[row.school_id].append(row)
    completed_ids = set()
    for school in schools:
        rows = rows_by_school[school.id]
        if any(row.service_status == "no_service" for row in rows):
            completed_ids.add(school.id)
            continue
        regular = next((row for row in rows if not row.plan.name.endswith("素食菜單")), None)
        vegetarian = next((row for row in rows if row.plan.name.endswith("素食菜單")), None)
        if (
            regular and regular.plan.items and regular.headcount > 0
            and (
                (vegetarian and (
                    vegetarian.headcount == 0
                    or (vegetarian.plan.items and vegetarian.headcount > 0)
                ))
                or (vegetarian is None and school.default_vegetarian_headcount <= 0)
            )
        ):
            completed_ids.add(school.id)
    return [school.name for school in schools if school.id not in completed_ids]


@order_bp.get("/summary/schools")
def school_menus():
    selected_date = _date(request.args.get("week"), default=date.today()) or date.today()
    week_start = selected_date - timedelta(days=selected_date.weekday())
    week_end = week_start + timedelta(days=6)
    schools = KitchenSchool.query.filter_by(active=True).order_by(KitchenSchool.name).all()
    requested_school_id = _int(request.args.get("school_id"), default=0) or 0
    selected_school = next((school for school in schools if school.id == requested_school_id), None)
    if selected_school is None and schools:
        selected_school = schools[0]

    central_plans = (
        KitchenMenuPlan.query.filter(
            KitchenMenuPlan.service_date.between(week_start, week_end),
            KitchenMenuPlan.meal_type == "午餐",
            db.or_(KitchenMenuPlan.name == "中央菜單", ~KitchenMenuPlan.assignments.any()),
        )
        .options(selectinload(KitchenMenuPlan.items).selectinload(KitchenMenuPlanItem.recipe))
        .order_by(KitchenMenuPlan.service_date, KitchenMenuPlan.id)
        .all()
    )
    confirmed_order_dates = {
        row[0] for row in db.session.query(KitchenPurchaseOrder.service_date).filter(
            KitchenPurchaseOrder.service_date.between(week_start, week_end),
            KitchenPurchaseOrder.status == "confirmed",
        ).all()
    }
    school_assignments = []
    if selected_school:
        school_assignments = (
            KitchenMenuAssignment.query.join(KitchenMenuPlan)
            .filter(
                KitchenMenuAssignment.school_id == selected_school.id,
                KitchenMenuPlan.service_date.between(week_start, week_end),
                KitchenMenuPlan.meal_type == "午餐",
            )
            .options(
                selectinload(KitchenMenuAssignment.plan)
                .selectinload(KitchenMenuPlan.items)
                .selectinload(KitchenMenuPlanItem.recipe)
            )
            .all()
        )
    assignment_by_date_variant = {
        (row.plan.service_date, "vegetarian" if row.plan.name.endswith("素食菜單") else "regular"): row
        for row in school_assignments
    }
    days = []
    for offset, weekday in enumerate(WEEKDAY_LABELS):
        day_date = week_start + timedelta(days=offset)
        source_items = []
        seen_recipe_ids = set()
        for plan in central_plans:
            if plan.service_date != day_date:
                continue
            for item in plan.items:
                if item.recipe_id not in seen_recipe_ids:
                    source_items.append(item)
                    seen_recipe_ids.add(item.recipe_id)
        regular_assignment = assignment_by_date_variant.get((day_date, "regular"))
        vegetarian_assignment = assignment_by_date_variant.get((day_date, "vegetarian"))
        assignments_for_day = [row for row in (regular_assignment, vegetarian_assignment) if row]
        selected_ids = {
            "regular": {item.recipe_id for item in regular_assignment.plan.items} if regular_assignment else set(),
            "vegetarian": {item.recipe_id for item in vegetarian_assignment.plan.items} if vegetarian_assignment else set(),
        }
        for assignment in assignments_for_day:
            for item in assignment.plan.items:
                if item.recipe_id not in seen_recipe_ids:
                    source_items.append(item)
                    seen_recipe_ids.add(item.recipe_id)
        # 選擇學校時的「總表」固定依菜色分類顯示，
        # 不受匯入次序或學校專屬菜單的排列影響。
        source_items.sort(key=_menu_item_category_sort_key)
        no_service = any(row.service_status == "no_service" for row in assignments_for_day)
        days.append({
            "date": day_date,
            "weekday": weekday,
            "source_items": source_items,
            "selected_ids": selected_ids,
            "assignments": assignments_for_day,
            "headcount": regular_assignment.headcount if regular_assignment else (selected_school.default_headcount if selected_school else 0),
            "vegetarian_headcount": vegetarian_assignment.headcount if vegetarian_assignment else (selected_school.default_vegetarian_headcount if selected_school else 0),
            "service_status": "no_service" if no_service else "serving",
            "locked": bool(
                day_date in confirmed_order_dates
                or any(row.plan.status != "draft" for row in assignments_for_day)
            ),
        })
    _all_schools, complete_school_ids, required_dates = _school_week_completion(week_start)
    return render_template(
        "kitchen/school_menus.html",
        schools=schools,
        selected_school=selected_school,
        days=days,
        week_start=week_start,
        week_end=week_end,
        previous_week=week_start - timedelta(days=7),
        next_week=week_start + timedelta(days=7),
        export_date=date.today() if week_start <= date.today() <= week_end else week_start,
        complete_school_ids=complete_school_ids,
        required_dates=required_dates,
        missing_schools_by_date={
            (week_start + timedelta(days=offset)).isoformat():
                _missing_school_names_for_date(week_start + timedelta(days=offset))
            for offset in range(7)
        },
    )


def _copy_nonregistered_template_row(sheet, source_row: int, target_row: int):
    for column in range(1, 28):
        source = sheet.cell(source_row, column)
        target = sheet.cell(target_row, column)
        target._style = copy(source._style)
        target.font = copy(source.font)
        target.fill = copy(source.fill)
        target.border = copy(source.border)
        target.alignment = copy(source.alignment)
        target.number_format = source.number_format
        target.protection = copy(source.protection)
    sheet.row_dimensions[target_row].height = sheet.row_dimensions[source_row].height


def _put_nonregistered_dishes(sheet, row_number: int, plans: list[KitchenMenuPlan]):
    dishes_by_category = defaultdict(list)
    for plan in plans:
        for item in sorted(plan.items, key=_menu_item_category_sort_key):
            category = item.recipe.category if item.recipe.category in NONREGISTERED_MENU_COLUMNS else "其他"
            dishes_by_category[category].append(item.recipe.name)
    for category, columns in NONREGISTERED_MENU_COLUMNS.items():
        if category == "其他":
            continue
        names = list(dishes_by_category[category])
        if category == "點心":
            names.extend(dishes_by_category["其他"])
        for index, column in enumerate(columns):
            if index >= len(names):
                break
            if index == len(columns) - 1 and len(names) > len(columns):
                sheet.cell(row_number, column).value = "、".join(names[index:])
            else:
                sheet.cell(row_number, column).value = names[index]


@order_bp.get("/summary/schools/nonregistered-menu.xlsx")
def school_menus_export():
    service_date = _date(request.args.get("date"), default=date.today()) or date.today()
    assignments = (
        KitchenMenuAssignment.query.join(KitchenMenuPlan)
        .filter(
            KitchenMenuPlan.service_date == service_date,
            KitchenMenuPlan.meal_type == "午餐",
            KitchenMenuAssignment.service_status == "serving",
            KitchenMenuAssignment.headcount > 0,
        )
        .options(
            selectinload(KitchenMenuAssignment.school),
            selectinload(KitchenMenuAssignment.plan)
            .selectinload(KitchenMenuPlan.items)
            .selectinload(KitchenMenuPlanItem.recipe),
        )
        .all()
    )
    grouped_assignments = defaultdict(list)
    for assignment in assignments:
        if assignment.plan.items:
            grouped_assignments[assignment.school_id].append(assignment)
    school_rows = []
    for school_assignments in grouped_assignments.values():
        # 同校只佔一列；葷食在前，素食緊接在後。
        school_assignments.sort(
            key=lambda assignment: (
                assignment.plan.name.endswith("素食菜單"),
                assignment.plan.id,
            )
        )
        school_rows.append((school_assignments[0].school, school_assignments))
    school_rows.sort(key=lambda row: row[0].name.casefold())

    template_path = Path(current_app.root_path) / NONREGISTERED_MENU_TEMPLATE
    if not template_path.is_file():
        abort(500, description="找不到非登合菜名 Excel 範本。")
    workbook = load_workbook(template_path)
    sheet = workbook.active
    first_data_row = 2
    template_last_row = max(sheet.max_row, 9)
    required_last_row = max(template_last_row, first_data_row + len(school_rows) - 1)
    for row_number in range(template_last_row + 1, required_last_row + 1):
        _copy_nonregistered_template_row(sheet, first_data_row, row_number)
    for row_number in range(first_data_row, required_last_row + 1):
        for column in range(1, 28):
            sheet.cell(row_number, column).value = None

    for row_number, (school, school_assignments) in enumerate(school_rows, start=first_data_row):
        sheet.cell(row_number, 1).value = school.name
        sheet.cell(row_number, 2).value = service_date
        sheet.cell(row_number, 3).value = school_assignments[0].plan.meal_type or "午餐"
        for column, value in enumerate(NONREGISTERED_MENU_DEFAULTS, start=4):
            sheet.cell(row_number, column).value = value
        _put_nonregistered_dishes(
            sheet,
            row_number,
            [assignment.plan for assignment in school_assignments],
        )

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name=f"非登合菜名-{service_date.isoformat()}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@order_bp.post("/summary/schools/save-day")
def school_menu_save_day():
    school_id = _int(request.form.get("school_id"), default=0) or 0
    school = db.session.get(KitchenSchool, school_id)
    service_date = _date(request.form.get("service_date"))
    headcount = _int(request.form.get("headcount"), default=None)
    service_status = request.form.get("service_status", "serving").strip()
    if not school or not school.active or not service_date:
        return {"message": "學校或日期不正確。"}, 400
    if headcount is None or headcount < 0:
        return {"message": "供餐人數不可為負數。"}, 400
    if service_status not in SCHOOL_SERVICE_STATUSES:
        return {"message": "供餐狀態不正確。"}, 400
    if _active_confirmed_orders(service_date):
        return {"message": "這一天已有確認採購單，請先重開採購草稿。"}, 409

    # 新介面會一次送出葷、素兩組資料，切換畫面時不會覆蓋另一組勾選。
    if "vegetarian_headcount" in request.form:
        regular_headcount = _int(request.form.get("headcount"), default=None)
        vegetarian_headcount = _int(request.form.get("vegetarian_headcount"), default=None)
        if regular_headcount is None or regular_headcount < 0 or vegetarian_headcount is None or vegetarian_headcount < 0:
            return {"message": "葷、素供餐人數不可為負數。"}, 400
        existing = [
            row for row in (
                _school_assignment_for_day(school.id, service_date, "regular"),
                _school_assignment_for_day(school.id, service_date, "vegetarian"),
            ) if row
        ]
        if any(row.plan.status != "draft" for row in existing):
            return {"message": "這一天已確認，無法修改。"}, 409
        if service_status == "no_service":
            plan = _editable_school_plan(school, service_date, "regular")
            assignment = KitchenMenuAssignment.query.filter_by(plan_id=plan.id, school_id=school.id).one()
            assignment.headcount = regular_headcount
            assignment.service_status = "no_service"
            for row in existing:
                row.service_status = "no_service"
            db.session.commit()
            return "", 204

        def sync_variant(variant: str, people: int, field_name: str):
            recipe_ids = {
                recipe_id for raw in request.form.getlist(field_name)
                if (recipe_id := _int(raw, default=0)) and db.session.get(KitchenRecipe, recipe_id)
            }
            assignment = _school_assignment_for_day(school.id, service_date, variant)
            if variant == "regular" and people <= 0 and not recipe_ids:
                if assignment:
                    db.session.delete(assignment.plan)
                return
            plan = _editable_school_plan(school, service_date, variant)
            for item in list(plan.items):
                db.session.delete(item)
            db.session.flush()
            recipes = KitchenRecipe.query.filter(KitchenRecipe.id.in_(recipe_ids or {-1})).order_by(
                KitchenRecipe.category, KitchenRecipe.name
            ).all()
            for sort_order, recipe in enumerate(recipes):
                db.session.add(KitchenMenuPlanItem(
                    plan_id=plan.id, recipe_id=recipe.id, sort_order=sort_order
                ))
            row = KitchenMenuAssignment.query.filter_by(plan_id=plan.id, school_id=school.id).one()
            row.headcount = people
            row.service_status = "serving"

        sync_variant("regular", regular_headcount, "regular_recipe_ids")
        sync_variant("vegetarian", vegetarian_headcount, "vegetarian_recipe_ids")
        db.session.commit()
        return "", 204

    assignment = _school_assignment_for_day(school.id, service_date)
    if assignment and assignment.plan.status != "draft":
        return {"message": "這一天已確認，無法修改。"}, 409

    selected_recipe_ids = {
        recipe_id for raw in request.form.getlist("recipe_ids")
        if (recipe_id := _int(raw, default=0)) and db.session.get(KitchenRecipe, recipe_id)
    }
    if service_status == "no_service":
        plan = _editable_school_plan(school, service_date)
        if plan.status != "draft":
            db.session.rollback()
            return {"message": "這一天已確認，無法修改。"}, 409
        assignment = KitchenMenuAssignment.query.filter_by(plan_id=plan.id, school_id=school.id).one()
        assignment.headcount = headcount
        assignment.service_status = "no_service"
        db.session.commit()
        return "", 204

    if not selected_recipe_ids:
        if assignment:
            plan = assignment.plan
            if plan.name == "中央菜單":
                db.session.delete(assignment)
            else:
                db.session.delete(plan)
        db.session.commit()
        return "", 204

    plan = _editable_school_plan(school, service_date)
    if plan.status != "draft":
        db.session.rollback()
        return {"message": "這一天已確認，無法修改。"}, 409
    for item in list(plan.items):
        db.session.delete(item)
    db.session.flush()
    recipes = KitchenRecipe.query.filter(KitchenRecipe.id.in_(selected_recipe_ids)).order_by(
        KitchenRecipe.category, KitchenRecipe.name
    ).all()
    for sort_order, recipe in enumerate(recipes):
        db.session.add(KitchenMenuPlanItem(
            plan_id=plan.id,
            recipe_id=recipe.id,
            sort_order=sort_order,
        ))
    assignment = KitchenMenuAssignment.query.filter_by(plan_id=plan.id, school_id=school.id).one()
    assignment.headcount = headcount
    assignment.service_status = "serving"
    db.session.commit()
    return "", 204


@order_bp.post("/summary/schools/save")
def school_menus_save():
    school_id = _int(request.form.get("school_id"), default=0) or 0
    school = db.session.get(KitchenSchool, school_id)
    selected_date = _date(request.form.get("week"), default=date.today()) or date.today()
    week_start = selected_date - timedelta(days=selected_date.weekday())
    redirect_to = url_for("order_tool.school_menus", week=week_start.isoformat(), school_id=school_id)
    if not school or not school.active:
        flash("請先選擇有效的學校。", "error")
        return redirect(redirect_to)

    saved_days = 0
    locked_days = 0
    for offset in range(7):
        service_date = week_start + timedelta(days=offset)
        field_suffix = service_date.isoformat()
        assignment = _school_assignment_for_day(school.id, service_date)
        if _active_confirmed_orders(service_date) or (assignment and assignment.plan.status != "draft"):
            locked_days += 1
            continue
        headcount = _int(request.form.get(f"headcount_{field_suffix}"), default=None)
        if headcount is None or headcount < 0:
            db.session.rollback()
            flash(f"{service_date.strftime('%m/%d')} 的人數不可為負數。", "error")
            return redirect(redirect_to)
        selected_recipe_ids = {
            recipe_id for raw in request.form.getlist(f"recipes_{field_suffix}")
            if (recipe_id := _int(raw, default=0)) and db.session.get(KitchenRecipe, recipe_id)
        }
        if not selected_recipe_ids:
            if assignment:
                plan = assignment.plan
                if plan.name == "中央菜單":
                    db.session.delete(assignment)
                else:
                    db.session.delete(plan)
                saved_days += 1
            continue

        plan = _editable_school_plan(school, service_date)
        if plan.status != "draft":
            locked_days += 1
            continue
        for item in list(plan.items):
            db.session.delete(item)
        db.session.flush()
        recipes = KitchenRecipe.query.filter(KitchenRecipe.id.in_(selected_recipe_ids)).order_by(
            KitchenRecipe.category, KitchenRecipe.name
        ).all()
        for sort_order, recipe in enumerate(recipes):
            db.session.add(KitchenMenuPlanItem(
                plan_id=plan.id,
                recipe_id=recipe.id,
                sort_order=sort_order,
            ))
        assignment = KitchenMenuAssignment.query.filter_by(plan_id=plan.id, school_id=school.id).one()
        assignment.headcount = headcount
        assignment.service_status = "serving"
        saved_days += 1
    db.session.commit()
    message = f"已儲存 {school.name} 本週菜單與人數（{saved_days} 天）。"
    if locked_days:
        message += f" {locked_days} 天已確認，未修改。"
    flash(message, "success")
    return redirect(redirect_to)


@order_bp.post("/summary/dishes")
def summary_dish_add():
    service_date = _date(request.form.get("service_date"))
    recipe_id = _int(request.form.get("recipe_id"), default=0) or 0
    dish_name = request.form.get("dish_name", "").strip()[:120]
    category = request.form.get("category", "其他").strip()
    plan_id = _int(request.form.get("plan_id"), default=0) or 0
    week_start = _date(request.form.get("week"), default=service_date or date.today()) or date.today()
    redirect_to = url_for("order_tool.summary", week=week_start.isoformat())
    if not service_date or (not recipe_id and not dish_name):
        flash("請搜尋或輸入一道菜色。", "error")
        return redirect(redirect_to)

    recipe = db.session.get(KitchenRecipe, recipe_id) if recipe_id else None
    if recipe is not None and not recipe.active:
        recipe = None
    if recipe is None and dish_name:
        recipe = KitchenRecipe.query.filter(
            db.func.lower(KitchenRecipe.name) == dish_name.lower()
        ).first()
        if recipe is not None and not recipe.active:
            flash(f"「{recipe.name}」目前已停用，請先到菜色配方重新啟用。", "error")
            return redirect(redirect_to)
    if recipe is None:
        if category not in CATEGORIES:
            category = "其他"
        recipe = KitchenRecipe(name=dish_name, category=category, active=True)
        db.session.add(recipe)
        try:
            db.session.flush()
        except IntegrityError:
            db.session.rollback()
            recipe = KitchenRecipe.query.filter(
                db.func.lower(KitchenRecipe.name) == dish_name.lower()
            ).first()
            if recipe is None or not recipe.active:
                flash("菜色建立失敗，請重新整理後再試一次。", "error")
                return redirect(redirect_to)

    plan = db.session.get(KitchenMenuPlan, plan_id) if plan_id else None
    if plan and plan.service_date != service_date:
        flash("菜單日期不相符，請重新選擇。", "error")
        return redirect(redirect_to)
    if plan is None:
        plan = KitchenMenuPlan.query.filter_by(
            service_date=service_date, meal_type="午餐", name="中央菜單", status="draft"
        ).order_by(KitchenMenuPlan.id).first()
    if plan is None:
        existing_central = KitchenMenuPlan.query.filter_by(
            service_date=service_date, meal_type="午餐", name="中央菜單"
        ).first()
        if existing_central:
            flash("當天的中央菜單已確認，不能再新增菜色。", "error")
            return redirect(redirect_to)
        plan = KitchenMenuPlan(
            service_date=service_date,
            meal_type="午餐",
            name="中央菜單",
            status="draft",
        )
        db.session.add(plan)
        db.session.flush()
    if plan.status != "draft":
        flash("已確認的菜單不能再新增菜色。", "error")
        return redirect(redirect_to)
    if KitchenMenuPlanItem.query.filter_by(plan_id=plan.id, recipe_id=recipe.id).first():
        flash(f"「{recipe.name}」已經在這張菜單中。", "warning")
        return redirect(redirect_to)

    max_order = max((item.sort_order for item in plan.items), default=-1)
    db.session.add(KitchenMenuPlanItem(plan_id=plan.id, recipe_id=recipe.id, sort_order=max_order + 1))
    db.session.commit()
    flash(f"已把「{recipe.name}」加入 {service_date.strftime('%m/%d')} 菜單。", "success")
    return redirect(redirect_to)


@order_bp.post("/summary/dishes/<int:row_id>/delete")
def summary_dish_delete(row_id: int):
    row = db.session.get(KitchenMenuPlanItem, row_id)
    if not row:
        abort(404)
    week_start = _date(request.form.get("week"), default=row.plan.service_date) or row.plan.service_date
    redirect_to = url_for("order_tool.summary", week=week_start.isoformat())
    if not _require_draft_plan(row.plan):
        return redirect(redirect_to)
    recipe_name = row.recipe.name
    db.session.delete(row)
    db.session.commit()
    flash(f"已從總表移除「{recipe_name}」。", "success")
    return redirect(redirect_to)


def _menu_text(value) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def _menu_name_key(value: str) -> str:
    return re.sub(r"[\s　]+", "", value).lower()


def _menu_category(header: str) -> str | None:
    raw = _menu_text(header)
    if "份" in raw or any(marker in raw for marker in ("全穀", "豆魚", "油脂", "水果", "乳品", "熱量", "三章")):
        return None
    normalized = raw.replace("類", "")
    for marker, category in MENU_HEADER_CATEGORIES.items():
        if marker in normalized:
            return category
    return None


def _infer_menu_year(sheet, filename: str) -> int:
    texts = [filename]
    for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 6), values_only=True):
        texts.extend(_menu_text(value) for value in row if value is not None)
    joined = " ".join(texts)
    match = re.search(r"(?<!\d)(20\d{2})\s*年", joined)
    if match:
        return int(match.group(1))
    match = re.search(r"(?<!\d)(1\d{2})\s*年", joined)
    if match:
        return int(match.group(1)) + 1911
    for row in sheet.iter_rows(values_only=True):
        for value in row:
            if isinstance(value, (date, datetime)):
                return value.year
    return date.today().year


def _infer_menu_month(sheet, filename: str) -> int | None:
    texts = [filename]
    for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 6), values_only=True):
        texts.extend(_menu_text(value) for value in row if value is not None)
    match = re.search(r"(?<!\d)(1[0-2]|[1-9])\s*月", " ".join(texts))
    return int(match.group(1)) if match else None


def _parse_menu_date(value, inferred_year: int, inferred_month: int | None) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)) and inferred_month and 1 <= int(value) <= 31:
        try:
            return date(inferred_year, inferred_month, int(value))
        except ValueError:
            return None
    text = _menu_text(value)
    if not text:
        return None
    numbers = [int(part) for part in re.findall(r"\d+", text)]
    try:
        if len(numbers) >= 3:
            year, month, day = numbers[:3]
            if 100 <= year < 1911:
                year += 1911
            return date(year, month, day)
        if len(numbers) == 2:
            return date(inferred_year, numbers[0], numbers[1])
        if len(numbers) == 1 and inferred_month:
            return date(inferred_year, inferred_month, numbers[0])
    except ValueError:
        return None
    return None


def _find_menu_header(sheet):
    best = None
    for row_number in range(1, min(sheet.max_row, 30) + 1):
        headers = [_menu_text(sheet.cell(row_number, column).value) for column in range(1, sheet.max_column + 1)]
        date_columns = [index + 1 for index, value in enumerate(headers) if "日期" in value]
        weekday_columns = [index + 1 for index, value in enumerate(headers) if "星期" in value or "週次" in value]
        dish_columns = [index + 1 for index, value in enumerate(headers) if _menu_category(value)]
        score = (4 if date_columns else 0) + (2 if weekday_columns else 0) + len(dish_columns)
        if date_columns and dish_columns and (best is None or score > best[0]):
            best = (score, row_number, date_columns[0], weekday_columns[0] if weekday_columns else None, headers, dish_columns)
    return best


def parse_menu_workbook(raw: bytes, filename: str, sheet_kind: str = "regular") -> dict:
    """從可見工作表自動定位日期與菜名欄，回傳依日期去重的菜色。"""

    if not raw:
        raise ValueError("檔案是空的。")
    if len(raw) > 10 * 1024 * 1024:
        raise ValueError("檔案超過 10MB，請拆成較小的菜單再匯入。")
    if not filename.lower().endswith((".xlsx", ".xlsm")):
        raise ValueError("目前請上傳 .xlsx 或 .xlsm 菜單。")
    try:
        # 菜單匯入只需要儲存格。唯讀模式不載入圖片／文字框，也能避開
        # 部分外部製表軟體產生、openpyxl 無法解析的 DrawingML 字型屬性。
        workbook = load_workbook(BytesIO(raw), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError("Excel 無法讀取，請確認檔案沒有損壞或加密。") from exc

    sheet_kind = sheet_kind if sheet_kind in {"regular", "vegetarian", "all"} else "regular"
    visible_worksheets = [sheet for sheet in workbook.worksheets if sheet.sheet_state == "visible"]
    if sheet_kind == "all":
        selected_worksheets = visible_worksheets
    elif sheet_kind == "vegetarian":
        selected_worksheets = [sheet for sheet in visible_worksheets if "素食" in _menu_text(sheet.title)]
    else:
        selected_worksheets = [sheet for sheet in visible_worksheets if "素食" not in _menu_text(sheet.title)]
    if visible_worksheets and not selected_worksheets:
        label = "素食" if sheet_kind == "vegetarian" else "一般"
        raise ValueError(f"找不到{label}菜單工作表，請改選其他工作表類型。")

    by_date: dict[date, dict[str, dict]] = defaultdict(dict)
    parsed_sheets = 0
    duplicate_count = 0
    raw_dish_count = 0
    for sheet in selected_worksheets:
        header = _find_menu_header(sheet)
        if not header:
            continue
        parsed_sheets += 1
        _, header_row, date_column, _weekday_column, headers, dish_columns = header
        first_dish = min(dish_columns)
        last_dish = max(dish_columns)
        inferred_year = _infer_menu_year(sheet, filename)
        inferred_month = _infer_menu_month(sheet, filename)
        column_categories = {}
        current_category = "其他"
        for column in range(first_dish, last_dish + 1):
            detected = _menu_category(headers[column - 1])
            if detected:
                current_category = detected
            column_categories[column] = current_category

        for row_number in range(header_row + 1, sheet.max_row + 1):
            service_date = _parse_menu_date(
                sheet.cell(row_number, date_column).value,
                inferred_year,
                inferred_month,
            )
            if not service_date:
                continue
            row_dishes = []
            for column in range(first_dish, last_dish + 1):
                name = _menu_text(sheet.cell(row_number, column).value)
                if not name or name.startswith("="):
                    continue
                row_dishes.append({"name": name[:120], "category": column_categories[column]})
            if len(row_dishes) == 1 and re.search(r"節|放假|停餐|補假|快樂", row_dishes[0]["name"]):
                continue
            for dish in row_dishes:
                raw_dish_count += 1
                key = _menu_name_key(dish["name"])
                if key in by_date[service_date]:
                    duplicate_count += 1
                    continue
                by_date[service_date][key] = dish

    if not by_date:
        if not visible_worksheets:
            raise ValueError("Excel 沒有可見的工作表。")
        if not parsed_sheets:
            raise ValueError("找不到「日期」與菜名欄位。請提供這種格式作為新模板。")
        raise ValueError("有找到欄位，但沒有辨識到可匯入的日期與菜名。")
    return {
        "days": [
            {"date": service_date, "dishes": list(dishes.values())}
            for service_date, dishes in sorted(by_date.items())
        ],
        "parsed_sheets": parsed_sheets,
        "raw_dish_count": raw_dish_count,
        "duplicate_count": duplicate_count,
        "sheet_names": [sheet.title for sheet in selected_worksheets],
    }


@order_bp.post("/summary/import")
def summary_import():
    upload = request.files.get("menu_file")
    if not upload or not upload.filename:
        flash("請先選擇要匯入的 Excel 菜單。", "error")
        return redirect(url_for("order_tool.summary"))
    try:
        parsed = parse_menu_workbook(
            upload.read(),
            upload.filename,
            request.form.get("menu_sheet_kind", "regular"),
        )
    except ValueError as exc:
        flash(str(exc), "error")
        return redirect(url_for("order_tool.summary"))

    recipes_by_name = {
        _menu_name_key(recipe.name): recipe
        for recipe in KitchenRecipe.query.order_by(KitchenRecipe.id).all()
    }
    added_items = 0
    existing_items = parsed["duplicate_count"]
    created_recipes = 0
    locked_days = 0
    imported_days = 0
    for day in parsed["days"]:
        service_date = day["date"]
        plan = KitchenMenuPlan.query.filter_by(
            service_date=service_date, meal_type="午餐", name="中央菜單", status="draft"
        ).order_by(KitchenMenuPlan.id).first()
        if plan is None:
            locked_plan = KitchenMenuPlan.query.filter_by(
                service_date=service_date, meal_type="午餐", name="中央菜單"
            ).first()
            if locked_plan:
                locked_days += 1
                continue
            plan = KitchenMenuPlan(
                service_date=service_date,
                meal_type="午餐",
                name="中央菜單",
                status="draft",
            )
            db.session.add(plan)
            db.session.flush()
        existing_recipe_ids = {item.recipe_id for item in plan.items}
        day_added = 0
        for dish in day["dishes"]:
            key = _menu_name_key(dish["name"])
            recipe = recipes_by_name.get(key)
            if recipe is None:
                recipe = KitchenRecipe(
                    name=dish["name"],
                    category=dish["category"] if dish["category"] in CATEGORIES else "其他",
                    active=True,
                )
                db.session.add(recipe)
                db.session.flush()
                recipes_by_name[key] = recipe
                created_recipes += 1
            elif not recipe.active:
                recipe.active = True
            if recipe.id in existing_recipe_ids:
                existing_items += 1
                continue
            db.session.add(KitchenMenuPlanItem(
                plan_id=plan.id,
                recipe_id=recipe.id,
                sort_order=len(existing_recipe_ids),
            ))
            existing_recipe_ids.add(recipe.id)
            added_items += 1
            day_added += 1
        if day_added:
            imported_days += 1
    db.session.commit()

    first_day = parsed["days"][0]["date"]
    sheet_names = "、".join(parsed["sheet_names"])
    message = f"匯入完成（{sheet_names}）：{imported_days} 天新增 {added_items} 道菜，略過 {existing_items} 筆重複；建立 {created_recipes} 個新菜色。"
    if locked_days:
        message += f" 另有 {locked_days} 天已確認，未修改。"
    flash(message, "success")
    week_start = first_day - timedelta(days=first_day.weekday())
    return redirect(url_for("order_tool.summary", week=week_start.isoformat()))


@order_bp.get("/summary.xlsx")
def summary_export():
    service_date = _date(request.args.get("date"), default=date.today()) or date.today()
    data = _summary_data(service_date)
    workbook = Workbook()
    materials = workbook.active
    materials.title = "材料總表"
    materials.append(["供餐日期", "餐數（用餐人）", "廠商", "食材", "基本需求量", "基本單位", "換算需求", "採購單位"])
    for row in data["material_rows"]:
        ingredient = row["ingredient"]
        materials.append([
            service_date.isoformat(),
            data["total_people"],
            row["supplier_name"],
            ingredient.name,
            float(row["required_amount"]),
            ingredient.base_unit,
            float(row["required_qty"]),
            ingredient.purchase_unit,
        ])
    schools = workbook.create_sheet("學校菜單")
    schools.append(["日期", "餐別", "學校", "人數", "菜單", "狀態"])
    for row in data["school_rows"]:
        schools.append([
            service_date.isoformat(),
            row["plan"].meal_type,
            row["school"].name,
            row["headcount"],
            row["dishes"],
            _status_label(row["plan"].status),
        ])
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name=f"材料總表-{service_date.isoformat()}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


CATALOG_IMPORT_COLUMNS = (
    "菜色名稱", "分類", "材料名稱", "每人用量", "基本單位",
    "採購單位", "1採購單位換算", "單價", "廠商", "備註",
)


def _catalog_serializer():
    return URLSafeSerializer(current_app.secret_key, salt="kitchen-catalog-import-v1")


def _catalog_rows_from_xlsx(raw: bytes):
    if not raw:
        return [], ["檔案是空的。"]
    if len(raw) > 5 * 1024 * 1024:
        return [], ["檔案超過 5MB，請拆成多份匯入。"]
    try:
        workbook = load_workbook(BytesIO(raw), read_only=True, data_only=True)
    except Exception:
        return [], ["無法讀取 Excel，請使用系統下載的 .xlsx 範本。"]
    sheet = workbook["匯入資料"] if "匯入資料" in workbook.sheetnames else workbook.active
    values = list(sheet.iter_rows(values_only=True))
    if not values:
        return [], ["Excel 沒有資料。"]
    headers = {str(value).strip(): index for index, value in enumerate(values[0]) if value is not None}
    missing = [column for column in ("菜色名稱", "材料名稱") if column not in headers]
    if missing:
        return [], [f"缺少欄位：{'、'.join(missing)}。請重新下載範本。"]

    rows = []
    errors = []
    for row_number, values_row in enumerate(values[1:5001], start=2):
        def value(column):
            index = headers.get(column)
            return values_row[index] if index is not None and index < len(values_row) else None

        recipe_name = str(value("菜色名稱") or "").strip()
        ingredient_name = str(value("材料名稱") or "").strip()
        if not recipe_name and not ingredient_name:
            continue
        if not recipe_name:
            errors.append(f"第 {row_number} 列缺少菜色名稱。")
            continue
        if not ingredient_name:
            errors.append(f"第 {row_number} 列缺少材料名稱。")
            continue
        category = str(value("分類") or "主菜").strip()
        if category not in CATEGORIES:
            errors.append(f"第 {row_number} 列分類「{category}」不正確。")
            continue
        base_unit = str(value("基本單位") or "g").strip()
        purchase_unit = str(value("採購單位") or ("kg" if base_unit == "g" else "個")).strip()
        if base_unit not in BASE_UNITS or purchase_unit not in PURCHASE_UNITS:
            errors.append(f"第 {row_number} 列單位不正確。")
            continue
        amount = _decimal(value("每人用量"), default=Decimal("0")) or Decimal("0")
        conversion = _decimal(value("1採購單位換算"), default=Decimal("1000") if base_unit == "g" and purchase_unit == "kg" else Decimal("1"))
        unit_price = _decimal(value("單價"), default=Decimal("0")) or Decimal("0")
        if amount < 0 or conversion is None or conversion <= 0 or unit_price < 0:
            errors.append(f"第 {row_number} 列數量、換算或單價不正確。")
            continue
        rows.append({
            "recipe_name": recipe_name[:120],
            "category": category,
            "ingredient_name": ingredient_name[:100],
            "amount": str(amount),
            "base_unit": base_unit,
            "purchase_unit": purchase_unit,
            "conversion": str(conversion),
            "unit_price": str(unit_price),
            "supplier_name": str(value("廠商") or "").strip()[:100],
            "note": str(value("備註") or "").strip()[:255],
        })
    return rows, errors


@order_bp.get("/catalog-template.xlsx")
def catalog_template():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "匯入資料"
    sheet.append(CATALOG_IMPORT_COLUMNS)
    sheet.freeze_panes = "A2"
    widths = (22, 12, 22, 14, 12, 12, 18, 12, 22, 28)
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width
    example = workbook.create_sheet("填寫範例")
    example.append(CATALOG_IMPORT_COLUMNS)
    example.append(["茄汁豬柳", "主菜", "肉絲", 45, "g", "kg", 1000, 0, "肉品廠商", "同一道菜有多種材料就分多列"])
    example.append(["茄汁豬柳", "主菜", "洋蔥", 12, "g", "kg", 1000, 0, "蔬菜廠商", "菜色名稱重複列沒關係，系統會合併材料"])
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name="菜色材料總表匯入範本.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@order_bp.route("/catalog-import", methods=["GET", "POST"])
def catalog_import():
    if request.method == "GET":
        return render_template("kitchen/catalog_import.html")
    upload = request.files.get("file")
    if not upload or not upload.filename.lower().endswith(".xlsx"):
        flash("請選擇 .xlsx Excel 檔。", "error")
        return redirect(url_for("order_tool.catalog_import"))
    rows, errors = _catalog_rows_from_xlsx(upload.read())
    grouped = defaultdict(list)
    for row in rows:
        grouped[row["recipe_name"]].append(row)
    existing_names = {
        name for (name,) in db.session.query(KitchenRecipe.name).filter(KitchenRecipe.name.in_(list(grouped))).all()
    } if grouped else set()
    new_names = [name for name in grouped if name not in existing_names]
    new_rows = [row for name in new_names for row in grouped[name]]
    token = _catalog_serializer().dumps(new_rows) if new_rows and not errors else None
    return render_template(
        "kitchen/catalog_import.html",
        preview=True,
        errors=errors,
        new_names=new_names,
        duplicate_names=sorted(existing_names),
        new_material_count=len({row["ingredient_name"] for row in new_rows}),
        token=token,
    )


@order_bp.post("/catalog-import/apply")
def catalog_import_apply():
    try:
        rows = _catalog_serializer().loads(request.form.get("token", ""))
    except BadSignature:
        flash("匯入預覽已失效，請重新上傳 Excel。", "error")
        return redirect(url_for("order_tool.catalog_import"))
    grouped = defaultdict(list)
    for row in rows:
        grouped[row["recipe_name"]].append(row)
    created_recipes = created_components = skipped = 0
    for recipe_name, recipe_rows in grouped.items():
        if KitchenRecipe.query.filter_by(name=recipe_name).first():
            skipped += 1
            continue
        first = recipe_rows[0]
        recipe = KitchenRecipe(name=recipe_name, category=first["category"], note=first["note"] or "由總表 Excel 匯入")
        db.session.add(recipe)
        db.session.flush()
        created_recipes += 1
        seen_ingredients = set()
        for item in recipe_rows:
            if item["ingredient_name"] in seen_ingredients:
                continue
            seen_ingredients.add(item["ingredient_name"])
            supplier = None
            if item["supplier_name"]:
                supplier = KitchenSupplier.query.filter_by(name=item["supplier_name"]).one_or_none()
                if supplier is None:
                    supplier = KitchenSupplier(name=item["supplier_name"], note="由菜色總表匯入")
                    db.session.add(supplier)
                    db.session.flush()
            ingredient = KitchenIngredient.query.filter_by(name=item["ingredient_name"]).one_or_none()
            if ingredient is None:
                ingredient = KitchenIngredient(
                    name=item["ingredient_name"], supplier_id=supplier.id if supplier else None,
                    base_unit=item["base_unit"], purchase_unit=item["purchase_unit"],
                    grams_per_purchase_unit=Decimal(item["conversion"]), unit_price=Decimal(item["unit_price"]),
                    order_increment=Decimal("0.001") if item["purchase_unit"] == "kg" else Decimal("1"),
                    note="由菜色總表匯入",
                )
                db.session.add(ingredient)
                db.session.flush()
            amount = Decimal(item["amount"])
            db.session.add(KitchenRecipeIngredient(
                recipe_id=recipe.id,
                ingredient_id=ingredient.id,
                grams_per_person=amount,
                quantity_status="manual" if amount > 0 else "pending",
                source_note="總表 Excel 匯入" if amount > 0 else "總表 Excel；克數待確認",
            ))
            created_components += 1
    db.session.commit()
    flash(f"已新增 {created_recipes} 道菜、{created_components} 筆材料；另跳過 {skipped} 道重複菜。", "success")
    return redirect(url_for("order_tool.catalog_import"))


def _generate_date_orders(service_date: date):
    if _active_confirmed_orders(service_date):
        return 0, True

    requirements = _requirements_for_date(service_date)
    existing_orders = KitchenPurchaseOrder.query.filter_by(service_date=service_date).all()
    order = next((row for row in existing_orders if row.supplier_key == "daily"), None)
    if order is None and existing_orders:
        order = existing_orders[0]

    ingredient_rows = {
        data["ingredient"].id: data
        for supplier_rows in requirements.values()
        for data in supplier_rows.values()
    }
    if not ingredient_rows:
        for old_order in existing_orders:
            if old_order.status == "draft":
                db.session.delete(old_order)
        db.session.commit()
        return 0, False

    if order is None:
        order = KitchenPurchaseOrder(
            service_date=service_date,
            supplier_key="daily",
            supplier_name_snapshot="每日採購單",
            supplier_overridden=False,
            status="draft",
        )
        db.session.add(order)
        db.session.flush()
    else:
        order.supplier_id = None
        order.supplier_key = "daily"
        order.supplier_name_snapshot = "每日採購單"
        order.supplier_overridden = False
        order.status = "draft"

    existing_items = {item.ingredient_id: item for item in order.items}
    seen_ingredient_ids = set()
    for data in ingredient_rows.values():
        ing = data["ingredient"]
        units_per_purchase = ing.grams_per_purchase_unit or Decimal("0")
        increment = ing.order_increment or Decimal("0")
        if units_per_purchase <= 0 or increment <= 0:
            continue
        required_amount = data["required_amount"]
        required_qty = required_amount / units_per_purchase
        recommended = _round_up_increment(required_qty, increment)
        supplier_item = _supplier_item_match(data["supplier_id"], ing.id, ing.name)
        conversion_rule = _package_conversion_rule(
            supplier_item.package_conversion if supplier_item else None,
            ing.purchase_unit,
        )
        package_qty = _package_qty_from_rule(recommended, conversion_rule)
        seen_ingredient_ids.add(ing.id)
        item = existing_items.get(ing.id)
        if item is None:
            unit_price = ing.unit_price or Decimal("0")
            item = KitchenPurchaseOrderItem(
                order_id=order.id,
                ingredient_id=ing.id,
                supplier_id=data["supplier_id"],
                supplier_item_id=supplier_item.id if supplier_item else None,
                supplier_name_snapshot=data["supplier_name"],
                ingredient_name_snapshot=ing.name,
                base_unit_snapshot=ing.base_unit or "g",
                required_grams=required_amount,
                required_qty=required_qty,
                purchase_unit_snapshot=ing.purchase_unit,
                grams_per_purchase_unit_snapshot=units_per_purchase,
                recommended_order_qty=recommended,
                actual_order_qty=recommended,
                package_qty=package_qty,
                package_unit=conversion_rule["package_unit"] if conversion_rule else None,
                package_conversion_snapshot=supplier_item.package_conversion if supplier_item else None,
                unit_price_snapshot=unit_price,
                amount=recommended * unit_price,
                manual_override=False,
            )
            db.session.add(item)
        else:
            item.supplier_id = data["supplier_id"]
            item.supplier_item_id = supplier_item.id if supplier_item else None
            item.supplier_name_snapshot = data["supplier_name"]
            item.ingredient_name_snapshot = ing.name
            item.base_unit_snapshot = ing.base_unit or "g"
            item.required_grams = required_amount
            item.required_qty = required_qty
            item.purchase_unit_snapshot = ing.purchase_unit
            item.grams_per_purchase_unit_snapshot = units_per_purchase
            item.recommended_order_qty = recommended
            if not item.manual_override:
                item.actual_order_qty = recommended
                item.package_qty = package_qty
                item.package_unit = conversion_rule["package_unit"] if conversion_rule else None
                item.package_conversion_snapshot = supplier_item.package_conversion if supplier_item else None
                item.unit_price_snapshot = ing.unit_price or Decimal("0")
                item.amount = item.actual_order_qty * item.unit_price_snapshot

    for ingredient_id, item in existing_items.items():
        if ingredient_id not in seen_ingredient_ids:
            db.session.delete(item)

    for old_order in existing_orders:
        if old_order.id != order.id:
            db.session.delete(old_order)

    db.session.commit()
    return 1, False


def _purchase_item_supplier_name(item: KitchenPurchaseOrderItem) -> str:
    snapshot = (item.supplier_name_snapshot or "").strip()
    if snapshot and not snapshot.startswith("⚠"):
        return snapshot
    if item.supplier:
        return item.supplier.name
    return "⚠ 未指定供應商"


def _purchase_item_sort_key(item: KitchenPurchaseOrderItem):
    supplier_name = _purchase_item_supplier_name(item)
    unassigned = supplier_name.startswith("⚠")
    return (
        1 if unassigned else 0,
        supplier_name.casefold(),
        (item.ingredient_name_snapshot or "").casefold(),
    )


def _sorted_purchase_items(items):
    return sorted(items, key=_purchase_item_sort_key)


def _purchase_item_delivery_key(item: KitchenPurchaseOrderItem, service_date: date):
    delivery_date = item.delivery_date or service_date
    delivery_slot = item.delivery_slot or "上午"
    return delivery_date, delivery_slot


def _procurement_export_sort_key(row, service_date: date):
    item = row["item"]
    delivery_date, delivery_slot = _purchase_item_delivery_key(item, service_date)
    slot_order = 0 if delivery_slot == "上午" else 1
    return (
        delivery_date,
        slot_order,
        delivery_slot,
        _purchase_item_sort_key(item),
    )


def _procurement_rows(service_date: date):
    orders = (
        KitchenPurchaseOrder.query.filter(
            KitchenPurchaseOrder.service_date == service_date,
            KitchenPurchaseOrder.status == "draft",
        )
        .options(
            selectinload(KitchenPurchaseOrder.items).selectinload(KitchenPurchaseOrderItem.ingredient)
        )
        .order_by(KitchenPurchaseOrder.service_date, KitchenPurchaseOrder.supplier_name_snapshot)
        .all()
    )
    requirements_by_key = {}
    for ingredient_rows in _requirements_for_date(service_date).values():
        for data in ingredient_rows.values():
            requirements_by_key[data["ingredient"].id] = data
    rows = []
    for order in orders:
        for item in _sorted_purchase_items(order.items):
            requirement = requirements_by_key.get(item.ingredient_id, {})
            rows.append({
                "order": order,
                "item": item,
                "total_people": requirement.get("total_people", 0),
                "school_names": sorted(requirement.get("school_names", set())),
                "supplier_name": _purchase_item_supplier_name(item),
            })
    rows.sort(key=lambda row: (row["order"].service_date, _purchase_item_sort_key(row["item"])))
    return rows


def _procurement_conversion_options(rows):
    """Return matching supplier catalog conversions for each procurement row."""
    catalog = (
        KitchenSupplierItem.query.filter(
            KitchenSupplierItem.active.is_(True),
            KitchenSupplierItem.package_conversion.isnot(None),
        )
        .options(selectinload(KitchenSupplierItem.supplier))
        .all()
    )
    result = {}
    for row in rows:
        item = row["item"]
        item_name = (item.ingredient_name_snapshot or "").strip().casefold()
        matches = {}
        for supplier_item in catalog:
            exact_id = item.ingredient_id and supplier_item.ingredient_id == item.ingredient_id
            exact_name = supplier_item.name.strip().casefold() == item_name
            if not exact_id and not exact_name:
                continue
            supplier_name = supplier_item.supplier.name
            current = matches.get(supplier_name)
            if current and current[0]:
                continue
            rule = _package_conversion_rule(supplier_item.package_conversion, item.purchase_unit_snapshot)
            matches[supplier_name] = (bool(exact_id), {
                "supplierItemId": supplier_item.id,
                "label": supplier_item.package_conversion,
                "packageUnit": rule["package_unit"] if rule else None,
                "purchasePerPackage": _trim_decimal(rule["purchase_per_package"]) if rule else None,
            })
        result[str(item.id)] = {name: value for name, (_, value) in matches.items()}
    return result


def _production_sheet_data(service_date: date):
    """Build a dish-level daily usage sheet without duplicating procurement data."""
    plans = (
        KitchenMenuPlan.query.filter_by(service_date=service_date)
        .options(
            selectinload(KitchenMenuPlan.assignments).selectinload(KitchenMenuAssignment.school),
            selectinload(KitchenMenuPlan.items)
            .selectinload(KitchenMenuPlanItem.recipe)
            .selectinload(KitchenRecipe.ingredients)
            .selectinload(KitchenRecipeIngredient.ingredient),
        )
        .order_by(KitchenMenuPlan.meal_type, KitchenMenuPlan.name)
        .all()
    )

    purchase_orders = (
        KitchenPurchaseOrder.query.filter(
            KitchenPurchaseOrder.service_date == service_date,
            KitchenPurchaseOrder.status != "cancelled",
        )
        .options(selectinload(KitchenPurchaseOrder.items))
        .order_by(KitchenPurchaseOrder.id.desc())
        .all()
    )
    purchase_items = {}
    for order in sorted(purchase_orders, key=lambda row: (row.status != "draft", -row.id)):
        for item in order.items:
            if item.ingredient_id:
                purchase_items.setdefault(item.ingredient_id, item)

    grouped = {"regular": {}, "vegetarian": {}}
    for plan in plans:
        assignments = [
            assignment
            for assignment in plan.assignments
            if assignment.service_status == "serving" and assignment.headcount > 0
        ]
        people = sum(assignment.headcount for assignment in assignments)
        if people <= 0:
            continue
        variant = "vegetarian" if plan.name.endswith("素食菜單") else "regular"
        for menu_item in plan.items:
            recipe = menu_item.recipe
            dish = grouped[variant].setdefault(recipe.id, {
                "recipe": recipe,
                "headcount": 0,
                "school_names": set(),
                "sort_key": (menu_item.sort_order, recipe.category or "其他", recipe.name),
            })
            dish["headcount"] += people
            dish["school_names"].update(assignment.school.name for assignment in assignments)

    result = {"regular": [], "vegetarian": []}
    for variant, dishes in grouped.items():
        for dish in sorted(dishes.values(), key=lambda row: row["sort_key"]):
            components = []
            for component in dish["recipe"].ingredients:
                ingredient = component.ingredient
                per_person = component.grams_per_person or Decimal("0")
                base_amount = per_person * dish["headcount"]
                divisor = ingredient.grams_per_purchase_unit or Decimal("0")
                has_conversion = divisor > 0
                purchase_item = purchase_items.get(ingredient.id)
                components.append({
                    "ingredient": ingredient,
                    "per_person": per_person,
                    "per_person_unit": ingredient.base_unit or "g",
                    "theoretical_qty": base_amount / divisor if has_conversion else base_amount,
                    "theoretical_unit": ingredient.purchase_unit if has_conversion else (ingredient.base_unit or "g"),
                    "purchase_unit": ingredient.purchase_unit,
                    "purchase_item": purchase_item,
                    "pending": component.quantity_status == "pending" or per_person <= 0,
                    "conversion_missing": not has_conversion,
                })
            dish["school_names"] = sorted(dish["school_names"])
            dish["components"] = components
            result[variant].append(dish)
    return result


def _daily_kitchen_bucket(school_name: str) -> str:
    """把學校分到每日廚房表的三種出餐方式。"""
    normalized = re.sub(r"[\s　]+", "", school_name or "")
    if "小便當" in normalized and ("平鎮" in normalized or "鎮高" in normalized):
        return "small_bento"
    if "廣豐" in normalized:
        return "bento"
    if "小便當" not in normalized and ("平鎮高中" in normalized or "鎮高" in normalized):
        return "bento"
    return "combo"


def _recipe_ingredient_names(recipe: KitchenRecipe) -> str:
    return "、".join(
        component.ingredient.name
        for component in recipe.ingredients
        if component.ingredient and component.ingredient.name
    )


def _daily_kitchen_sheet_data(service_date: date):
    """產生「每日廚房表格」，人數依學校實際勾選的菜色分流。"""
    plans = (
        KitchenMenuPlan.query.filter_by(service_date=service_date, meal_type="午餐")
        .options(
            selectinload(KitchenMenuPlan.assignments).selectinload(KitchenMenuAssignment.school),
            selectinload(KitchenMenuPlan.items)
            .selectinload(KitchenMenuPlanItem.recipe)
            .selectinload(KitchenRecipe.ingredients)
            .selectinload(KitchenRecipeIngredient.ingredient),
        )
        .order_by(KitchenMenuPlan.name, KitchenMenuPlan.id)
        .all()
    )
    grouped = {"regular": {}, "vegetarian": {}}
    for plan in plans:
        assignments = [
            assignment for assignment in plan.assignments
            if assignment.service_status == "serving" and assignment.headcount > 0
        ]
        if not assignments:
            continue
        variant = "vegetarian" if plan.name.endswith("素食菜單") else "regular"
        for menu_item in plan.items:
            recipe = menu_item.recipe
            dish = grouped[variant].setdefault(recipe.id, {
                "recipe": recipe,
                "sort_key": (
                    CATEGORY_ORDER.get(recipe.category or "其他", len(CATEGORIES)),
                    menu_item.sort_order,
                    recipe.name.casefold(),
                ),
                "combo": 0,
                "bento": 0,
                "small_bento": 0,
                "school_counts": defaultdict(int),
            })
            if menu_item.sort_order < dish["sort_key"][1]:
                dish["sort_key"] = (
                    CATEGORY_ORDER.get(recipe.category or "其他", len(CATEGORIES)),
                    menu_item.sort_order,
                    recipe.name.casefold(),
                )
            for assignment in assignments:
                bucket = _daily_kitchen_bucket(assignment.school.name)
                dish[bucket] += max(assignment.headcount, 0)
                dish["school_counts"][assignment.school.name] += max(assignment.headcount, 0)

    recipe_ids = [recipe_id for dishes in grouped.values() for recipe_id in dishes]
    notes = {
        (row.variant, row.recipe_id): row
        for row in KitchenDailyDishNote.query.filter(
            KitchenDailyDishNote.service_date == service_date,
            KitchenDailyDishNote.recipe_id.in_(recipe_ids or [-1]),
        ).all()
    }
    result = {"regular": [], "vegetarian": []}
    for variant, dishes in grouped.items():
        for dish in sorted(dishes.values(), key=lambda row: row["sort_key"]):
            # 小便當 300 份不移轉；達 360 份才抽 60 份改用合菜出餐。
            if variant == "regular" and dish["small_bento"] >= 360:
                transferred = 60
                dish["small_bento"] -= transferred
                dish["combo"] += transferred
            default_ingredients = _recipe_ingredient_names(dish["recipe"])
            note = notes.get((variant, dish["recipe"].id))
            dish["ingredients_text"] = (
                note.ingredients_text if note and note.ingredients_text else default_ingredients
            )
            if note and note.combo_count is not None:
                dish["combo"] = note.combo_count
            if note and note.bento_count is not None:
                dish["bento"] = note.bento_count
            if note and note.small_bento_count is not None:
                dish["small_bento"] = note.small_bento_count
            dish["school_rows"] = [
                {"name": name, "headcount": headcount}
                for name, headcount in sorted(
                    dish["school_counts"].items(), key=lambda row: row[0].casefold()
                )
            ]
            dish["school_summary"] = "、".join(
                f"{row['name']} {row['headcount']} 人" for row in dish["school_rows"]
            )
            dish["total"] = dish["combo"] + dish["bento"] + dish["small_bento"]
            result[variant].append(dish)
    return result


@order_bp.route("/summary/daily-kitchen-sheet", methods=["GET", "POST"])
def daily_kitchen_sheet():
    service_date = _date(
        request.values.get("date"), default=date.today()
    ) or date.today()
    sheets = _daily_kitchen_sheet_data(service_date)
    if request.method == "POST":
        autosave = request.headers.get("X-Requested-With") == "daily-kitchen-autosave"
        updates = []
        for variant, dishes in sheets.items():
            for dish in dishes:
                field_name = f"ingredients_{variant}_{dish['recipe'].id}"
                if field_name not in request.form:
                    continue
                ingredients_text = request.form.get(field_name, "").strip()[:2000]
                if not ingredients_text:
                    ingredients_text = _recipe_ingredient_names(dish["recipe"])
                counts = {}
                for key in ("combo", "bento", "small_bento"):
                    value = _int(
                        request.form.get(f"{key}_{variant}_{dish['recipe'].id}"),
                        default=None,
                    )
                    if value is None or value < 0:
                        db.session.rollback()
                        if autosave:
                            return {"message": "合菜、便當與小便當數量必須是 0 以上的整數。"}, 400
                        flash("合菜、便當與小便當數量必須是 0 以上的整數。", "error")
                        return redirect(url_for(
                            "order_tool.daily_kitchen_sheet", date=service_date.isoformat()
                        ))
                    counts[key] = value
                updates.append((variant, dish, ingredients_text, counts))

        for variant, dish, ingredients_text, counts in updates:
            note = KitchenDailyDishNote.query.filter_by(
                service_date=service_date,
                variant=variant,
                recipe_id=dish["recipe"].id,
            ).one_or_none()
            if note is None:
                note = KitchenDailyDishNote(
                    service_date=service_date,
                    variant=variant,
                    recipe_id=dish["recipe"].id,
                )
                db.session.add(note)
            note.ingredients_text = ingredients_text
            note.combo_count = counts["combo"]
            note.bento_count = counts["bento"]
            note.small_bento_count = counts["small_bento"]
        db.session.commit()
        if autosave:
            return {"message": "已儲存"}
        flash("每日廚房表格的食材與數字已儲存，現在可以匯出 Excel。", "success")
        return redirect(url_for("order_tool.daily_kitchen_sheet", date=service_date.isoformat()))
    return render_template(
        "kitchen/daily_kitchen_sheet.html",
        service_date=service_date,
        previous_date=service_date - timedelta(days=1),
        next_date=service_date + timedelta(days=1),
        week_start=service_date - timedelta(days=service_date.weekday()),
        sheets=sheets,
    )


def _write_daily_kitchen_export_section(sheet, start_row: int, label: str, dishes: list[dict]):
    header_row = start_row + 1
    headers = [label, "", "合菜", "班級數", "便當", "小便當", "總計", "供餐學校／人數"]
    for column, value in enumerate(headers, start=1):
        cell = sheet.cell(header_row, column, value)
        cell.font = Font(name="Microsoft JhengHei", size=15, bold=True)
        cell.alignment = Alignment(vertical="center")
        cell.border = Border(bottom=Side(style="medium", color="000000"))
    sheet.row_dimensions[header_row].height = 27

    for row_number, dish in enumerate(dishes, start=header_row + 1):
        values = [
            dish["recipe"].name,
            dish["ingredients_text"],
            dish["combo"] or None,
            None,
            dish["bento"] or None,
            dish["small_bento"] or None,
            dish["total"] or None,
            dish["school_summary"],
        ]
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row_number, column, value)
            cell.font = Font(name="Microsoft JhengHei", size=14, bold=True)
            cell.alignment = Alignment(
                horizontal="right" if 3 <= column <= 7 else "left",
                vertical="center",
                wrap_text=column in {2, 8},
            )
            cell.border = Border(bottom=Side(style="thin", color="000000"))
            if 3 <= column <= 7:
                cell.number_format = "#,##0"
        sheet.row_dimensions[row_number].height = 31
    return max(header_row + len(dishes), header_row)


@order_bp.get("/summary/daily-kitchen-sheet.xlsx")
def daily_kitchen_sheet_export():
    service_date = _date(request.args.get("date"), default=date.today()) or date.today()
    sheets = _daily_kitchen_sheet_data(service_date)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = service_date.strftime("%m%d")
    sheet.sheet_view.showGridLines = True
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 1
    sheet.page_margins.left = 0.2
    sheet.page_margins.right = 0.2
    sheet.column_dimensions["A"].width = 22
    sheet.column_dimensions["B"].width = 58
    for column in ("C", "D", "E", "F", "G"):
        sheet.column_dimensions[column].width = 13
    sheet.column_dimensions["H"].width = 48

    weekday = "一二三四五六日"[service_date.weekday()]
    current_row = 1
    for index, (variant, label) in enumerate((("regular", "葷"), ("vegetarian", "素"))):
        if index:
            current_row += 2
        sheet.cell(current_row, 1, service_date)
        sheet.cell(current_row, 1).number_format = 'm"月"d"日"'
        sheet.cell(current_row, 2, f"（{weekday}）")
        for column in (1, 2):
            sheet.cell(current_row, column).font = Font(
                name="Microsoft JhengHei", size=19, bold=True
            )
            sheet.cell(current_row, column).alignment = Alignment(vertical="center")
        sheet.row_dimensions[current_row].height = 34
        current_row = _write_daily_kitchen_export_section(
            sheet, current_row, label, sheets[variant]
        )
    sheet.print_area = f"A1:H{current_row}"

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name=f"每日廚房表格_{service_date.isoformat()}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _production_order_for_date(service_date: date):
    return (
        KitchenPurchaseOrder.query.filter(
            KitchenPurchaseOrder.service_date == service_date,
            KitchenPurchaseOrder.status != "cancelled",
            KitchenPurchaseOrder.items.any(),
        )
        .order_by(KitchenPurchaseOrder.id.desc())
        .first()
    )


def _production_nav_state():
    requested_date = _date(request.args.get("date"))
    if requested_date:
        return requested_date, _production_order_for_date(requested_date) is not None
    latest_order = (
        KitchenPurchaseOrder.query.filter(
            KitchenPurchaseOrder.status != "cancelled",
            KitchenPurchaseOrder.items.any(),
        )
        .order_by(KitchenPurchaseOrder.service_date.desc(), KitchenPurchaseOrder.id.desc())
        .first()
    )
    return (latest_order.service_date, True) if latest_order else (date.today(), False)


def _require_production_order(service_date: date):
    if _production_order_for_date(service_date):
        return None
    flash("請先產生這一天的採購明細，才能查看菜色用量表。", "warning")
    return redirect(url_for("order_tool.procurement", date=service_date.isoformat()))


@order_bp.get("/summary/production-sheet")
def production_sheet():
    service_date = _date(request.args.get("date"), default=date.today()) or date.today()
    blocked = _require_production_order(service_date)
    if blocked:
        return blocked
    variant = request.args.get("variant", "regular")
    if variant not in {"regular", "vegetarian"}:
        variant = "regular"
    sheets = _production_sheet_data(service_date)
    return render_template(
        "kitchen/production_sheet.html",
        service_date=service_date,
        previous_date=service_date - timedelta(days=1),
        next_date=service_date + timedelta(days=1),
        week_start=service_date - timedelta(days=service_date.weekday()),
        variant=variant,
        dishes=sheets[variant],
        variant_counts={key: len(value) for key, value in sheets.items()},
    )


def _write_production_export_sheet(sheet, service_date: date, variant_label: str, dishes: list[dict]):
    """Match the two-column, dish-by-dish production worksheet used on site."""
    weekday = "一二三四五六日"[service_date.weekday()]
    sheet.sheet_properties.tabColor = "1769AA" if variant_label == "葷食" else "08735C"
    sheet.sheet_view.showGridLines = True
    sheet.freeze_panes = None
    sheet.auto_filter.ref = None
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.outlinePr.summaryBelow = True
    sheet.print_options.horizontalCentered = True
    sheet.page_margins.left = 0.25
    sheet.page_margins.right = 0.25
    sheet.page_margins.top = 0.35
    sheet.page_margins.bottom = 0.35

    default_font = Font(name="Microsoft JhengHei", size=10, color="000000")
    title_font = Font(name="Microsoft JhengHei", size=16, color="000000")
    main_title_font = Font(name="Microsoft JhengHei", size=20, color="000000")
    actual_header_font = Font(name="Microsoft JhengHei", size=10, bold=True, color="FF0000")
    actual_value_font = Font(name="Microsoft JhengHei", size=10, color="FF0000")
    thin_black = Side(style="thin", color="000000")
    table_border = Border(left=thin_black, right=thin_black, top=thin_black, bottom=thin_black)

    sheet["B1"] = service_date
    sheet["B1"].number_format = 'm"月"d"日"'
    sheet["B1"].font = main_title_font
    sheet["B1"].alignment = Alignment(horizontal="center")
    sheet["D1"] = weekday
    sheet["D1"].font = main_title_font
    sheet["D1"].alignment = Alignment(horizontal="center")
    sheet["E1"] = "供應份數："
    sheet["E1"].font = title_font
    sheet["H1"] = max((dish["headcount"] for dish in dishes), default=0)
    sheet["H1"].font = title_font
    sheet["H1"].number_format = "#,##0"
    sheet["B2"] = "採購叫貨量"
    sheet["B2"].font = title_font
    if variant_label == "素食":
        sheet["D2"] = "(素)"
        sheet["D2"].font = title_font
    sheet.row_dimensions[1].height = 28
    sheet.row_dimensions[2].height = 24

    widths = {
        "A": 3.5, "B": 18, "C": 10, "D": 12, "E": 15, "F": 12,
        "G": 6, "H": 15, "I": 6, "J": 24,
        "K": 3.5, "L": 18, "M": 10, "N": 12, "O": 15, "P": 12,
        "Q": 6, "R": 15, "S": 6, "T": 24,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    midpoint = (len(dishes) + 1) // 2
    columns = (dishes[:midpoint], dishes[midpoint:])
    final_row = 3
    for side_index, side_dishes in enumerate(columns):
        row = 3
        index_col = 1 if side_index == 0 else 11
        material_col = index_col + 1
        note_col = index_col + 9
        for dish_index, dish in enumerate(side_dishes, 1 + (midpoint if side_index else 0)):
            slots = max(4, len(dish["components"]))
            title_row = row
            header_row = row + 1
            first_data_row = row + 2
            last_data_row = first_data_row + slots - 1
            total_row = last_data_row + 1

            sheet.cell(title_row, index_col, dish_index)
            sheet.cell(title_row, material_col, dish["recipe"].name)
            sheet.cell(title_row, material_col + 3, "供應份數：")
            sheet.cell(title_row, material_col + 4, dish["headcount"])
            for column in (index_col, material_col, material_col + 3, material_col + 4):
                sheet.cell(title_row, column).font = title_font
                sheet.cell(title_row, column).alignment = Alignment(vertical="center")
            sheet.cell(title_row, material_col + 4).number_format = "#,##0"
            sheet.row_dimensions[title_row].height = 24

            headers = ["材料明細", "單量", "單份用量", "生產用量(總餐數)", "總量", "", "實際叫貨量", ""]
            for offset, header in enumerate(headers):
                cell = sheet.cell(header_row, material_col + offset, header)
                cell.font = actual_header_font if offset == 6 else default_font
                cell.alignment = Alignment(horizontal="center", vertical="center", shrink_to_fit=True)
                cell.border = table_border
            note_header = sheet.cell(header_row, note_col, "現場備註")
            note_header.font = default_font
            note_header.alignment = Alignment(horizontal="center", vertical="center")
            sheet.row_dimensions[header_row].height = 22

            for slot in range(slots):
                data_row = first_data_row + slot
                component = dish["components"][slot] if slot < len(dish["components"]) else None
                for offset in range(8):
                    cell = sheet.cell(data_row, material_col + offset)
                    cell.font = default_font
                    cell.alignment = Alignment(
                        horizontal="right" if offset in {1, 2, 3, 4, 6} else "left",
                        vertical="center",
                        shrink_to_fit=offset not in {0},
                    )
                    cell.border = table_border
                if component:
                    item = component["purchase_item"]
                    divisor = component["ingredient"].grams_per_purchase_unit or Decimal("0")
                    sheet.cell(data_row, material_col, component["ingredient"].name)
                    sheet.cell(data_row, material_col + 1, float(component["per_person"]) if not component["pending"] else None)
                    if divisor > 0 and not component["pending"]:
                        amount_ref = sheet.cell(data_row, material_col + 1).coordinate
                        sheet.cell(data_row, material_col + 2, f"={amount_ref}/{_trim_decimal(divisor)}")
                    sheet.cell(data_row, material_col + 3, f"={sheet.cell(title_row, material_col + 4).coordinate}")
                    portion_ref = sheet.cell(data_row, material_col + 2).coordinate
                    people_ref = sheet.cell(data_row, material_col + 3).coordinate
                    sheet.cell(data_row, material_col + 4, f"={portion_ref}*{people_ref}")
                    sheet.cell(data_row, material_col + 5, component["theoretical_unit"])
                    actual_uses_package = bool(item and item.package_qty is not None and item.package_unit)
                    actual_qty = item.package_qty if actual_uses_package else (item.actual_order_qty if item else None)
                    actual_unit = item.package_unit if actual_uses_package else (
                        item.purchase_unit_snapshot if item else component["purchase_unit"]
                    )
                    sheet.cell(data_row, material_col + 6, float(actual_qty) if actual_qty is not None else None)
                    sheet.cell(data_row, material_col + 6).font = actual_value_font
                    sheet.cell(data_row, material_col + 7, actual_unit)
                    if item and item.note:
                        note = sheet.cell(data_row, note_col, item.note)
                        note.font = default_font
                        note.alignment = Alignment(vertical="center", wrap_text=True)
                else:
                    sheet.cell(data_row, material_col + 4, 0)
                    sheet.cell(data_row, material_col + 5, "Kg")
                    sheet.cell(data_row, material_col + 7, "Kg")
                sheet.cell(data_row, material_col + 1).number_format = "#,##0.###"
                sheet.cell(data_row, material_col + 2).number_format = "0.####"
                sheet.cell(data_row, material_col + 3).number_format = "#,##0"
                sheet.cell(data_row, material_col + 4).number_format = "#,##0.####"
                sheet.cell(data_row, material_col + 6).number_format = "#,##0.####"
                sheet.row_dimensions[data_row].height = 20

            sum_cell = sheet.cell(total_row, material_col + 1)
            sum_cell.value = f"=SUM({sheet.cell(first_data_row, material_col + 1).coordinate}:{sheet.cell(last_data_row, material_col + 1).coordinate})"
            sum_cell.font = default_font
            sum_cell.alignment = Alignment(horizontal="right")
            sum_cell.number_format = "#,##0.###"
            sheet.row_dimensions[total_row].height = 20
            row = total_row + 1
            final_row = max(final_row, total_row)

    if not dishes:
        sheet["B4"] = f"本日無{variant_label}菜色資料"
        sheet["B4"].font = title_font
        final_row = 4

    sheet.print_area = f"A1:T{final_row}"


@order_bp.get("/summary/production-sheet.xlsx")
def production_sheet_export():
    service_date = _date(request.args.get("date"), default=date.today()) or date.today()
    blocked = _require_production_order(service_date)
    if blocked:
        return blocked
    sheets = _production_sheet_data(service_date)
    workbook = Workbook()
    workbook.calculation.calcMode = "auto"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    regular_sheet = workbook.active
    regular_sheet.title = "葷食"
    vegetarian_sheet = workbook.create_sheet("素食")
    _write_production_export_sheet(regular_sheet, service_date, "葷食", sheets["regular"])
    _write_production_export_sheet(vegetarian_sheet, service_date, "素食", sheets["vegetarian"])
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name=f"菜色用量表_{service_date.isoformat()}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@order_bp.get("/summary/procurement")
def procurement():
    service_date = _date(request.args.get("date"), default=date.today()) or date.today()
    rows = _procurement_rows(service_date)
    suppliers = KitchenSupplier.query.filter_by(active=True).order_by(KitchenSupplier.name).all()
    pending_recipes = set()
    plans = KitchenMenuPlan.query.filter_by(service_date=service_date).all()
    for plan in plans:
        if not any(x.service_status == "serving" and x.headcount > 0 for x in plan.assignments):
            continue
        for menu_item in plan.items:
            if any(component.quantity_status == "pending" or (component.grams_per_person or 0) <= 0
                   for component in menu_item.recipe.ingredients):
                pending_recipes.add(menu_item.recipe.name)
    return render_template(
        "kitchen/procurement.html",
        rows=rows,
        conversion_options=_procurement_conversion_options(rows),
        suppliers=suppliers,
        package_units=PACKAGE_UNITS,
        pending_recipes=sorted(pending_recipes),
        service_date=service_date,
        previous_date=service_date - timedelta(days=1),
        next_date=service_date + timedelta(days=1),
        week_start=service_date - timedelta(days=service_date.weekday()),
    )


@order_bp.get("/summary/procurement.xlsx")
def procurement_export():
    service_date = _date(request.args.get("date"), default=date.today()) or date.today()
    rows = sorted(
        _procurement_rows(service_date),
        key=lambda row: _procurement_export_sort_key(row, service_date),
    )

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "每日訂購單"

    headers = ["廠商", "品名", "數量", "單位", "備註"]
    header_fill = PatternFill("solid", fgColor="F2F2F2")
    ordered_fill = PatternFill("solid", fgColor="FFF1A8")
    thin = Side(style="thin", color="000000")
    table_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    delivery_groups = []
    for row in rows:
        delivery_key = _purchase_item_delivery_key(row["item"], service_date)
        if not delivery_groups or delivery_groups[-1][0] != delivery_key:
            delivery_groups.append((delivery_key, []))
        delivery_groups[-1][1].append(row)
    if not delivery_groups:
        delivery_groups.append(((service_date, "上午"), []))

    for group_index, ((delivery_date, delivery_slot), group_rows) in enumerate(delivery_groups):
        title_row = 1 if group_index == 0 else sheet.max_row + 2
        sheet.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=2)
        sheet.merge_cells(start_row=title_row, start_column=3, end_row=title_row, end_column=5)
        sheet.cell(title_row, 1, "每日訂購單")
        sheet.cell(
            title_row,
            3,
            f"進貨日期：{delivery_date.year - 1911:03d}/{delivery_date.month:02d}/{delivery_date.day:02d} {delivery_slot}",
        )
        sheet.cell(title_row, 1).font = Font(size=16, bold=True)
        sheet.cell(title_row, 3).font = Font(size=12, bold=True)
        sheet.cell(title_row, 1).alignment = Alignment(horizontal="center", vertical="center")
        sheet.cell(title_row, 3).alignment = Alignment(horizontal="right", vertical="center")
        sheet.row_dimensions[title_row].height = 28

        sheet.append(headers)
        header_row = sheet.max_row
        for cell in sheet[header_row]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = table_border

        previous_supplier = None
        for row in group_rows:
            item = row["item"]
            supplier_name = row["supplier_name"]
            use_package = item.package_qty is not None and bool(item.package_unit)
            quantity = item.package_qty if use_package else item.actual_order_qty
            unit = item.package_unit if use_package else item.purchase_unit_snapshot
            notes = []
            if item.note:
                notes.append(item.note.strip())
            if item.package_conversion_snapshot:
                notes.append(item.package_conversion_snapshot.strip())
            sheet.append([
                supplier_name if supplier_name != previous_supplier else "",
                item.ingredient_name_snapshot,
                float(quantity or 0),
                unit,
                "；".join(dict.fromkeys(note for note in notes if note)),
            ])
            previous_supplier = supplier_name
            data_row = sheet.max_row
            sheet.cell(data_row, 3).number_format = "0.00"
            for cell in sheet[data_row]:
                cell.border = table_border
                cell.alignment = Alignment(
                    horizontal="right" if cell.column == 3 else "left",
                    vertical="center",
                    wrap_text=True,
                )
            if item.ordered:
                for cell in sheet[data_row]:
                    cell.fill = ordered_fill

    sheet.freeze_panes = "A3"
    widths = (18, 28, 14, 10, 42)
    for column, width in zip("ABCDE", widths):
        sheet.column_dimensions[column].width = width
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.orientation = "portrait"
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.print_title_rows = "2:2"
    sheet.print_area = f"A1:E{max(sheet.max_row, 2)}"
    sheet.sheet_view.showGridLines = False
    sheet.oddFooter.center.text = "第 &P / &N 頁"
    sheet.oddFooter.center.size = 9

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name=f"每日訂購單-{service_date.isoformat()}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@order_bp.post("/summary/procurement/generate")
def procurement_generate():
    service_date = _date(request.form.get("date"), default=date.today()) or date.today()
    missing_schools = _missing_school_names_for_date(service_date)
    if missing_schools:
        flash(
            "尚有學校未完成菜單勾選：" + "、".join(missing_schools) + "。請先完成後再產生採購單。",
            "warning",
        )
        return redirect(url_for(
            "order_tool.school_menus",
            week=(service_date - timedelta(days=service_date.weekday())).isoformat(),
        ))
    created, blocked = _generate_date_orders(service_date)
    if created:
        flash(f"已重新計算 {service_date.strftime('%m/%d')} 每日採購單。", "success")
    if blocked:
        flash("這一天已有已確認採購單，未重新計算。", "warning")
    if not created and not blocked:
        flash("目前沒有可計算的食材需求，請先確認各校菜單、人數與配方。", "warning")
    return redirect(url_for("order_tool.procurement", date=service_date.isoformat()))


@order_bp.post("/summary/procurement/save")
def procurement_save():
    service_date = _date(request.form.get("date"), default=date.today()) or date.today()
    item_ids = {_int(raw, default=0) or 0 for raw in request.form.getlist("item_ids")}
    updated = 0
    conversion_updates = 0
    for item_id in item_ids:
        item = db.session.get(KitchenPurchaseOrderItem, item_id)
        if not item or item.order.status != "draft" or item.order.service_date != service_date:
            continue
        result, error = _apply_procurement_item_values(
            item,
            actual_raw=request.form.get(f"actual_{item.id}"),
            package_qty_raw=request.form.get(f"package_qty_{item.id}"),
            package_unit_raw=request.form.get(f"package_unit_{item.id}"),
            delivery_date_raw=request.form.get(f"delivery_date_{item.id}"),
            delivery_slot_raw=request.form.get(f"delivery_slot_{item.id}"),
            supplier_name_raw=request.form.get(f"supplier_{item.id}"),
        )
        if error:
            db.session.rollback()
            flash(error, "error")
            return redirect(url_for("order_tool.procurement", date=service_date.isoformat()))
        item.ordered = request.form.get(f"ordered_{item.id}") == "1"
        conversion_updates += int(result["conversion_changed"])
        updated += 1
    db.session.commit()
    message = f"已儲存 {updated} 筆採購資料。"
    if conversion_updates:
        message += f" 同步記住 {conversion_updates} 筆廠商換算。"
    flash(message, "success")
    return redirect(url_for("order_tool.procurement", date=service_date.isoformat()))


def _apply_procurement_item_values(
    item: KitchenPurchaseOrderItem,
    *,
    actual_raw,
    package_qty_raw,
    package_unit_raw,
    delivery_date_raw,
    delivery_slot_raw,
    supplier_name_raw,
):
    actual = _decimal(actual_raw, default=None)
    package_qty_text = str(package_qty_raw or "").strip()
    package_qty = _decimal(package_qty_raw, default=None)
    package_unit = str(package_unit_raw or "").strip()[:20] or None
    delivery_date = _date(delivery_date_raw, default=None)
    delivery_slot = str(delivery_slot_raw or "上午").strip()
    supplier_name = str(supplier_name_raw or "").strip()[:100]
    if supplier_name.startswith("⚠") or supplier_name == "未指定供應商":
        supplier_name = ""

    if actual is None or actual < 0:
        return None, f"「{item.ingredient_name_snapshot}」的實際採購量不正確。"
    if package_qty_text and package_qty is None:
        return None, f"「{item.ingredient_name_snapshot}」的包裝數量不正確。"
    if package_qty is not None and package_qty < 0:
        return None, f"「{item.ingredient_name_snapshot}」的包裝數量不可小於 0。"
    if not delivery_date or delivery_slot not in ("上午", "下午"):
        return None, f"「{item.ingredient_name_snapshot}」的交貨日期或時段不正確。"

    supplier = None
    supplier_created = False
    if supplier_name:
        supplier = KitchenSupplier.query.filter(
            db.func.lower(KitchenSupplier.name) == supplier_name.lower()
        ).first()
        if supplier is None:
            supplier = KitchenSupplier(name=supplier_name, note="由採購明細新增", active=True)
            db.session.add(supplier)
            db.session.flush()
            supplier_created = True
        elif not supplier.active:
            supplier.active = True

    supplier_item = _supplier_item_match(
        supplier.id if supplier else None,
        item.ingredient_id,
        item.ingredient_name_snapshot,
    )
    conversion_rule = _package_conversion_rule(
        supplier_item.package_conversion if supplier_item else None,
        item.purchase_unit_snapshot,
    )
    if package_qty is None and conversion_rule:
        package_qty = _package_qty_from_rule(actual, conversion_rule)
        package_unit = conversion_rule["package_unit"]

    conversion_changed = False
    if supplier and package_qty is not None and package_qty > 0 and package_unit and actual > 0:
        supplier_item, conversion_changed = _remember_supplier_conversion(
            supplier, item, actual, package_qty, package_unit
        )

    item.supplier_id = supplier.id if supplier else None
    item.supplier_item_id = supplier_item.id if supplier_item else None
    item.supplier_name_snapshot = supplier.name if supplier else "⚠ 未指定供應商"
    item.actual_order_qty = actual
    item.package_qty = package_qty
    item.package_unit = package_unit
    item.package_conversion_snapshot = supplier_item.package_conversion if supplier_item else None
    item.amount = actual * (item.unit_price_snapshot or Decimal("0"))
    item.delivery_date = delivery_date
    item.delivery_slot = delivery_slot
    item.manual_override = True
    if item.ingredient:
        item.ingredient.supplier_id = supplier.id if supplier else None

    return {
        "supplier": supplier,
        "supplier_created": supplier_created,
        "supplier_item": supplier_item,
        "conversion_changed": conversion_changed,
    }, None


@order_bp.post("/summary/procurement/items/<int:item_id>/save")
def procurement_item_autosave(item_id: int):
    item = db.session.get(KitchenPurchaseOrderItem, item_id)
    if not item:
        return {"message": "找不到採購品項。"}, 404
    if item.order.status != "draft":
        return {"message": "已確認的採購單不可直接修改。"}, 409

    result, error = _apply_procurement_item_values(
        item,
        actual_raw=request.form.get("actual"),
        package_qty_raw=request.form.get("package_qty"),
        package_unit_raw=request.form.get("package_unit"),
        delivery_date_raw=request.form.get("delivery_date"),
        delivery_slot_raw=request.form.get("delivery_slot"),
        supplier_name_raw=request.form.get("supplier_name"),
    )
    if error:
        db.session.rollback()
        return {"message": error}, 400
    db.session.commit()
    supplier = result["supplier"]
    return {
        "message": "已儲存",
        "supplierName": supplier.name if supplier else "⚠ 未指定供應商",
        "supplierCreated": result["supplier_created"],
        "packageQty": _trim_decimal(item.package_qty) if item.package_qty is not None else "",
        "packageUnit": item.package_unit or "",
        "conversionLabel": item.package_conversion_snapshot or "",
        "amount": _trim_decimal(item.amount),
    }


@order_bp.get("/purchases")
def purchases():
    start = _date(request.args.get("start"), default=date.today() - timedelta(days=7)) or date.today()
    end = _date(request.args.get("end"), default=date.today() + timedelta(days=14)) or start
    if end < start:
        start, end = end, start
    orders = KitchenPurchaseOrder.query.filter(KitchenPurchaseOrder.service_date.between(start, end)).order_by(
        KitchenPurchaseOrder.service_date.desc()
    ).all()
    supplier_names_by_order = {}
    for order in orders:
        names = []
        for item in _sorted_purchase_items(order.items):
            name = _purchase_item_supplier_name(item)
            if name not in names:
                names.append(name)
        supplier_names_by_order[order.id] = names
    return render_template(
        "kitchen/purchases.html",
        orders=orders,
        supplier_names_by_order=supplier_names_by_order,
        start=start,
        end=end,
    )


@order_bp.get("/purchase")
def purchase_legacy_redirect():
    args = {}
    if request.args.get("start"):
        args["start"] = request.args["start"]
    if request.args.get("end"):
        args["end"] = request.args["end"]
    return redirect(url_for("order_tool.purchases", **args))


@order_bp.post("/purchases/generate")
def generate_purchases():
    start = _date(request.form.get("start"), default=date.today()) or date.today()
    end = _date(request.form.get("end"), default=start) or start
    if end < start:
        start, end = end, start
    if (end - start).days > 31:
        flash("一次最多產生 32 天的採購草稿。", "error")
        return redirect(url_for("order_tool.purchases", start=start, end=end))
    created = 0
    blocked_dates = []
    day = start
    while day <= end:
        count, blocked = _generate_date_orders(day)
        created += count
        if blocked:
            blocked_dates.append(str(day))
        day += timedelta(days=1)
    if created:
        flash(f"已建立 / 更新 {created} 天的採購草稿；每一天只會有一張。", "success")
    if blocked_dates:
        flash("以下日期已有已確認採購單，因此沒有自動重算：" + "、".join(blocked_dates), "warning")
    if not created and not blocked_dates:
        flash("此期間沒有可計算的菜單需求。請確認菜色配方與學校人數。", "warning")
    return redirect(url_for("order_tool.purchases", start=start, end=end))


@order_bp.get("/purchases/<int:order_id>")
def purchase_detail(order_id: int):
    order = db.session.get(KitchenPurchaseOrder, order_id)
    if not order:
        abort(404)
    return render_template(
        "kitchen/purchase_detail.html",
        order=order,
        items=_sorted_purchase_items(order.items),
        suppliers=KitchenSupplier.query.order_by(KitchenSupplier.active.desc(), KitchenSupplier.name).all(),
        total=_order_total(order),
        do_print=request.args.get("print") == "1",
    )


@order_bp.post("/purchases/<int:order_id>/update")
def purchase_update(order_id: int):
    order = db.session.get(KitchenPurchaseOrder, order_id)
    if not order:
        abort(404)
    if order.status != "draft":
        flash("只有草稿採購單可以修改。", "error")
        return redirect(url_for("order_tool.purchase_detail", order_id=order_id))
    order.note = request.form.get("note", "").strip() or None
    return _commit("採購單備註已更新。", "order_tool.purchase_detail", order_id=order_id)


@order_bp.post("/purchase-items/<int:item_id>/update")
def purchase_item_update(item_id: int):
    item = db.session.get(KitchenPurchaseOrderItem, item_id)
    if not item:
        abort(404)
    if item.order.status != "draft":
        flash("已確認的採購單不可直接修改。", "error")
        return redirect(url_for("order_tool.purchase_detail", order_id=item.order_id))
    actual = _decimal(request.form.get("actual_order_qty"))
    price = _decimal(request.form.get("unit_price"))
    if actual is None or actual < 0 or price is None or price < 0:
        flash("實際叫貨量與單價不可為負數。", "error")
        return redirect(url_for("order_tool.purchase_detail", order_id=item.order_id))
    item.actual_order_qty = actual
    item.unit_price_snapshot = price
    item.amount = actual * price
    item.note = request.form.get("note", "").strip() or None
    item.manual_override = True
    db.session.commit()
    flash("採購項目已更新；之後重新產生需求也不會洗掉這次人工調整。", "success")
    return redirect(url_for("order_tool.purchase_detail", order_id=item.order_id))


@order_bp.post("/purchase-items/<int:item_id>/ordered")
def purchase_item_ordered(item_id: int):
    item = db.session.get(KitchenPurchaseOrderItem, item_id)
    if not item:
        abort(404)
    item.ordered = request.form.get("ordered") == "1"
    db.session.commit()
    if request.headers.get("X-Requested-With") == "procurement-tracking":
        return "", 204
    flash(f"「{item.ingredient_name_snapshot}」叫貨狀態已更新。", "success")
    return redirect(request.referrer or url_for("order_tool.purchase_detail", order_id=item.order_id))


@order_bp.post("/purchases/<int:order_id>/ordered")
def purchase_order_ordered(order_id: int):
    order = db.session.get(KitchenPurchaseOrder, order_id)
    if not order:
        abort(404)
    ordered = request.form.get("ordered") == "1"
    for item in order.items:
        item.ordered = ordered
    db.session.commit()
    flash(f"{order.service_date} 全部品項已標記為{'已叫貨' if ordered else '未叫貨'}。", "success")
    return redirect(request.referrer or url_for("order_tool.purchases", start=order.service_date, end=order.service_date))


@order_bp.post("/purchases/<int:order_id>/delete")
def purchase_delete(order_id: int):
    order = db.session.get(KitchenPurchaseOrder, order_id)
    if not order:
        abort(404)

    return_to = request.form.get("return_to")
    if return_to == "dashboard":
        redirect_url = url_for("order_tool.index")
    else:
        redirect_url = url_for(
            "order_tool.purchases",
            start=request.form.get("start") or order.service_date,
            end=request.form.get("end") or order.service_date,
        )

    # 除了前端確認視窗，後端也要求明確的確認值，避免誤送 POST 就刪除。
    if request.form.get("confirm_delete") != "1":
        flash("未完成刪除確認，採購單已保留。", "error")
        return redirect(redirect_url)

    service_date = order.service_date
    item_count = len(order.items)
    db.session.delete(order)
    db.session.commit()
    flash(f"已刪除 {service_date} 的採購單與 {item_count} 筆採購品項。", "success")
    return redirect(redirect_url)


@order_bp.post("/purchases/<int:order_id>/confirm")
def purchase_confirm(order_id: int):
    order = db.session.get(KitchenPurchaseOrder, order_id)
    if not order:
        abort(404)
    if order.status != "draft":
        flash("這張採購單目前不是草稿。", "error")
        return redirect(url_for("order_tool.purchase_detail", order_id=order_id))
    if not order.items:
        flash("空白採購單不可確認。", "error")
        return redirect(url_for("order_tool.purchase_detail", order_id=order_id))
    if any(not item.supplier_id for item in order.items):
        flash("正式確認前，每個食材都必須指定供應商。", "error")
        return redirect(url_for("order_tool.purchase_detail", order_id=order_id))
    order.status = "confirmed"
    for plan in KitchenMenuPlan.query.filter_by(service_date=order.service_date).all():
        plan.status = "confirmed"
    db.session.commit()
    flash("採購單已確認並保存歷史 snapshot。", "success")
    return redirect(url_for("order_tool.purchase_detail", order_id=order_id))


@order_bp.post("/purchases/<int:order_id>/reopen")
def purchase_reopen(order_id: int):
    order = db.session.get(KitchenPurchaseOrder, order_id)
    if not order:
        abort(404)
    order.status = "draft"
    db.session.commit()
    flash("採購單已重開草稿；既有 snapshot 與人工調整仍保留。", "warning")
    return redirect(url_for("order_tool.purchase_detail", order_id=order_id))


@order_bp.post("/purchases/<int:order_id>/cancel")
def purchase_cancel(order_id: int):
    order = db.session.get(KitchenPurchaseOrder, order_id)
    if not order:
        abort(404)
    order.status = "cancelled"
    db.session.commit()
    flash("採購單已取消，歷史資料仍保留。", "warning")
    return redirect(url_for("order_tool.purchase_detail", order_id=order_id))
