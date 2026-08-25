"""學校食材登錄 Excel 匯出。

不新增資料表／欄位：直接使用既有學校、菜單、配方與人數。
製造商依使用者提供的 schoolingredient Excel 模板各食材帶入；供應商名稱沿用模板固定值。
"""

from __future__ import annotations

from copy import copy
from datetime import date
from decimal import Decimal
from io import BytesIO
from pathlib import Path

from flask import Blueprint, current_app, request, send_file
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, PatternFill
from sqlalchemy.orm import selectinload

from models import (
    KitchenMenuAssignment,
    KitchenMenuPlan,
    KitchenMenuPlanItem,
    KitchenRecipe,
    KitchenRecipeIngredient,
)

school_ingredient_export_bp = Blueprint("school_ingredient_export", __name__)

TEMPLATE_PATH = Path("static/schoolingredient_template.xlsx")
EXPECTED_HEADERS = (
    "供餐日期", "學校", "菜色名稱", "食材名稱", "進貨日期", "生產日期", "有效日期", "批號",
    "製造商", "供應商名稱", "食材驗證標章", "驗證號碼", "產品名稱", "重量(公斤)",
    "非基改玉米", "非基改黃豆", "加工品", "食材原產地",
)

# 模板讀取失敗時的備援值；內容來自使用者提供的 schoolingredient_20260601.xlsx。
FALLBACK_MANUFACTURER = {
    "粳米": "陸穀實業股份有限公司",
    "帶皮雞胸肉": "保證責任台灣省北台肉雞運銷合作社",
    "洋蔥": "徐匯恩",
    "豆腐": "津悅",
    "絞肉": "大湖畜牧場",
    "敏豆": "豐誠冷凍食品有限公司",
    "甜不辣": "品豐國際企業有限公司",
    "蚵白菜": "盛綻農業農產品初級加工場",
    "玉米粒": "富士鮮品股份有限公司-二廠",
    "雞蛋(白殼)": "炎稜畜牧場",
    "荷葉白菜": "羅勻晨",
    "素排": "佛心素食材料行",
    "三色丁": "富士鮮品股份有限公司",
    "香菇": "彰化縣菇類生產合作社",
    "杏鮑菇": "彰化縣菇類生產合作社",
    "甜椒": "劉明仁",
}
FALLBACK_CERTIFICATION = {
    "粳米": ("產銷履歷", "2605190098801217"),
    "帶皮雞胸肉": ("CAS台灣優良農產品", "016683"),
    "洋蔥": ("生產追溯-農產品", "1201004693"),
    "豆腐": ("", ""),
    "絞肉": ("生產追溯-豬肉", "LE300431"),
    "敏豆": ("產銷履歷", "2604155169113265"),
    "甜不辣": ("生產追溯-水產品", "0316600001"),
    "蚵白菜": ("產銷履歷", "00993663601066"),
    "玉米粒": ("CAS台灣優良農產品", "123701"),
    "雞蛋(白殼)": ("雞蛋噴印-洗選鮮蛋", "D43003260530C"),
    "荷葉白菜": ("台灣有機農產品", "1-010-100311"),
    "素排": ("", ""),
    "三色丁": ("CAS台灣優良農產品", "123706"),
    "香菇": ("生產追溯-農產品", "1004000002"),
    "杏鮑菇": ("生產追溯-農產品", "1004000002"),
    "甜椒": ("生產追溯-農產品", "1101003260"),
}
FALLBACK_FIXED = {
    "supplier": "廣豐食品有限公司",
    "corn": "Y",
    "soy": "Y",
    "processed": "N",
    "origin": "臺灣",
}


def _parse_date(raw: str | None) -> date:
    try:
        return date.fromisoformat(raw or "")
    except ValueError:
        return date.today()


def _ingredient_weight_kg(component: KitchenRecipeIngredient, headcount: int) -> Decimal | None:
    ingredient = component.ingredient
    per_person = component.grams_per_person or Decimal("0")
    base_amount = per_person * Decimal(max(headcount, 0))
    base_unit = (ingredient.base_unit or "g").strip().lower()
    purchase_unit = (ingredient.purchase_unit or "").strip().lower()

    if base_unit in {"g", "公克", "克"}:
        return base_amount / Decimal("1000")
    if purchase_unit in {"kg", "公斤"}:
        units_per_kg = ingredient.grams_per_purchase_unit or Decimal("0")
        if units_per_kg > 0:
            return base_amount / units_per_kg
    return None


def _template_values(sheet):
    manufacturers = dict(FALLBACK_MANUFACTURER)
    certification = dict(FALLBACK_CERTIFICATION)
    fixed = dict(FALLBACK_FIXED)

    for row_number in range(2, sheet.max_row + 1):
        ingredient_name = str(sheet.cell(row_number, 4).value or "").strip()
        if ingredient_name:
            manufacturer = sheet.cell(row_number, 9).value
            mark = sheet.cell(row_number, 11).value
            number = sheet.cell(row_number, 12).value
            if manufacturer is not None and str(manufacturer).strip():
                manufacturers[ingredient_name] = str(manufacturer).strip()
            certification[ingredient_name] = (
                "" if mark is None else str(mark).strip(),
                "" if number is None else str(number).strip(),
            )

        if row_number == 2:
            fixed = {
                "supplier": str(sheet.cell(row_number, 10).value or FALLBACK_FIXED["supplier"]).strip(),
                "corn": sheet.cell(row_number, 15).value or "Y",
                "soy": sheet.cell(row_number, 16).value or "Y",
                "processed": sheet.cell(row_number, 17).value or "N",
                "origin": sheet.cell(row_number, 18).value or "臺灣",
            }

    return manufacturers, certification, fixed


def _rows_for_date(service_date: date, manufacturers, certification, fixed):
    assignments = (
        KitchenMenuAssignment.query.join(KitchenMenuPlan)
        .filter(
            KitchenMenuPlan.service_date == service_date,
            KitchenMenuAssignment.service_status == "serving",
            KitchenMenuAssignment.headcount > 0,
        )
        .options(
            selectinload(KitchenMenuAssignment.school),
            selectinload(KitchenMenuAssignment.plan)
            .selectinload(KitchenMenuPlan.items)
            .selectinload(KitchenMenuPlanItem.recipe)
            .selectinload(KitchenRecipe.ingredients)
            .selectinload(KitchenRecipeIngredient.ingredient),
        )
        .all()
    )
    assignments.sort(key=lambda row: (row.school.name.casefold(), row.plan.meal_type or "", row.plan.name.casefold()))

    rows = []
    unresolved = set()
    for assignment in assignments:
        for menu_item in sorted(assignment.plan.items, key=lambda item: item.sort_order):
            recipe = menu_item.recipe
            for component in recipe.ingredients:
                if (component.grams_per_person or Decimal("0")) <= 0:
                    continue
                ingredient = component.ingredient
                weight_kg = _ingredient_weight_kg(component, assignment.headcount)
                if weight_kg is None:
                    unresolved.add(ingredient.name)
                    continue

                ingredient_key = ingredient.name.strip()
                manufacturer_name = manufacturers.get(ingredient_key, "")
                mark, verification_number = certification.get(ingredient_key, ("", ""))
                headcount = max(assignment.headcount, 0)
                per_person_kg = weight_kg / Decimal(headcount) if headcount else Decimal("0")
                rows.append({
                    "values": (
                        service_date,
                        assignment.school.name,
                        recipe.name,
                        ingredient.name,
                        service_date,
                        None,
                        None,
                        None,
                        manufacturer_name,
                        fixed["supplier"],
                        mark,
                        verification_number,
                        None,
                        weight_kg,
                        fixed["corn"],
                        fixed["soy"],
                        fixed["processed"],
                        fixed["origin"],
                    ),
                    "per_person_kg": per_person_kg,
                    "headcount": headcount,
                })
    return rows, sorted(unresolved)


def _new_template_workbook():
    """依原 Excel 欄位結構建立備援模板；F/G/M 隱藏，S/T/U 保留原本計算輔助欄。"""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    for column, header in enumerate(EXPECTED_HEADERS, start=1):
        sheet.cell(1, column).value = header
    widths = {
        "A": 10.45, "B": 19.63, "C": 13.36, "D": 9, "E": 11.09, "F": 6.63,
        "G": 9, "H": 13, "I": 15.09, "J": 9.36, "K": 19, "L": 21.36,
        "M": 0.09, "N": 11.09, "O": 9, "P": 13, "Q": 13, "R": 13,
        "S": 9, "T": 9, "U": 13,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    sheet.column_dimensions["F"].hidden = True
    sheet.column_dimensions["G"].hidden = True
    sheet.column_dimensions["M"].hidden = True
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in sheet[1][:18]:
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    return workbook


def _load_template():
    template_path = Path(current_app.root_path) / TEMPLATE_PATH
    if template_path.is_file():
        try:
            workbook = load_workbook(template_path)
            sheet = workbook.active
            headers = tuple(sheet.cell(1, column).value for column in range(1, 19))
            if headers == EXPECTED_HEADERS:
                return workbook
            current_app.logger.warning("schoolingredient template header mismatch; using fallback template")
        except Exception:
            current_app.logger.exception("schoolingredient template could not be loaded; using fallback template")
    else:
        current_app.logger.warning("schoolingredient template missing; using fallback template")
    return _new_template_workbook()


def _copy_template_row(sheet, source_row: int, target_row: int):
    for column in range(1, 22):
        source = sheet.cell(source_row, column)
        target = sheet.cell(target_row, column)
        target._style = copy(source._style)
        target.font = copy(source.font)
        target.fill = copy(source.fill)
        target.border = copy(source.border)
        target.alignment = copy(source.alignment)
        target.number_format = source.number_format
        target.protection = copy(source.protection)
    sheet.row_dimensions[target_row].height = sheet.row_dimensions[source_row].height


def _build_workbook(service_date: date):
    workbook = _load_template()
    sheet = workbook.active
    manufacturers, certification, fixed = _template_values(sheet)
    rows, unresolved = _rows_for_date(service_date, manufacturers, certification, fixed)
    if unresolved:
        names = "、".join(unresolved)
        return None, (
            f"以下食材目前沒有可靠的公斤換算，為避免匯出錯誤重量已停止：{names}。"
            "請先確認其基本單位／採購換算。"
        )

    first_data_row = 2
    template_last_row = max(sheet.max_row, first_data_row)
    required_last_row = max(template_last_row, first_data_row + len(rows) - 1)
    for row_number in range(template_last_row + 1, required_last_row + 1):
        _copy_template_row(sheet, first_data_row, row_number)

    for row_number in range(first_data_row, required_last_row + 1):
        for column in range(1, 22):
            sheet.cell(row_number, column).value = None

    for row_number, row in enumerate(rows, start=first_data_row):
        for column, value in enumerate(row["values"], start=1):
            sheet.cell(row_number, column).value = value
        sheet.cell(row_number, 19).value = row["per_person_kg"]
        sheet.cell(row_number, 20).value = row["headcount"]
        sheet.cell(row_number, 21).value = f"=S{row_number}*T{row_number}"

    return workbook, None


@school_ingredient_export_bp.get("/summary/school-ingredient.xlsx")
def school_ingredient_export():
    service_date = _parse_date(request.args.get("date"))
    workbook, error = _build_workbook(service_date)
    if error:
        return error, 400

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name=f"學校食材登錄-{service_date.isoformat()}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
