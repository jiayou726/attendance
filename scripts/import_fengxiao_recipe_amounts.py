"""Conservatively import explicit per-person grams from Fengxiao workbooks.

Only recipe and ingredient names that already exist in the Kitchen database are
eligible.  Existing positive quantities always win, conflicting source values
are skipped, and missing recipes/ingredients are reported without being created.
"""

from __future__ import annotations

import argparse
import json
import re
import sqlite3
from collections import defaultdict
from pathlib import Path
from statistics import median

from openpyxl import load_workbook


NON_DISH_LABELS = {"熱量"}
SOURCE_NOTE = "鳳小週食譜 1人g 匯入"


def _text(value) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def _normalized(value) -> str:
    value = re.sub(
        r"CAS|非基因改造|非基改|有機|產銷履歷|生產追溯|洗選|上等|大成|卜蜂|洽富",
        "",
        _text(value),
        flags=re.I,
    )
    value = re.sub(r"\d+(?:\.\d+)?(?:\*\d+(?:\.\d+)?)?|[\s\-_/.,，、（）()]", "", value)
    aliases = {
        "油丁": "油豆腐",
        "三角油": "油豆腐",
        "四角油": "油豆腐",
        "雞胸丁": "雞丁",
        "洗選蛋": "雞蛋",
        "蛋": "雞蛋",
        "青油菜": "青菜",
        "小松菜": "青菜",
        "高麗": "高麗菜",
        "洋菇罐頭": "洋菇",
        "玉米粒": "玉米",
    }
    for source, target in aliases.items():
        if source in value:
            return target
    return value


def _clean_source_ingredient(value) -> str:
    raw = _text(value)
    raw = re.sub(r"[（(].*?[）)]", "", raw)
    raw = re.sub(r"(?:CAS|冷凍|生鮮|調理|IQF)", "", raw, flags=re.I)
    raw = re.sub(
        r"[-_/](?:卜蜂|洽富|大成|上等|CAS|東豪|富士鮮品|復進|勝大|豐誠).*$",
        "",
        raw,
        flags=re.I,
    )
    raw = re.sub(r"(?<=[\u4e00-\u9fff])Q$", "", raw, flags=re.I)
    raw = re.sub(
        r"\d+(?:\.\d+)?\s*(?:kg|k|公斤|g|克)(?:\s*/\s*(?:板|包|袋|箱))?",
        "",
        raw,
        flags=re.I,
    )
    aliases = {
        "馬鈴薯": "洋芋",
        "牛番茄": "番茄",
        "去皮洋蔥": "洋蔥",
        "生香菇": "香菇",
        "琇珍菇": "秀珍菇",
    }
    return aliases.get(raw, raw).strip(" .-_、")


def _direct_grams(value):
    if isinstance(value, (int, float)) and not isinstance(value, bool) and value > 0:
        return float(value)
    match = re.fullmatch(r"\s*(\d+(?:\.\d+)?)\s*(?:g|克)\s*", _text(value), re.I)
    return float(match.group(1)) if match else None


def _fallback_amount(ingredient_name: str, recipe_name: str, category: str | None) -> float:
    """Return a deliberately ordinary estimate when no comparable sample exists."""
    ingredient = ingredient_name
    dish = f"{recipe_name}{category or ''}"
    rules = (
        (("鹽", "胡椒", "香料", "咖哩粉", "五香", "味精"), 1.0),
        (("糖", "蒜", "薑", "蔥", "九層塔", "香菜", "芹菜"), 2.0),
        (("豆腐", "豆干", "豆皮", "麵輪", "黑輪", "甜不辣", "米血", "素排"), 25.0),
        (("油", "醬", "醋", "酒", "沙茶", "番茄醬"), 3.0),
        (("紫菜", "海帶芽", "蝦皮", "柴魚"), 1.5),
        (("蛋",), 20.0),
        (("米", "麵", "冬粉", "米粉", "粿", "年糕", "麥", "薏仁", "紅豆", "綠豆"), 45.0),
        (("肉", "豬", "雞", "魚", "排", "腿", "翅", "丸", "羹", "蝦", "蛤", "鴨"), 35.0),
    )
    for keywords, amount in rules:
        if any(keyword in ingredient for keyword in keywords):
            return amount
    if "湯" in dish:
        return 12.0
    if any(word in dish for word in ("飯", "麵", "粥")):
        return 35.0
    if any(word in dish for word in ("主菜", "主食")):
        return 35.0
    return 25.0


def read_source(root: Path):
    """Return direct gram observations and valid dish names from all workbooks."""
    observations = []
    valid_dishes = set()
    files = sorted(root.glob("*.xlsx"))
    for path in files:
        workbook = load_workbook(path, read_only=True, data_only=True)
        for sheet in workbook.worksheets:
            rows = list(sheet.iter_rows(values_only=True))
            width = max((len(row) for row in rows), default=0)
            for base in (0, 6, 12, 18, 24):
                if len(rows) < 5 or base + 4 >= width:
                    continue
                current_dish = ""
                for row_number, row in enumerate(rows[4:], start=5):
                    padded = list(row) + [None] * max(0, base + 5 - len(row))
                    candidate = _text(padded[base])
                    if candidate and not re.search(r"假|連假|停餐", candidate):
                        current_dish = candidate
                    ingredient = _text(padded[base + 2])
                    if (
                        not current_dish
                        or current_dish in NON_DISH_LABELS
                        or not ingredient
                        or any(mark in ingredient for mark in ("、", "，", ",", "算人數"))
                    ):
                        continue
                    valid_dishes.add(current_dish)
                    amount = _direct_grams(padded[base + 3])
                    if amount is None:
                        continue
                    observations.append(
                        {
                            "file": path.name,
                            "sheet": sheet.title,
                            "row": row_number,
                            "dish": current_dish,
                            "ingredient": ingredient,
                            "amount": amount,
                        }
                    )
        workbook.close()
    return files, observations, valid_dishes


def import_amounts(
    database: Path,
    source_root: Path,
    *,
    apply: bool = False,
    fill_estimates: bool = False,
):
    files, observations, valid_dishes = read_source(source_root)
    connection = sqlite3.connect(database)
    connection.row_factory = sqlite3.Row
    try:
        recipes = {
            row["name"]: dict(row)
            for row in connection.execute("select id, name, category from kitchen_recipe")
        }
        recipes_by_id = {row["id"]: row for row in recipes.values()}
        ingredients = [
            dict(row)
            for row in connection.execute("select id, name, base_unit from kitchen_ingredient")
        ]
        ingredient_by_exact = {row["name"]: row for row in ingredients}
        ingredient_by_normalized = defaultdict(list)
        for ingredient in ingredients:
            ingredient_by_normalized[_normalized(ingredient["name"])].append(ingredient)

        grouped = defaultdict(list)
        source_samples_by_ingredient = defaultdict(list)
        unmatched_ingredient_observations = 0
        for row in observations:
            cleaned = _clean_source_ingredient(row["ingredient"])
            exact = ingredient_by_exact.get(cleaned)
            matches = [exact] if exact else ingredient_by_normalized.get(_normalized(cleaned), [])
            if len(matches) != 1 or matches[0]["base_unit"] != "g":
                unmatched_ingredient_observations += 1
                continue
            source_samples_by_ingredient[matches[0]["id"]].append(row["amount"])
            recipe = recipes.get(row["dish"])
            if recipe is None:
                continue
            grouped[(recipe["id"], matches[0]["id"])].append(row)

        existing = {
            (row["recipe_id"], row["ingredient_id"]): dict(row)
            for row in connection.execute(
                """select id, recipe_id, ingredient_id, grams_per_person,
                          quantity_status, source_note
                     from kitchen_recipe_ingredient"""
            )
        }

        updated_pending = 0
        added_components = 0
        preserved_existing = 0
        conflicting_pairs = 0
        stable_pairs = 0
        direct_keys = set()
        for key, rows in grouped.items():
            amounts = {round(row["amount"], 6) for row in rows}
            if len(amounts) != 1:
                conflicting_pairs += 1
                continue
            stable_pairs += 1
            amount = amounts.pop()
            component = existing.get(key)
            if component and float(component["grams_per_person"] or 0) > 0:
                preserved_existing += 1
                continue
            direct_keys.add(key)
            if component:
                updated_pending += 1
                if apply:
                    connection.execute(
                        """update kitchen_recipe_ingredient
                              set grams_per_person = ?, quantity_status = 'manual', source_note = ?
                            where id = ?""",
                        (amount, SOURCE_NOTE, component["id"]),
                    )
            else:
                added_components += 1
                if apply:
                    connection.execute(
                        """insert into kitchen_recipe_ingredient
                           (recipe_id, ingredient_id, grams_per_person, quantity_status, source_note)
                           values (?, ?, ?, 'manual', ?)""",
                        (key[0], key[1], amount, SOURCE_NOTE),
                    )

        estimated_components = 0
        estimate_methods = defaultdict(int)
        if fill_estimates:
            database_samples_by_ingredient = defaultdict(list)
            for component in existing.values():
                amount = float(component["grams_per_person"] or 0)
                if amount > 0:
                    database_samples_by_ingredient[component["ingredient_id"]].append(amount)

            ingredient_by_id = {row["id"]: row for row in ingredients}
            for key, component in existing.items():
                if key in direct_keys or float(component["grams_per_person"] or 0) > 0:
                    continue
                recipe = recipes_by_id[component["recipe_id"]]
                ingredient = ingredient_by_id[component["ingredient_id"]]
                source_samples = source_samples_by_ingredient.get(component["ingredient_id"], [])
                database_samples = database_samples_by_ingredient.get(component["ingredient_id"], [])
                if source_samples:
                    amount = float(median(source_samples))
                    method = "同食材鳳小菜單中位數"
                elif database_samples:
                    amount = float(median(database_samples))
                    method = "同食材既有配方中位數"
                else:
                    amount = _fallback_amount(
                        ingredient["name"], recipe["name"], recipe.get("category")
                    )
                    method = "依食材與菜色類型估算"
                estimated_components += 1
                estimate_methods[method] += 1
                if apply:
                    connection.execute(
                        """update kitchen_recipe_ingredient
                              set grams_per_person = ?, quantity_status = 'estimated', source_note = ?
                            where id = ?""",
                        (round(amount, 3), method, component["id"]),
                    )

        if apply:
            connection.commit()
        else:
            connection.rollback()

        missing_dishes = sorted(valid_dishes - set(recipes))
        return {
            "mode": "apply" if apply else "dry-run",
            "files": len(files),
            "source_dishes": len(valid_dishes),
            "missing_dishes": len(missing_dishes),
            "missing_dish_names": missing_dishes,
            "direct_gram_observations": len(observations),
            "stable_pairs": stable_pairs,
            "conflicting_pairs_skipped": conflicting_pairs,
            "unmatched_ingredient_observations": unmatched_ingredient_observations,
            "updated_pending": updated_pending,
            "added_components": added_components,
            "preserved_existing": preserved_existing,
            "estimated_components": estimated_components,
            "estimate_methods": dict(estimate_methods),
            "total_changes": updated_pending + added_components + estimated_components,
        }
    finally:
        connection.close()


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--database", type=Path, default=Path("attendance.db"))
    parser.add_argument("--root", type=Path, required=True)
    parser.add_argument("--apply", action="store_true", help="commit safe changes")
    parser.add_argument(
        "--fill-estimates",
        action="store_true",
        help="fill every remaining zero component with a clearly marked estimate",
    )
    args = parser.parse_args()
    result = import_amounts(
        args.database,
        args.root,
        apply=args.apply,
        fill_estimates=args.fill_estimates,
    )
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
