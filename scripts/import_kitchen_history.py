"""Import the user's historical kitchen files into kitchen_* tables only.

The importer is intentionally conservative:
- every menu dish and listed material is imported;
- grams/person is filled only when one manufacturing row can be matched to
  exactly one dish on that service date;
- all other material quantities remain 0 and are marked ``pending`` so they
  cannot silently inflate or corrupt purchase calculations.

Run against a test/staging DATABASE_URL first. Existing attendance tables are
never queried or mutated by this script.
"""

from __future__ import annotations

import argparse
import re
import statistics
import sys
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook
import xlrd

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from app import create_app
from extensions import db
from models import (
    KitchenIngredient,
    KitchenRecipe,
    KitchenRecipeIngredient,
    KitchenSchool,
    KitchenSupplier,
    KitchenSupplierItem,
)


PHOTO_SCHOOLS = {
    "信義": 600,
    "新屋": 600,
    "平鎮": 500,
    "新明": 300,
    "中平": 550,
}

NON_FOOD_WORDS = (
    "工程", "機電", "營造", "鋼鋁", "不銹鋼", "橡膠", "文具", "冷凍空調", "鋁門窗", "瓦斯",
    "煤氣", "體育", "室內設計", "汽車", "金儀", "檢驗", "水電", "制服", "包裝", "輸送帶",
    "資訊", "化學", "印刷", "五金", "家具", "電器", "氣體", "病媒", "快篩", "冷氣", "電話",
    "保溫車", "土木", "建材", "環境維護", "度量衡", "酒精", "年糕紙", "餐盒", "帳單空白",
)
FOOD_WORDS = (
    "食品", "食材", "肉", "豬", "雞", "蛋", "豆", "蔬", "菜", "果", "農產", "農會", "漁會",
    "米", "麵", "麥", "包子", "餐包", "麵包", "蛋糕", "乳", "飲料", "油脂", "醬油", "調味",
    "素食", "素菜", "水果", "青菓", "菓菜", "洋芋", "洋蔥", "紅蘿蔔", "菇", "蒜頭", "紅豆",
    "蛤", "酸菜", "豬血", "海帶", "木耳", "蝦", "魚", "年糕", "優酪乳", "保久乳",
)

UNIT_ALIASES = {
    "kg": "kg", "kgs": "kg", "公斤": "kg", "千克": "kg",
    "斤": "斤", "台斤": "斤",
    "箱": "箱", "包": "包", "袋": "袋", "盒": "盒", "桶": "桶", "瓶": "瓶", "罐": "罐",
    "個": "個", "顆": "個", "粒": "個", "支": "支", "隻": "隻", "片": "片", "條": "條",
    "板": "板", "籃": "籃", "簍": "簍", "捲": "捲", "件": "件", "組": "組", "打": "打",
}

PROCESS_WORDS = {"煮", "炒", "燒", "炸", "蒸", "燙", "滷", "烤", "拌", "切", "絲", "丁"}
CERT_RE = re.compile(r"[（(][^）)]{0,12}[）)]")
LEADING_NO_RE = re.compile(r"^\d+[.、_-]*")


def text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def clean_supplier_name(filename: str) -> str:
    name = Path(filename).stem
    name = LEADING_NO_RE.sub("", name)
    name = re.sub(r"[（(].*?[）)]", "", name)
    name = re.sub(r"(?:有限公司|股份有限公司)$", "", name).strip(" .-_、")
    return name[:100]


def supplier_is_food(filename: str) -> bool:
    name = Path(filename).stem
    if name.startswith("00") or "空白" in name:
        return False
    if any(word in name for word in NON_FOOD_WORDS):
        return False
    return any(word in name for word in FOOD_WORDS)


def purchase_unit(value):
    raw = text(value).lower().replace(".", "").replace(" ", "")
    return UNIT_ALIASES.get(raw)


def _short_number(value: str) -> str:
    number = float(value)
    return str(int(number)) if number.is_integer() else str(number).rstrip("0").rstrip(".")


def extract_package_conversion(item_name: str, purchase_unit_name: str | None = None) -> str:
    """Extract a readable pack conversion from historical product-name notes."""
    raw = text(item_name).replace("／", "/").replace("×", "*").replace("X", "*")
    outer_default = purchase_unit_name if purchase_unit_name in UNIT_ALIASES.values() else None
    unit_alias = {"公斤": "kg", "KG": "kg", "Kg": "kg", "G": "g", "克": "g", "顆": "個"}

    approximate = re.search(
        r"(\d+(?:\.\d+)?)\s*(包|袋|盒|瓶|罐|桶)\s*(?:約|大約)\s*(\d+(?:\.\d+)?)\s*(粒|個|顆|片|支)",
        raw, re.I,
    )
    if approximate:
        outer_count, outer, inner_count, inner = approximate.groups()
        if float(outer_count) == 1:
            return f"1{outer}≈{_short_number(inner_count)}{unit_alias.get(inner, inner)}"

    multiplied = re.search(
        r"(\d+(?:\.\d+)?)\s*(kg|公斤|g|克|斤)\s*\*\s*(\d+(?:\.\d+)?)\s*(包|袋|盒|瓶|罐)",
        raw, re.I,
    )
    if multiplied and outer_default:
        weight, weight_unit, count, inner = multiplied.groups()
        weight_unit = unit_alias.get(weight_unit, weight_unit.lower())
        return f"1{outer_default}＝{_short_number(count)}{inner}（每{inner}{_short_number(weight)}{weight_unit}）"

    explicit = re.search(
        r"(\d+(?:\.\d+)?)\s*(kg|公斤|g|克|斤|包|袋|盒|瓶|罐|粒|個|顆|片|支|入)\s*/\s*(箱|件|包|袋|盒|桶|瓶|罐)",
        raw, re.I,
    )
    if explicit:
        count, inner, outer = explicit.groups()
        return f"1{outer}＝{_short_number(count)}{unit_alias.get(inner, inner.lower() if inner.lower() in {'kg', 'g'} else inner)}"

    reversed_weight = re.search(r"(包|袋|盒|桶)\s*/\s*(\d+(?:\.\d+)?)\s*(kg|公斤|g|克|斤)", raw, re.I)
    if reversed_weight:
        outer, count, inner = reversed_weight.groups()
        return f"1{outer}＝{_short_number(count)}{unit_alias.get(inner, inner.lower())}"

    multiplied_count = re.search(r"\*\s*(\d+(?:\.\d+)?)\s*(包|袋|盒|瓶|罐|粒|個|顆|片|支)", raw, re.I)
    if multiplied_count and outer_default:
        count, inner = multiplied_count.groups()
        return f"1{outer_default}＝{_short_number(count)}{unit_alias.get(inner, inner)}"

    standalone_weight = re.search(r"(\d+(?:\.\d+)?)\s*(kg|公斤|g|克|斤)\b", raw, re.I)
    if standalone_weight and outer_default not in {"kg", "斤"}:
        count, inner = standalone_weight.groups()
        return f"1{outer_default}＝{_short_number(count)}{unit_alias.get(inner, inner.lower())}"

    count_in = re.search(r"(\d+(?:\.\d+)?)\s*入", raw)
    if count_in and outer_default:
        return f"1{outer_default}＝{_short_number(count_in.group(1))}入"
    return ""


def _workbook_rows(path: Path):
    """Yield worksheet rows from either historical BIFF .xls or modern .xlsx."""
    if path.suffix.lower() == ".xls":
        book = xlrd.open_workbook(path, on_demand=True)
        for sheet in book.sheets():
            for row_index in range(sheet.nrows):
                values = []
                for cell in sheet.row(row_index):
                    value = cell.value
                    if cell.ctype == xlrd.XL_CELL_DATE:
                        try:
                            value = xlrd.xldate_as_datetime(value, book.datemode)
                        except (ValueError, OverflowError):
                            pass
                    values.append(value)
                yield values
        book.release_resources()
        return
    workbook = load_workbook(path, read_only=True, data_only=True)
    for sheet in workbook.worksheets:
        yield from sheet.iter_rows(values_only=True)


def _row_date(value, current_year):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    raw = text(value)
    match = re.fullmatch(r"(\d{1,2})[/-](\d{1,2})", raw)
    if match and current_year:
        try:
            return date(current_year, int(match.group(1)), int(match.group(2)))
        except ValueError:
            return None
    return None


def read_supplier_orders(path: Path):
    """Extract item/quantity/unit/price rows without trusting workbook formulas."""
    orders = []
    current_date = None
    current_year = None
    sequence = 0
    try:
        rows = _workbook_rows(path)
        for raw_row in rows:
            sequence += 1
            values = [value for value in raw_row if value not in (None, "")]
            if not values:
                continue
            for value in values:
                year_match = re.fullmatch(r"\s*(\d{2,4})年\s*", text(value))
                if year_match:
                    year = int(year_match.group(1))
                    current_year = year + 1911 if year < 1911 else year
            found_date = _row_date(values[0], current_year)
            if found_date:
                current_date = found_date
                current_year = found_date.year

            unit_index = None
            unit = None
            for index, value in enumerate(values):
                candidate = purchase_unit(value)
                if candidate:
                    unit_index, unit = index, candidate
                    break
            if unit_index is None:
                continue
            quantity_index = next(
                (index for index in range(unit_index - 1, -1, -1) if isinstance(values[index], (int, float)) and not isinstance(values[index], bool)),
                None,
            )
            if quantity_index is None:
                continue
            item = None
            for index in range(quantity_index - 1, -1, -1):
                candidate = text(values[index])
                if not candidate or _row_date(values[index], current_year):
                    continue
                if re.fullmatch(r"\d{2,4}年", candidate) or candidate in {"品名", "食材", "日期", "數量", "單位", "單價"}:
                    continue
                item = candidate
                break
            if not item or len(item) > 120 or item in {"合計", "總計", "小計"}:
                continue
            price = next(
                (value for value in values[unit_index + 1:] if isinstance(value, (int, float)) and not isinstance(value, bool)),
                None,
            )
            quantity = values[quantity_index]
            if quantity < 0 or (price is not None and price < 0):
                continue
            orders.append({
                "name": re.sub(r"\s+", " ", item).strip(),
                "unit": unit,
                "package_conversion": extract_package_conversion(item, unit),
                "quantity": float(quantity),
                "unit_price": float(price) if price is not None else None,
                "purchase_date": current_date,
                "sequence": sequence,
                "source_file": path.name,
                "file_mtime": path.stat().st_mtime,
            })
    except Exception as exc:
        print(f"supplier_parse_error={path.name}: {exc}")
    return orders


def normalized(value: str) -> str:
    value = CERT_RE.sub("", text(value))
    value = re.sub(r"CAS|非基因改造|非基改|有機|產銷履歷|生產追溯|洗選|上等|大成|卜蜂|洽富", "", value, flags=re.I)
    value = re.sub(r"\d+(?:\.\d+)?(?:\*\d+(?:\.\d+)?)?|[\s\-_/.,，、（）()]", "", value)
    aliases = {
        "油丁": "油豆腐", "三角油": "油豆腐", "四角油": "油豆腐", "雞胸丁": "雞丁",
        "洗選蛋": "雞蛋", "蛋": "雞蛋", "青油菜": "青菜", "小松菜": "青菜",
        "高麗": "高麗菜", "洋菇罐頭": "洋菇", "玉米粒": "玉米",
    }
    for source, target in aliases.items():
        if source in value:
            return target
    return value


def clean_supplier_product_name(value: str) -> str:
    """Return a short operational name, dropping historical brand/package noise."""
    raw = text(value)
    raw = re.sub(r"[（(].*?[）)]", "", raw)
    raw = re.sub(r"(?:CAS|冷凍|生鮮|調理|IQF)", "", raw, flags=re.I)
    raw = re.sub(r"\b\d+(?:\.\d+)?\s*(?:kg|公斤|g|克|斤|兩)(?:\s*[x×*]\s*\d+)?\b", "", raw, flags=re.I)
    raw = re.sub(r"\d+\s*包\s*/\s*件", "", raw, flags=re.I)
    raw = re.sub(r"[-_/](?:卜蜂|洽富|大成|上等|CAS|東豪).*$", "", raw, flags=re.I)
    raw = re.sub(r"(?<=[\u4e00-\u9fff])[A-Z]$", "", raw, flags=re.I)
    compact = re.sub(r"\s+", "", raw)
    if (re.match(r"^棒棒腿", compact, flags=re.I)
            or re.fullmatch(r"TS\d+\+?", compact, flags=re.I)
            or re.fullmatch(r"棒\d+", compact)
            or re.fullmatch(r"\d+棒(?:進口)?", compact)):
        raw = "棒棒腿"
    elif re.match(r"^(?:雞)?排\d*$", compact):
        raw = "雞排"
    elif re.match(r"^骨腿(?:T?\d+)?$", compact, flags=re.I):
        raw = "骨腿"
    elif re.match(r"^檸檬(?:雞)?翅", compact):
        raw = "檸檬雞翅"
    elif re.match(r"^菲力(?:雞排)?\d*", compact):
        raw = "菲力雞排"
    elif re.match(r"^照燒(?:里肌)?\d*", compact):
        raw = "照燒里肌"
    elif re.match(r"^日式(?:豬排)?\d*", compact):
        raw = "日式豬排"
    elif re.match(r"^(?:棒丁|棒腿丁)$", compact):
        raw = "棒腿丁"
    aliases = {
        "馬鈴薯": "洋芋",
        "洋芋大丁": "芋大丁",
        "洋芋小丁": "芋小丁",
        "紅卜": "紅蘿蔔",
        "豬火": "豬火鍋片",
        "豬火片": "豬火鍋片",
        "豬肉片": "豬火鍋片",
        "牛火": "牛火鍋片",
        "牛火片": "牛火鍋片",
        "牛南丁": "牛腩丁",
    }
    compact = re.sub(r"\s+", "", raw)
    raw = aliases.get(compact, raw)
    raw = raw.strip(" .-_、")
    if len(raw) < 2 or not re.search(r"[\u4e00-\u9fffA-Za-z0-9]", raw) or "?" in raw:
        return ""
    return raw[:120]


def supplier_product_key(value: str) -> str:
    """Collapse historical brand/package spellings into one editable catalog item."""
    return normalized(clean_supplier_product_name(value))


def split_materials(raw: str) -> list[str]:
    raw = text(raw).split("/")[0]
    raw = CERT_RE.sub("", raw)
    parts = re.split(r"[.．、,，+＋]", raw)
    result = []
    for part in parts:
        item = part.strip(" /-_")
        if not item or item in PROCESS_WORDS or item.isdigit():
            continue
        item = re.sub(r"^(非基因改造|非基改)", "", item).strip()
        if item and item not in result:
            result.append(item[:100])
    return result


def infer_category(dish: str, material_line: str) -> str:
    if any(word in dish for word in ("飯", "麵", "粥")) and not any(word in dish for word in ("湯", "甜")):
        return "主食"
    if any(word in dish for word in ("湯", "羹", "露", "西米露")):
        return "湯品"
    if any(word in dish for word in ("蔬菜", "青菜")):
        return "青菜"
    if any(word in dish for word in ("雞", "豬", "肉", "魚", "排", "腿", "翅", "蛋")):
        return "主菜"
    if any(word in dish for word in ("包", "糕", "果", "奶", "布丁", "飲")):
        return "點心"
    return "副菜"


def read_menu_files(root: Path):
    descriptions: dict[str, Counter] = defaultdict(Counter)
    categories: dict[str, Counter] = defaultdict(Counter)
    date_dishes: dict = defaultdict(lambda: defaultdict(set))
    file_count = 0
    for path in root.rglob("*.xlsx"):
        if path.name.startswith("~$"):
            continue
        try:
            workbook = load_workbook(path, read_only=True, data_only=True)
        except Exception:
            continue
        file_count += 1
        for sheet in workbook.worksheets:
            rows = list(sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 180), min_col=1, max_col=min(sheet.max_column, 12), values_only=True))
            for index, row in enumerate(rows[:-1]):
                if not row or not isinstance(row[0], datetime):
                    continue
                following = rows[index + 1]
                service_date = row[0].date()
                for column in range(2, len(row)):
                    dish = text(row[column])
                    material_line = text(following[column] if column < len(following) else "")
                    if not dish or not material_line or len(dish) > 40 or dish.startswith("="):
                        continue
                    materials = split_materials(material_line)
                    if not materials:
                        continue
                    descriptions[dish][material_line] += 1
                    categories[dish][infer_category(dish, material_line)] += 1
                    date_dishes[service_date][dish].update(materials)
    return descriptions, categories, date_dishes, file_count


def read_manufacturing(root: Path, date_dishes):
    estimates = defaultdict(list)
    used_rows = 0
    ambiguous_rows = 0
    for path in sorted(root.glob("2026.*月製造資料.xlsx")):
        workbook = load_workbook(path, read_only=True, data_only=True)
        for sheet in workbook.worksheets:
            raw_date = sheet.cell(1, 3).value
            if not isinstance(raw_date, datetime):
                continue
            service_date = raw_date.date()
            total_people = sheet.cell(2, 12).value
            if not isinstance(total_people, (int, float)) or total_people <= 0:
                continue
            dishes = date_dishes.get(service_date, {})
            if not dishes:
                continue
            for row in sheet.iter_rows(min_row=6, max_row=min(sheet.max_row, 80), min_col=1, max_col=4, values_only=True):
                school_scope, item_code, material_name, kg = row
                if not material_name or not isinstance(kg, (int, float)) or kg <= 0:
                    continue
                material_norm = normalized(text(material_name))
                candidates = set()
                candidate_material = {}
                for dish, materials in dishes.items():
                    for material in materials:
                        token = normalized(material)
                        generic_veg = text(item_code).startswith("B") and any(v in token for v in ("青菜", "蔬菜"))
                        if generic_veg or (token and material_norm and (token in material_norm or material_norm in token)):
                            candidates.add(dish)
                            candidate_material[dish] = material
                if len(candidates) != 1:
                    ambiguous_rows += 1
                    continue
                dish = next(iter(candidates))
                grams_per_person = float(kg) * 1000 / float(total_people)
                if not 0.05 <= grams_per_person <= 500:
                    continue
                estimates[(dish, candidate_material[dish])].append(grams_per_person)
                used_rows += 1
    return estimates, used_rows, ambiguous_rows


def read_supplier_master(path: Path | None):
    rows = []
    if not path or not path.exists():
        return rows
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.worksheets[0]
    for row in sheet.iter_rows(min_row=2, values_only=True):
        name = text(row[0] if len(row) > 0 else "")
        if not name:
            continue
        rows.append({
            "name": name,
            "phone": text(row[1] if len(row) > 1 else ""),
            "mobile": text(row[2] if len(row) > 2 else ""),
            "fax": text(row[3] if len(row) > 3 else ""),
            "contact": text(row[4] if len(row) > 4 else ""),
            "address": text(row[5] if len(row) > 5 else ""),
        })
    return rows


def match_contact(name: str, contacts):
    key = normalized(name)
    matches = [row for row in contacts if normalized(row["name"]) in key or key in normalized(row["name"])]
    return max(matches, key=lambda row: len(normalized(row["name"])), default={})


def upsert_supplier_files(root: Path, contacts):
    imported = 0
    seen = set()
    for path in sorted(root.rglob("*")):
        if path.suffix.lower() not in {".xls", ".xlsx"} or not supplier_is_food(path.name):
            continue
        name = clean_supplier_name(path.name)
        if not name or name in seen:
            continue
        seen.add(name)
        row = KitchenSupplier.query.filter_by(name=name).one_or_none()
        if row is None:
            row = KitchenSupplier(name=name)
            db.session.add(row)
        contact = match_contact(name, contacts)
        for field in ("phone", "mobile", "fax", "contact", "address"):
            value = contact.get(field)
            if value and not getattr(row, field):
                setattr(row, field, value[:255] if field == "address" else value[:100])
        row.source_file = path.name[:255]
        row.note = row.note or "由歷史食品廠商 Excel 匯入"
        imported += 1
    return imported


def upsert_supplier_catalog(root: Path, minimum_orders: int, minimum_item_orders: int, max_items_per_supplier: int):
    grouped = defaultdict(lambda: defaultdict(lambda: {
        "names": Counter(), "units": Counter(), "orders": [], "sources": set(),
    }))
    files_read = 0
    rows_read = 0
    seen_orders = defaultdict(set)
    for path in sorted(root.rglob("*")):
        if path.suffix.lower() not in {".xls", ".xlsx"} or not supplier_is_food(path.name):
            continue
        supplier_name = clean_supplier_name(path.name)
        orders = read_supplier_orders(path)
        files_read += 1
        for order in orders:
            display_name = clean_supplier_product_name(order["name"])
            key = supplier_product_key(display_name)
            if not key:
                continue
            if order["purchase_date"]:
                order_key = (
                    order["purchase_date"].isoformat(), key, order["quantity"],
                    order["unit"], order["unit_price"],
                )
            else:
                order_key = (path.name, order["sequence"], key, order["quantity"], order["unit"], order["unit_price"])
            if order_key in seen_orders[supplier_name]:
                continue
            seen_orders[supplier_name].add(order_key)
            rows_read += 1
            bucket = grouped[supplier_name][key]
            bucket["names"][display_name] += 1
            bucket["units"][order["unit"]] += 1
            bucket["orders"].append(order)
            bucket["sources"].add(order["source_file"])

    ingredient_by_key = defaultdict(list)
    for ingredient in KitchenIngredient.query.all():
        ingredient_by_key[supplier_product_key(ingredient.name)].append(ingredient)

    kept_suppliers = removed_suppliers = item_count = 0
    for supplier in KitchenSupplier.query.filter(KitchenSupplier.source_file.isnot(None)).all():
        products = grouped.get(supplier.name, {})
        total_orders = sum(len(bucket["orders"]) for bucket in products.values())
        if total_orders < minimum_orders:
            for ingredient in KitchenIngredient.query.filter_by(supplier_id=supplier.id).all():
                ingredient.supplier_id = None
            db.session.delete(supplier)
            removed_suppliers += 1
            continue

        kept_suppliers += 1
        ranked_products = sorted(products.items(), key=lambda pair: (-len(pair[1]["orders"]), pair[0]))
        selected_products = [pair for pair in ranked_products if len(pair[1]["orders"]) >= minimum_item_orders]
        if not selected_products:
            selected_products = ranked_products[:3]
        selected_products = selected_products[:max_items_per_supplier]
        seen_item_ids = set()
        for key, bucket in selected_products:
            display_name = bucket["names"].most_common(1)[0][0]
            latest = max(
                bucket["orders"],
                key=lambda order: (
                    order["purchase_date"].toordinal() if order["purchase_date"] else 0,
                    order["file_mtime"],
                    order["sequence"],
                ),
            )
            package_orders = [order for order in bucket["orders"] if order.get("package_conversion")]
            latest_package = max(
                package_orders,
                key=lambda order: (
                    order["purchase_date"].toordinal() if order["purchase_date"] else 0,
                    order["file_mtime"],
                    order["sequence"],
                ),
            ) if package_orders else None
            unit = latest["unit"] or bucket["units"].most_common(1)[0][0]
            item = KitchenSupplierItem.query.filter_by(supplier_id=supplier.id, source_key=key).one_or_none()
            if item is None:
                item = KitchenSupplierItem.query.filter_by(supplier_id=supplier.id, name=display_name).one_or_none()
            if item is None:
                # Reuse an older spelling row when normalization finds one.
                item = next((row for row in supplier.items if supplier_product_key(row.name) == key), None)
            if item is None:
                item = KitchenSupplierItem(supplier_id=supplier.id, source_key=key, name=display_name, unit=unit)
                db.session.add(item)
                db.session.flush()
            item.source_key = key
            if not item.manual_override:
                item.name = display_name
                item.unit = unit
                item.package_conversion = latest_package["package_conversion"] if latest_package else None
                item.last_unit_price = latest["unit_price"]
            item.last_quantity = latest["quantity"]
            item.last_purchase_date = latest["purchase_date"]
            item.order_count = len(bucket["orders"])
            item.source_file = "、".join(sorted(bucket["sources"]))[:255]
            matches = ingredient_by_key.get(key, [])
            if len(matches) == 1:
                item.ingredient_id = matches[0].id
                if matches[0].supplier_id is None:
                    matches[0].supplier_id = supplier.id
            seen_item_ids.add(item.id)
            item_count += 1
        for old_item in list(supplier.items):
            if old_item.id not in seen_item_ids and not old_item.manual_override:
                old_item.active = False
    return {
        "files_read": files_read,
        "rows_read": rows_read,
        "kept_suppliers": kept_suppliers,
        "removed_suppliers": removed_suppliers,
        "supplier_items": item_count,
    }


def upsert_schools():
    for name, headcount in PHOTO_SCHOOLS.items():
        row = KitchenSchool.query.filter_by(name=name).one_or_none()
        if row is None:
            row = KitchenSchool(name=name)
            db.session.add(row)
        row.default_headcount = headcount


def upsert_recipes(descriptions, categories, estimates):
    recipe_count = ingredient_count = component_count = estimated_count = 0
    for dish, variants in sorted(descriptions.items()):
        primary_line = variants.most_common(1)[0][0]
        materials = split_materials(primary_line)
        recipe = KitchenRecipe.query.filter_by(name=dish).one_or_none()
        if recipe is None:
            recipe = KitchenRecipe(name=dish)
            db.session.add(recipe)
            db.session.flush()
            recipe_count += 1
        recipe.category = categories[dish].most_common(1)[0][0]
        recipe.note = f"115年菜單匯入；主要材料：{primary_line}"[:255]
        for material in materials:
            ingredient = KitchenIngredient.query.filter_by(name=material).one_or_none()
            if ingredient is None:
                ingredient = KitchenIngredient(
                    name=material,
                    base_unit="g",
                    purchase_unit="kg",
                    grams_per_purchase_unit=1000,
                    unit_price=0,
                    order_increment=0.001,
                    note="由115年菜單材料文字匯入",
                )
                db.session.add(ingredient)
                db.session.flush()
                ingredient_count += 1
            component = KitchenRecipeIngredient.query.filter_by(recipe_id=recipe.id, ingredient_id=ingredient.id).one_or_none()
            if component is None:
                component = KitchenRecipeIngredient(
                    recipe_id=recipe.id,
                    ingredient_id=ingredient.id,
                    grams_per_person=0,
                    quantity_status="pending",
                    source_note="115年菜單；克數待確認",
                )
                db.session.add(component)
                component_count += 1
            samples = estimates.get((dish, material), [])
            if samples and component.quantity_status != "manual":
                component.grams_per_person = round(statistics.median(samples), 3)
                component.quantity_status = "estimated"
                component.source_note = f"2026製造表反推中位數（{len(samples)}次）"
                estimated_count += 1
    return recipe_count, ingredient_count, component_count, estimated_count


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--menu-root", type=Path, required=True)
    parser.add_argument("--manufacturing-root", type=Path, required=True)
    parser.add_argument("--supplier-root", type=Path, required=True)
    parser.add_argument("--supplier-master-xlsx", type=Path)
    parser.add_argument("--minimum-orders", type=int, default=5, help="remove imported suppliers below this historical row count")
    parser.add_argument("--minimum-item-orders", type=int, default=5, help="keep regularly ordered supplier items")
    parser.add_argument("--max-items-per-supplier", type=int, default=50, help="cap automatic catalog size per supplier")
    parser.add_argument("--apply", action="store_true", help="commit kitchen_* changes")
    args = parser.parse_args()

    descriptions, categories, date_dishes, menu_files = read_menu_files(args.menu_root)
    estimates, used_rows, ambiguous_rows = read_manufacturing(args.manufacturing_root, date_dishes)
    contacts = read_supplier_master(args.supplier_master_xlsx)

    app = create_app({"AUTO_CREATE_DB": True})
    with app.app_context():
        upsert_schools()
        supplier_candidates = upsert_supplier_files(args.supplier_root, contacts)
        recipe_stats = upsert_recipes(descriptions, categories, estimates)
        db.session.flush()
        catalog_stats = upsert_supplier_catalog(
            args.supplier_root,
            max(args.minimum_orders, 0),
            max(args.minimum_item_orders, 1),
            max(args.max_items_per_supplier, 1),
        )
        if args.apply:
            db.session.commit()
        else:
            db.session.rollback()
        print(f"menu_files={menu_files}")
        print(f"unique_dishes={len(descriptions)}")
        print(f"food_supplier_candidates={supplier_candidates}")
        for key, value in catalog_stats.items():
            print(f"{key}={value}")
        print(f"manufacturing_rows_used={used_rows}")
        print(f"manufacturing_rows_ambiguous={ambiguous_rows}")
        print("new_recipes=%d new_ingredients=%d new_components=%d estimated_components=%d" % recipe_stats)
        print("APPLIED" if args.apply else "DRY RUN ONLY")


if __name__ == "__main__":
    main()
