from datetime import date, timedelta
from decimal import Decimal
from io import BytesIO

import pytest
from openpyxl import Workbook, load_workbook

import blueprints.order_tool as order_tool_module
from app import create_app
from extensions import db
from models import (
    Checkin,
    Employee,
    KitchenIngredient,
    KitchenMenuAssignment,
    KitchenMenuPlan,
    KitchenMenuPlanItem,
    KitchenPurchaseOrder,
    KitchenPurchaseOrderItem,
    KitchenRecipe,
    KitchenRecipeIngredient,
    KitchenSchool,
    KitchenSupplier,
    KitchenSupplierItem,
)

TEST_DAY = date(2026, 8, 13)


@pytest.fixture()
def app(tmp_path):
    db_path = tmp_path / "kitchen_test.db"
    app = create_app({
        "TESTING": True,
        "SECRET_KEY": "test-only-secret",
        "SQLALCHEMY_DATABASE_URI": f"sqlite:///{db_path}",
        "AUTO_CREATE_DB": True,
        "KITCHEN_CSRF_ENABLED": False,
        "ADMIN_HR_PASSWORD": "test-hr-password",
        "ADMIN_MGR_PASSWORD": "test-mgr-password",
        "PRODUCTION": False,
    })
    yield app
    with app.app_context():
        db.session.remove()
        db.drop_all()


@pytest.fixture()
def client(app):
    return app.test_client()


@pytest.fixture()
def authed_client(client):
    with client.session_transaction() as sess:
        sess["role"] = "mgr"
    return client


def _ids(app):
    with app.app_context():
        return {
            "school": KitchenSchool.query.filter_by(name="內小").one().id,
            "supplier": KitchenSupplier.query.filter_by(name="測試肉品").one().id,
            "ingredient": KitchenIngredient.query.filter_by(name="骨腿丁").one().id,
            "recipe": KitchenRecipe.query.filter_by(name="南洋綠咖哩雞").one().id,
        }


def _ids_partial(app):
    with app.app_context():
        school = KitchenSchool.query.filter_by(name="內小").one_or_none()
        supplier = KitchenSupplier.query.filter_by(name="測試肉品").one_or_none()
        return {
            "school": school.id if school else None,
            "supplier": supplier.id if supplier else None,
        }


def _seed_core_via_routes(app, client):
    assert client.post("/admin/order-tool/schools", data={"name": "內小", "code": "1-08"}).status_code == 302
    assert client.post("/admin/order-tool/suppliers", data={"name": "測試肉品", "phone": "02-12345678"}).status_code == 302
    ids = _ids_partial(app)

    assert client.post("/admin/order-tool/ingredients", data={
        "name": "骨腿丁",
        "supplier_id": str(ids["supplier"]),
        "purchase_unit": "kg",
        "grams_per_purchase_unit": "1000",
        "unit_price": "82",
        "order_increment": "0.001",
        "note": "",
    }).status_code == 302

    assert client.post("/admin/order-tool/recipes", data={
        "name": "南洋綠咖哩雞",
        "category": "主菜",
        "serving_output_g": "95",
        "note": "",
    }).status_code == 302

    ids = _ids(app)
    assert client.post(f"/admin/order-tool/recipes/{ids['recipe']}/ingredients", data={
        "ingredient_id": str(ids["ingredient"]),
        "grams_per_person": "88",
    }).status_code == 302
    return ids


def _create_plan(app, client, ids, service_date="2026-08-13", headcount=801):
    assert client.post("/admin/order-tool/plans", data={
        "service_date": service_date,
        "meal_type": "午餐",
        "name": "中央菜單",
        "note": "",
    }).status_code == 302
    service_day = date.fromisoformat(service_date)
    with app.app_context():
        plan = KitchenMenuPlan.query.filter_by(service_date=service_day, meal_type="午餐", name="中央菜單").one()
        plan_id = plan.id
    assert client.post(f"/admin/order-tool/plans/{plan_id}/items", data={"recipe_id": str(ids["recipe"])}).status_code == 302
    assert client.post(f"/admin/order-tool/plans/{plan_id}/assignments", data={
        "school_id": str(ids["school"]),
        "headcount": str(headcount),
    }).status_code == 302
    return plan_id


def test_kitchen_does_not_require_login_but_other_admin_pages_do(client):
    response = client.get("/admin/order-tool/", follow_redirects=False)
    assert response.status_code == 200
    page = response.get_data(as_text=True)
    assert "開啟菜單" in page
    assert all(label in page for label in ("總表", "菜色配方", "食材", "學校", "廠商", "採購叫貨"))
    assert '<span class="nav-disabled" aria-disabled="true"' in page
    assert "未來 7 天菜單" not in page

    blocked = client.get(
        "/admin/order-tool/summary/production-sheet?date=2026-08-13",
        follow_redirects=False,
    )
    assert blocked.status_code == 302
    assert "/summary/procurement?date=2026-08-13" in blocked.headers["Location"]

    protected = client.get("/admin/not-a-kitchen-page", follow_redirects=False)
    assert protected.status_code == 302
    assert "/admin/login" in protected.headers["Location"]


def test_static_pages_render_and_attendance_models_survive(app, authed_client):
    with app.app_context():
        db.session.add(Employee(id=9001, name="原打卡員工", area="A", default_break=0.5))
        db.session.add(Checkin(employee_id=9001, work_date="2026-08-13", p_type="in", ts="2026-08-13 08:00:00"))
        db.session.commit()

    for path in (
        "/admin/order-tool/",
        "/admin/order-tool/schools",
        "/admin/order-tool/suppliers",
        "/admin/order-tool/ingredients",
        "/admin/order-tool/recipes",
        "/admin/order-tool/plans",
        "/admin/order-tool/purchases",
        "/admin/order-tool/purchase",
    ):
        response = authed_client.get(path)
        assert response.status_code in (200, 302), path

    response = authed_client.get("/admin/order-tool/")
    assert b"kitchen_mobile.css" in response.data

    with app.app_context():
        assert db.session.get(Employee, 9001).name == "原打卡員工"
        assert Checkin.query.filter_by(employee_id=9001).count() == 1


def test_ingredient_supplier_can_search_existing_or_create_new(app, authed_client):
    assert authed_client.post(
        "/admin/order-tool/suppliers",
        data={"name": "搜尋得到的肉品商"},
    ).status_code == 302

    page = authed_client.get("/admin/order-tool/ingredients").get_data(as_text=True)
    assert 'name="supplier_name"' in page
    assert 'list="ingredient-supplier-options"' in page
    assert '<option value="搜尋得到的肉品商"' in page
    assert "若沒有符合項目，送出後會直接新增廠商" in page

    assert authed_client.post("/admin/order-tool/ingredients", data={
        "name": "搜尋既有廠商食材",
        "supplier_name": "搜尋得到的肉品商",
        "base_unit": "g",
        "purchase_unit": "kg",
        "grams_per_purchase_unit": "1000",
        "unit_price": "10",
        "order_increment": "0.001",
        "note": "",
    }).status_code == 302
    assert authed_client.post("/admin/order-tool/ingredients", data={
        "name": "直接新增廠商食材",
        "supplier_name": "現場新廠商",
        "base_unit": "g",
        "purchase_unit": "kg",
        "grams_per_purchase_unit": "1000",
        "unit_price": "20",
        "order_increment": "0.001",
        "note": "",
    }).status_code == 302

    with app.app_context():
        existing = KitchenSupplier.query.filter_by(name="搜尋得到的肉品商").one()
        created = KitchenSupplier.query.filter_by(name="現場新廠商").one()
        assert KitchenSupplier.query.filter_by(name="搜尋得到的肉品商").count() == 1
        assert KitchenIngredient.query.filter_by(name="搜尋既有廠商食材").one().supplier_id == existing.id
        assert KitchenIngredient.query.filter_by(name="直接新增廠商食材").one().supplier_id == created.id
        assert created.note == "由食材主檔新增"


def test_summary_is_a_monday_to_sunday_grid_and_can_add_a_dish(app, authed_client):
    ids = _seed_core_via_routes(app, authed_client)

    response = authed_client.get("/admin/order-tool/summary?week=2026-08-13")
    assert response.status_code == 200
    page = response.get_data(as_text=True)
    assert "每週菜單總表" in page
    assert all(label in page for label in ("週一", "週二", "週三", "週四", "週五", "週六", "週日"))
    assert "08/10" in page and "08/16" in page
    assert "搜尋或輸入新菜色" in page

    response = authed_client.post("/admin/order-tool/summary/dishes", data={
        "service_date": "2026-08-13",
        "week": "2026-08-10",
        "recipe_id": str(ids["recipe"]),
    })
    assert response.status_code == 302
    assert "week=2026-08-10" in response.headers["Location"]

    with app.app_context():
        plan = KitchenMenuPlan.query.filter_by(
            service_date=TEST_DAY, meal_type="午餐", name="中央菜單"
        ).one()
        item = KitchenMenuPlanItem.query.filter_by(plan_id=plan.id, recipe_id=ids["recipe"]).one()
        assert item.recipe.name == "南洋綠咖哩雞"

    page = authed_client.get("/admin/order-tool/summary?week=2026-08-10").get_data(as_text=True)
    assert "南洋綠咖哩雞" in page

    response = authed_client.post("/admin/order-tool/summary/dishes", data={
        "service_date": "2026-08-14",
        "week": "2026-08-10",
        "dish_name": "香煎鯖魚",
        "category": "主菜",
    })
    assert response.status_code == 302
    with app.app_context():
        new_recipe = KitchenRecipe.query.filter_by(name="香煎鯖魚", category="主菜").one()
        friday_plan = KitchenMenuPlan.query.filter_by(service_date=date(2026, 8, 14)).one()
        assert KitchenMenuPlanItem.query.filter_by(
            plan_id=friday_plan.id, recipe_id=new_recipe.id
        ).count() == 1


def test_summary_recipe_category_search_picker_and_delete(app, authed_client):
    ids = _seed_core_via_routes(app, authed_client)
    authed_client.post("/admin/order-tool/summary/dishes", data={
        "service_date": "2026-08-13",
        "week": "2026-08-10",
        "recipe_id": str(ids["recipe"]),
    })

    with app.app_context():
        plan = KitchenMenuPlan.query.filter_by(service_date=TEST_DAY, name="中央菜單").one()
        plan_id = plan.id
        item_id = plan.items[0].id

    recipe_page = authed_client.get(
        f"/admin/order-tool/recipes/{ids['recipe']}"
    ).get_data(as_text=True)
    assert "菜色分類" in recipe_page
    assert "選擇後自動儲存" in recipe_page
    assert 'onchange="this.form.requestSubmit()"' in recipe_page
    assert "儲存分類" not in recipe_page
    assert "輸入食材名稱搜尋" in recipe_page
    assert "data-ingredient-search" in recipe_page
    assert "ingredient-search-data" in recipe_page

    response = authed_client.post(
        f"/admin/order-tool/recipes/{ids['recipe']}/category",
        data={"category": "副菜"},
    )
    assert response.status_code == 302
    with app.app_context():
        assert db.session.get(KitchenRecipe, ids["recipe"]).category == "副菜"

    plan_page = authed_client.get(
        f"/admin/order-tool/plans/{plan_id}"
    ).get_data(as_text=True)
    assert "輸入菜名搜尋" in plan_page
    assert 'data-search-mode="select"' in plan_page
    assert "南洋綠咖哩雞" in plan_page

    summary_page = authed_client.get(
        "/admin/order-tool/summary?week=2026-08-10"
    ).get_data(as_text=True)
    assert "副菜" in summary_page
    assert f"/summary/dishes/{item_id}/delete" in summary_page
    assert "移除南洋綠咖哩雞" in summary_page

    response = authed_client.post(
        f"/admin/order-tool/summary/dishes/{item_id}/delete",
        data={"week": "2026-08-10"},
    )
    assert response.status_code == 302
    assert "week=2026-08-10" in response.headers["Location"]
    with app.app_context():
        assert db.session.get(KitchenMenuPlanItem, item_id) is None


def test_summary_next_step_builds_school_week_menu_with_headcount(app, authed_client):
    ids = _seed_core_via_routes(app, authed_client)
    authed_client.post("/admin/order-tool/summary/dishes", data={
        "service_date": "2026-08-13",
        "week": "2026-08-10",
        "recipe_id": str(ids["recipe"]),
    })

    page = authed_client.get(
        f"/admin/order-tool/summary/schools?week=2026-08-10&school_id={ids['school']}"
    ).get_data(as_text=True)
    assert "各家學校菜單" in page
    assert "南洋綠咖哩雞" in page
    assert "供餐人數" in page
    assert "產生採購單" in page
    assert "儲存本週菜單與人數" not in page
    assert "data-school-menu-autosave" in page
    assert "今日停餐" in page
    assert "data-no-service-toggle" in page
    script = authed_client.get("/static/kitchen_ui.js").get_data(as_text=True)
    assert "headcount?.addEventListener('input'" in script
    assert "等待儲存…" in script
    assert "請輸入人數" in script
    assert "window.setTimeout(saveDay, 1000)" in script

    payload = {"school_id": str(ids["school"]), "week": "2026-08-10"}
    for offset in range(7):
        day = date(2026, 8, 10) + timedelta(days=offset)
        payload[f"headcount_{day.isoformat()}"] = "0"
    payload["headcount_2026-08-13"] = "586"
    payload["recipes_2026-08-13"] = str(ids["recipe"])
    response = authed_client.post("/admin/order-tool/summary/schools/save", data=payload)
    assert response.status_code == 302

    with app.app_context():
        plan = KitchenMenuPlan.query.filter_by(
            service_date=TEST_DAY, meal_type="午餐", name="內小菜單"
        ).one()
        assignment = KitchenMenuAssignment.query.filter_by(plan_id=plan.id, school_id=ids["school"]).one()
        assert assignment.headcount == 586
        assert [item.recipe_id for item in plan.items] == [ids["recipe"]]

    plans_page = authed_client.get("/admin/order-tool/plans", follow_redirects=False)
    assert plans_page.status_code == 302
    assert "/summary" in plans_page.headers["Location"]


def test_school_menu_auto_saves_one_day_and_blocks_incomplete_procurement(app, authed_client):
    ids = _seed_core_via_routes(app, authed_client)
    authed_client.post("/admin/order-tool/summary/dishes", data={
        "service_date": "2026-08-13",
        "week": "2026-08-10",
        "recipe_id": str(ids["recipe"]),
    })
    response = authed_client.post("/admin/order-tool/summary/schools/save-day", data={
        "school_id": str(ids["school"]),
        "service_date": "2026-08-13",
        "headcount": "586",
        "recipe_ids": str(ids["recipe"]),
    })
    assert response.status_code == 204
    with app.app_context():
        assignment = KitchenMenuAssignment.query.join(KitchenMenuPlan).filter(
            KitchenMenuAssignment.school_id == ids["school"],
            KitchenMenuPlan.service_date == TEST_DAY,
        ).one()
        assert assignment.headcount == 586
        assert [item.recipe_id for item in assignment.plan.items] == [ids["recipe"]]

    assert authed_client.post("/admin/order-tool/schools", data={
        "name": "尚未勾選國小",
        "code": "2-08",
    }).status_code == 302
    blocked = authed_client.post(
        "/admin/order-tool/summary/procurement/generate",
        data={"date": "2026-08-13"},
        follow_redirects=True,
    )
    assert "尚有學校未完成菜單勾選：尚未勾選國小" in blocked.get_data(as_text=True)
    with app.app_context():
        assert KitchenPurchaseOrder.query.count() == 0


def test_school_menu_saves_regular_and_vegetarian_separately_and_combines_procurement(app, authed_client):
    ids = _seed_core_via_routes(app, authed_client)
    response = authed_client.post(f"/admin/order-tool/schools/{ids['school']}/update", data={
        "name": "內小",
        "code": "1-08",
        "default_headcount": "40",
        "default_vegetarian_headcount": "3",
    })
    assert response.status_code == 302
    schools_page = authed_client.get("/admin/order-tool/schools").get_data(as_text=True)
    assert "平常葷食人數" in schools_page
    assert "平常素食人數" in schools_page
    authed_client.post("/admin/order-tool/summary/dishes", data={
        "service_date": "2026-08-13",
        "week": "2026-08-10",
        "recipe_id": str(ids["recipe"]),
    })

    page = authed_client.get(
        f"/admin/order-tool/summary/schools?week=2026-08-10&school_id={ids['school']}"
    ).get_data(as_text=True)
    assert "data-meal-variant=\"regular\"" in page
    assert "data-meal-variant=\"vegetarian\"" in page
    assert "葷食 40 人" in page
    assert "素食 3 人" in page
    css = authed_client.get("/static/kitchen_mobile.css").get_data(as_text=True)
    assert "[data-variant-panel][hidden]{display:none!important}" in css

    response = authed_client.post("/admin/order-tool/summary/schools/save-day", data={
        "school_id": str(ids["school"]),
        "service_date": "2026-08-13",
        "headcount": "40",
        "vegetarian_headcount": "3",
        "regular_recipe_ids": str(ids["recipe"]),
        "vegetarian_recipe_ids": str(ids["recipe"]),
    })
    assert response.status_code == 204

    with app.app_context():
        assignments = KitchenMenuAssignment.query.join(KitchenMenuPlan).filter(
            KitchenMenuAssignment.school_id == ids["school"],
            KitchenMenuPlan.service_date == TEST_DAY,
        ).all()
        assert sorted(row.headcount for row in assignments) == [3, 40]
        assert {row.plan.name for row in assignments} == {"內小菜單", "內小素食菜單"}
        requirements = order_tool_module._requirements_for_date(TEST_DAY)
        ingredient_row = next(iter(next(iter(requirements.values())).values()))
        assert ingredient_row["total_people"] == 43
        assert ingredient_row["school_names"] == {"內小"}


def test_daily_production_sheet_splits_meal_variants_and_shows_purchase_total(app, authed_client):
    ids = _seed_core_via_routes(app, authed_client)
    authed_client.post("/admin/order-tool/summary/dishes", data={
        "service_date": "2026-08-13",
        "week": "2026-08-10",
        "recipe_id": str(ids["recipe"]),
    })
    saved = authed_client.post("/admin/order-tool/summary/schools/save-day", data={
        "school_id": str(ids["school"]),
        "service_date": "2026-08-13",
        "headcount": "40",
        "vegetarian_headcount": "3",
        "regular_recipe_ids": str(ids["recipe"]),
        "vegetarian_recipe_ids": str(ids["recipe"]),
    })
    assert saved.status_code == 204
    generated = authed_client.post(
        "/admin/order-tool/summary/procurement/generate",
        data={"date": "2026-08-13"},
    )
    assert generated.status_code == 302

    with app.app_context():
        item = KitchenPurchaseOrderItem.query.filter_by(ingredient_id=ids["ingredient"]).one()
        item.actual_order_qty = Decimal("9.5")
        item.package_qty = Decimal("1")
        item.package_unit = "箱"
        item.note = "冷藏，上午先到"
        db.session.commit()

    regular = authed_client.get(
        "/admin/order-tool/summary/production-sheet?date=2026-08-13&variant=regular"
    ).get_data(as_text=True)
    assert all(label in regular for label in (
        "菜色用量表", "食材", "每人用量", "供餐人數", "理論總量",
        "採購單位", "當日總採購量", "現場備註",
    ))
    assert "南洋綠咖哩雞" in regular
    assert "88" in regular and "g/人" in regular
    assert "<b>40</b> 人" in regular
    assert "<b>3.52</b> kg" in regular
    assert "<b>9.5 kg</b>" in regular
    assert "＝ 1 箱" in regular
    assert "冷藏，上午先到" in regular
    assert "匯出 Excel（葷／素分頁）" in regular
    assert "/admin/order-tool/summary/production-sheet.xlsx?date=2026-08-13" in regular

    vegetarian = authed_client.get(
        "/admin/order-tool/summary/production-sheet?date=2026-08-13&variant=vegetarian"
    ).get_data(as_text=True)
    assert "素食 <b>1</b> 道" in vegetarian
    assert "<b>3</b> 人" in vegetarian
    assert "<b>0.264</b> kg" in vegetarian

    exported = authed_client.get(
        "/admin/order-tool/summary/production-sheet.xlsx?date=2026-08-13"
    )
    assert exported.status_code == 200
    assert exported.mimetype == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    workbook = load_workbook(BytesIO(exported.data), data_only=False)
    assert workbook.sheetnames == ["葷食", "素食"]
    expected_headers = [
        "菜色類別", "菜色", "食材", "每人用量", "每人單位", "供餐人數",
        "理論總量", "採購單位", "當日總採購量", "現場備註", "供餐學校",
    ]
    assert [cell.value for cell in workbook["葷食"][4]] == expected_headers
    regular_row = [cell.value for cell in workbook["葷食"][5]]
    vegetarian_row = [cell.value for cell in workbook["素食"][5]]
    assert regular_row[:6] == ["主菜", "南洋綠咖哩雞", "骨腿丁", 88, "g/人", 40]
    assert regular_row[6] == pytest.approx(3.52)
    assert regular_row[7:11] == ["kg", 9.5, "冷藏，上午先到", "內小"]
    assert vegetarian_row[:6] == ["主菜", "南洋綠咖哩雞", "骨腿丁", 88, "g/人", 3]
    assert vegetarian_row[6] == pytest.approx(0.264)
    assert vegetarian_row[8] == pytest.approx(9.5)

    school_menu = authed_client.get(
        f"/admin/order-tool/summary/schools?week=2026-08-10&school_id={ids['school']}"
    ).get_data(as_text=True)
    assert "菜色用量" in school_menu
    assert "/admin/order-tool/summary/production-sheet?date=2026-08-13" in school_menu


def test_no_service_school_is_complete_and_excluded_from_procurement(app, authed_client):
    ids = _seed_core_via_routes(app, authed_client)
    authed_client.post("/admin/order-tool/summary/dishes", data={
        "service_date": "2026-08-13",
        "week": "2026-08-10",
        "recipe_id": str(ids["recipe"]),
    })
    authed_client.post("/admin/order-tool/summary/schools/save-day", data={
        "school_id": str(ids["school"]),
        "service_date": "2026-08-13",
        "headcount": "586",
        "service_status": "serving",
        "recipe_ids": str(ids["recipe"]),
    })
    assert authed_client.post("/admin/order-tool/schools", data={
        "name": "今日停餐國小",
        "code": "3-08",
        "default_headcount": "320",
    }).status_code == 302
    with app.app_context():
        stopped_school_id = KitchenSchool.query.filter_by(name="今日停餐國小").one().id

    response = authed_client.post("/admin/order-tool/summary/schools/save-day", data={
        "school_id": str(stopped_school_id),
        "service_date": "2026-08-13",
        "headcount": "320",
        "service_status": "no_service",
    })
    assert response.status_code == 204
    with app.app_context():
        stopped = KitchenMenuAssignment.query.join(KitchenMenuPlan).filter(
            KitchenMenuAssignment.school_id == stopped_school_id,
            KitchenMenuPlan.service_date == TEST_DAY,
        ).one()
        assert stopped.service_status == "no_service"
        assert stopped.headcount == 320
        assert stopped.plan.items == []

    generated = authed_client.post(
        "/admin/order-tool/summary/procurement/generate",
        data={"date": "2026-08-13"},
        follow_redirects=True,
    )
    assert generated.status_code == 200
    assert "尚有學校未完成菜單勾選" not in generated.get_data(as_text=True)
    with app.app_context():
        assert KitchenPurchaseOrder.query.count() == 1
        item = KitchenPurchaseOrderItem.query.one()
        assert item.required_qty == Decimal("51.568")


def test_no_service_preserves_dishes_when_toggled_back(app, authed_client):
    ids = _seed_core_via_routes(app, authed_client)
    authed_client.post("/admin/order-tool/summary/dishes", data={
        "service_date": "2026-08-13",
        "week": "2026-08-10",
        "recipe_id": str(ids["recipe"]),
    })
    for status in ("serving", "no_service", "serving"):
        response = authed_client.post("/admin/order-tool/summary/schools/save-day", data={
            "school_id": str(ids["school"]),
            "service_date": "2026-08-13",
            "headcount": "586",
            "service_status": status,
            "recipe_ids": str(ids["recipe"]),
        })
        assert response.status_code == 204

    with app.app_context():
        assignment = KitchenMenuAssignment.query.join(KitchenMenuPlan).filter(
            KitchenMenuAssignment.school_id == ids["school"],
            KitchenMenuPlan.service_date == TEST_DAY,
        ).one()
        assert assignment.service_status == "serving"
        assert assignment.headcount == 586
        assert [item.recipe_id for item in assignment.plan.items] == [ids["recipe"]]


def test_school_menu_exports_selected_date_into_nonregistered_template(app, authed_client):
    ids = _seed_core_via_routes(app, authed_client)
    plan_id = _create_plan(app, authed_client, ids)
    with app.app_context():
        plan = db.session.get(KitchenMenuPlan, plan_id)
        recipes = [
            KitchenRecipe(name="香Q米飯", category="主食"),
            KitchenRecipe(name="麻婆豆腐", category="副菜"),
            KitchenRecipe(name="有機青菜", category="青菜"),
            KitchenRecipe(name="玉米蛋花湯", category="湯品"),
            KitchenRecipe(name="當季水果", category="點心"),
        ]
        db.session.add_all(recipes)
        db.session.flush()
        for sort_order, recipe in enumerate(recipes, start=1):
            db.session.add(KitchenMenuPlanItem(
                plan_id=plan.id,
                recipe_id=recipe.id,
                sort_order=sort_order,
            ))
        stopped_school = KitchenSchool(name="停餐校", default_headcount=100)
        db.session.add(stopped_school)
        db.session.flush()
        db.session.add(KitchenMenuAssignment(
            plan_id=plan.id,
            school_id=stopped_school.id,
            headcount=100,
            service_status="no_service",
        ))
        db.session.commit()

    page = authed_client.get(
        "/admin/order-tool/summary/schools?week=2026-08-10"
    ).get_data(as_text=True)
    assert "匯出非登合菜名" in page
    assert "菜單日期" in page
    assert "選擇該校 Excel" not in page
    assert "匯入這間學校" not in page

    response = authed_client.get(
        "/admin/order-tool/summary/schools/nonregistered-menu.xlsx?date=2026-08-13"
    )
    assert response.status_code == 200
    assert "2026-08-13.xlsx" in response.headers["Content-Disposition"]
    workbook = load_workbook(BytesIO(response.data))
    sheet = workbook["Sheet1"]
    assert sheet.max_column == 27
    assert sheet["A1"].value == "學校*"
    assert sheet["AA1"].value == "附餐"
    assert sheet["A2"].value == "內小"
    assert sheet["B2"].value.date() == TEST_DAY
    assert sheet["B2"].number_format == "m/d;@"
    assert sheet["C2"].value == "午餐"
    assert [sheet.cell(2, column).value for column in range(4, 11)] == [4, 2, 1.7, 0, 0, 2, 563]
    assert sheet["K2"].value == "香Q米飯"
    assert sheet["M2"].value == "南洋綠咖哩雞"
    assert sheet["Q2"].value == "麻婆豆腐"
    assert sheet["W2"].value == "有機青菜"
    assert sheet["X2"].value == "玉米蛋花湯"
    assert sheet["Y2"].value == "當季水果"
    assert sheet["A3"].value is None
    assert sheet.column_dimensions["P"].hidden is True
    assert sheet.column_dimensions["Z"].hidden is True

    assert authed_client.post("/admin/order-tool/summary/schools/import").status_code == 404


def test_single_day_procurement_has_simple_fields_and_searchable_supplier(app, authed_client):
    ids = _seed_core_via_routes(app, authed_client)
    assert authed_client.post(f"/admin/order-tool/suppliers/{ids['supplier']}/items", data={
        "name": "骨腿丁",
        "unit": "kg",
        "package_conversion": "1箱＝10kg",
        "last_unit_price": "82",
    }).status_code == 302
    authed_client.post("/admin/order-tool/summary/dishes", data={
        "service_date": "2026-08-13",
        "week": "2026-08-10",
        "recipe_id": str(ids["recipe"]),
    })
    payload = {"school_id": str(ids["school"]), "week": "2026-08-10"}
    for offset in range(7):
        day = date(2026, 8, 10) + timedelta(days=offset)
        payload[f"headcount_{day.isoformat()}"] = "0"
    payload["headcount_2026-08-13"] = "801"
    payload["recipes_2026-08-13"] = str(ids["recipe"])
    authed_client.post("/admin/order-tool/summary/schools/save", data=payload)

    response = authed_client.post(
        "/admin/order-tool/summary/procurement/generate",
        data={"date": "2026-08-13"},
        follow_redirects=True,
    )
    page = response.get_data(as_text=True)
    assert all(label in page for label in (
        "已叫貨", "食材名稱", "總供餐人次", "系統需求量", "實際採購量／包裝換算", "交貨日期／時段", "供應廠商"
    ))
    assert 'list="supplier-search-options"' in page
    assert "data-procurement-autosave" in page
    assert "data-auto-save-url" in page
    assert "每一列修改後自動儲存" in page
    assert ">801</b> 人次" in page
    assert '<small class="included-schools"><span>包含學校（1）</span>內小</small>' in page
    assert "1箱＝10kg" in page

    with app.app_context():
        item = KitchenPurchaseOrderItem.query.one()
        item_id, order_id = item.id, item.order_id
        supplier_item = KitchenSupplierItem.query.filter_by(name="骨腿丁").one()
        assert item.supplier_item_id == supplier_item.id
        assert item.package_conversion_snapshot == "1箱＝10kg"
        assert item.package_qty == Decimal("7.0488")
    saved = authed_client.post("/admin/order-tool/summary/procurement/save", data={
        "date": "2026-08-13",
        "item_ids": str(item_id),
        f"ordered_{item_id}": "1",
        f"actual_{item_id}": "71",
        f"package_qty_{item_id}": "3",
        f"package_unit_{item_id}": "箱",
        f"delivery_date_{item_id}": "2026-08-12",
        f"delivery_slot_{item_id}": "下午",
        f"supplier_{item_id}": "測試肉品",
    })
    assert saved.status_code == 302
    with app.app_context():
        item = db.session.get(KitchenPurchaseOrderItem, item_id)
        assert item.ordered is True
        assert item.actual_order_qty == Decimal("71")
        assert item.package_qty == Decimal("3")
        assert item.package_unit == "箱"
        assert item.package_conversion_snapshot == "1箱＝23.6667kg"
        assert item.supplier_item.name == "骨腿丁"
        assert item.supplier_item.package_conversion == "1箱＝23.6667kg"
        assert item.delivery_date == date(2026, 8, 12)
        assert item.delivery_slot == "下午"
        assert item.supplier.name == "測試肉品"
        assert KitchenPurchaseOrder.query.filter_by(service_date=TEST_DAY).count() == 1

    history = authed_client.get("/admin/order-tool/purchases?start=2026-08-13&end=2026-08-13").get_data(as_text=True)
    assert "全部已叫" in history and "1 / 1" in history
    detail = authed_client.get(f"/admin/order-tool/purchases/{order_id}").get_data(as_text=True)
    assert "＝ <b>3 箱</b>" in detail
    assert "廠商換算：1箱＝23.6667kg" in detail

    procurement_page = authed_client.get(
        "/admin/order-tool/summary/procurement?date=2026-08-13"
    ).get_data(as_text=True)
    assert "匯出 Excel" in procurement_page
    assert "儲存採購明細" not in procurement_page
    exported = authed_client.get(
        "/admin/order-tool/summary/procurement.xlsx?date=2026-08-13"
    )
    assert exported.status_code == 200
    assert exported.headers["Content-Disposition"].endswith(".xlsx")
    workbook = load_workbook(BytesIO(exported.data))
    sheet = workbook["每日訂購單"]
    assert sheet["A1"].value == "每日訂購單"
    assert sheet["C1"].value == "進貨日期：115/08/12 下午"
    assert [cell.value for cell in sheet[2]] == ["廠商", "品名", "數量", "單位", "備註"]
    assert sheet["A3"].value == "測試肉品"
    assert sheet["B3"].value == "骨腿丁"
    assert sheet["C3"].value == 3
    assert sheet["C3"].number_format == "0.00"
    assert sheet["D3"].value == "箱"
    assert sheet["E3"].value == "1箱＝23.6667kg"
    assert sheet["A3"].fill.fgColor.rgb == "00FFF1A8"
    assert sheet.print_title_rows == "$2:$2"
    assert sheet.page_setup.fitToWidth == 1

    assert authed_client.post(f"/admin/order-tool/purchase-items/{item_id}/ordered", data={}).status_code == 302
    with app.app_context():
        assert db.session.get(KitchenPurchaseOrderItem, item_id).ordered is False
    assert authed_client.post(
        f"/admin/order-tool/purchase-items/{item_id}/ordered",
        data={"ordered": "1"},
        headers={"X-Requested-With": "procurement-tracking"},
    ).status_code == 204
    assert authed_client.post(f"/admin/order-tool/purchases/{order_id}/ordered", data={"ordered": "1"}).status_code == 302
    with app.app_context():
        assert db.session.get(KitchenPurchaseOrderItem, item_id).ordered is True


def test_procurement_autosave_creates_suppliers_and_groups_every_output(app, authed_client):
    ids = _seed_core_via_routes(app, authed_client)
    assert authed_client.post("/admin/order-tool/ingredients", data={
        "name": "洋蔥",
        "supplier_id": "",
        "base_unit": "g",
        "purchase_unit": "kg",
        "grams_per_purchase_unit": "1000",
        "unit_price": "25",
        "order_increment": "0.001",
        "note": "",
    }).status_code == 302
    with app.app_context():
        onion_id = KitchenIngredient.query.filter_by(name="洋蔥").one().id
    assert authed_client.post(f"/admin/order-tool/recipes/{ids['recipe']}/ingredients", data={
        "ingredient_id": str(onion_id),
        "grams_per_person": "12",
    }).status_code == 302
    authed_client.post("/admin/order-tool/summary/dishes", data={
        "service_date": "2026-08-13",
        "week": "2026-08-10",
        "recipe_id": str(ids["recipe"]),
    })
    authed_client.post("/admin/order-tool/summary/schools/save-day", data={
        "school_id": str(ids["school"]),
        "service_date": "2026-08-13",
        "headcount": "100",
        "service_status": "serving",
        "recipe_ids": str(ids["recipe"]),
    })
    authed_client.post("/admin/order-tool/summary/procurement/generate", data={"date": "2026-08-13"})
    with app.app_context():
        items = {item.ingredient_name_snapshot: item.id for item in KitchenPurchaseOrderItem.query.all()}
        order_id = KitchenPurchaseOrder.query.one().id

    for ingredient_name, supplier_name in (("骨腿丁", "乙供應商"), ("洋蔥", "甲供應商")):
        response = authed_client.post(
            f"/admin/order-tool/summary/procurement/items/{items[ingredient_name]}/save",
            data={
                "actual": "10",
                "package_qty": "",
                "package_unit": "",
                "delivery_date": "2026-08-12",
                "delivery_slot": "上午",
                "supplier_name": supplier_name,
            },
        )
        assert response.status_code == 200
        assert response.json["supplierCreated"] is True

    # 包裝單位是選填；把原本不同廠商改成同一家後應立即歸在同一組。
    response = authed_client.post(
        f"/admin/order-tool/summary/procurement/items/{items['洋蔥']}/save",
        data={
            "actual": "10",
            "package_qty": "6",
            "package_unit": "",
            "delivery_date": "2026-08-12",
            "delivery_slot": "上午",
            "supplier_name": "乙供應商",
        },
    )
    assert response.status_code == 200
    assert response.json["supplierCreated"] is False
    assert response.json["packageQty"] == "6"
    assert response.json["packageUnit"] == ""

    with app.app_context():
        assert KitchenSupplier.query.filter_by(name="甲供應商", active=True).one()
        assert KitchenSupplier.query.filter_by(name="乙供應商", active=True).one()
        onion_item = db.session.get(KitchenPurchaseOrderItem, items["洋蔥"])
        assert onion_item.supplier.name == "乙供應商"
        assert onion_item.package_qty == Decimal("6.0000")
        assert onion_item.package_unit is None

    procurement_page = authed_client.get(
        "/admin/order-tool/summary/procurement?date=2026-08-13"
    ).get_data(as_text=True)
    assert '<select class="package-unit-input"' in procurement_page
    assert '<option value="">不指定</option>' in procurement_page
    assert procurement_page.count('<tr class="supplier-group-row"><td colspan="8">乙供應商') == 1
    assert '<tr class="supplier-group-row"><td colspan="8">甲供應商' not in procurement_page

    response = authed_client.post(
        f"/admin/order-tool/summary/procurement/items/{items['洋蔥']}/save",
        data={
            "actual": "10",
            "package_qty": "6",
            "package_unit": "",
            "delivery_date": "2026-08-12",
            "delivery_slot": "下午",
            "supplier_name": "乙供應商",
        },
    )
    assert response.status_code == 200

    exported = authed_client.get("/admin/order-tool/summary/procurement.xlsx?date=2026-08-13")
    workbook = load_workbook(BytesIO(exported.data))
    sheet = workbook["每日訂購單"]
    assert sheet["C1"].value == "進貨日期：115/08/12 上午"
    assert [cell.value for cell in sheet[2]] == ["廠商", "品名", "數量", "單位", "備註"]
    assert [sheet["A3"].value, sheet["B3"].value] == ["乙供應商", "骨腿丁"]
    assert sheet["C5"].value == "進貨日期：115/08/12 下午"
    assert [cell.value for cell in sheet[6]] == ["廠商", "品名", "數量", "單位", "備註"]
    assert [sheet["A7"].value, sheet["B7"].value] == ["乙供應商", "洋蔥"]
    assert "進貨：" not in (sheet["E3"].value or "")
    assert "進貨：" not in (sheet["E7"].value or "")

    history = authed_client.get(
        "/admin/order-tool/purchases?start=2026-08-13&end=2026-08-13"
    ).get_data(as_text=True)
    assert "乙供應商" in history
    assert "乙供應商、甲供應商" not in history
    detail = authed_client.get(f"/admin/order-tool/purchases/{order_id}").get_data(as_text=True)
    assert detail.count('<tr class="supplier-group-row"><td colspan="8">乙供應商') == 1


def _menu_upload_file():
    workbook = Workbook()
    hidden = workbook.active
    hidden.title = "舊菜單"
    hidden.sheet_state = "hidden"
    hidden.append(["日期", "星期", "主食", "主菜"])
    hidden.append(["6/1", "一", "不應匯入的舊菜", "舊主菜"])

    sheet = workbook.create_sheet("6月", 0)
    sheet["A1"] = "中平國小115年6月菜單"
    sheet.append(["日期", "星期", "主食", "主菜", "副菜", "", "蔬菜", "湯品", "全穀雜糧類(份)"])
    sheet.append(["6/1", "一", "白米飯", "沙茶雞肉", "麻婆豆腐", "麻婆豆腐", "有機蔬菜", "玉米蛋花湯", 5])
    sheet.append([None, None, "白米/煮", "雞肉/燒", "豆腐/煮", "豆腐/煮", "青菜/炒", "玉米/煮", None])
    sheet.append([date(2026, 6, 2), "二", "糙米飯", "日式炸豬排", "日式蒸蛋", "金茸蒲瓜", "有機蔬菜", "肉骨茶湯", 5])
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def _menu_upload_file_with_regular_and_vegetarian():
    workbook = Workbook()
    regular = workbook.active
    regular.title = "中平一般菜單"
    regular.append(["日期", "星期", "主食", "主菜", "副菜", "青菜", "湯品"])
    regular.append(["8/31", "一", "白米飯", "馬鈴薯燉肉", "炒四季豆", "有機蔬菜", "味噌湯"])

    vegetarian = workbook.create_sheet("中平菜單 (素食)")
    vegetarian.append(["日期", "星期", "主食", "主菜", "副菜", "青菜", "湯品"])
    vegetarian.append(["8/31", "一", "糙米飯", "馬鈴薯豆腸", "炒毛豆", "季節蔬菜", "紫菜湯"])

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def test_menu_parser_reads_cells_without_loading_drawings(monkeypatch):
    real_load_workbook = order_tool_module.load_workbook
    observed = {}

    def checked_load_workbook(*args, **kwargs):
        observed.update(kwargs)
        return real_load_workbook(*args, **kwargs)

    monkeypatch.setattr(order_tool_module, "load_workbook", checked_load_workbook)
    parsed = order_tool_module.parse_menu_workbook(
        _menu_upload_file_with_regular_and_vegetarian().read(),
        "中平115年8月菜單.xlsx",
    )

    assert observed["read_only"] is True
    assert parsed["sheet_names"] == ["中平一般菜單"]


def test_summary_import_can_choose_regular_or_vegetarian_sheet(app, authed_client):
    regular_response = authed_client.post(
        "/admin/order-tool/summary/import",
        data={
            "menu_file": (_menu_upload_file_with_regular_and_vegetarian(), "中平115年8月菜單.xlsx"),
            "menu_sheet_kind": "regular",
        },
        content_type="multipart/form-data",
        follow_redirects=True,
    )
    assert "中平一般菜單" in regular_response.get_data(as_text=True)
    with app.app_context():
        assert KitchenRecipe.query.filter_by(name="馬鈴薯燉肉").count() == 1
        assert KitchenRecipe.query.filter_by(name="馬鈴薯豆腸").count() == 0

    vegetarian_response = authed_client.post(
        "/admin/order-tool/summary/import",
        data={
            "menu_file": (_menu_upload_file_with_regular_and_vegetarian(), "中平115年8月菜單.xlsx"),
            "menu_sheet_kind": "vegetarian",
        },
        content_type="multipart/form-data",
        follow_redirects=True,
    )
    assert "中平菜單 (素食)" in vegetarian_response.get_data(as_text=True)
    with app.app_context():
        assert KitchenRecipe.query.filter_by(name="馬鈴薯豆腸").count() == 1


def test_summary_import_detects_template_dates_and_deduplicates(app, authed_client):
    response = authed_client.post(
        "/admin/order-tool/summary/import",
        data={"menu_file": (_menu_upload_file(), "中平國小115年6月菜單.xlsx")},
        content_type="multipart/form-data",
        follow_redirects=True,
    )
    assert response.status_code == 200
    page = response.get_data(as_text=True)
    assert "匯入完成" in page
    assert "2026/06/01 — 2026/06/07" in page

    with app.app_context():
        monday = KitchenMenuPlan.query.filter_by(service_date=date(2026, 6, 1)).one()
        tuesday = KitchenMenuPlan.query.filter_by(service_date=date(2026, 6, 2)).one()
        assert [item.recipe.name for item in monday.items] == [
            "白米飯", "沙茶雞肉", "麻婆豆腐", "有機蔬菜", "玉米蛋花湯"
        ]
        assert len(tuesday.items) == 6
        assert KitchenRecipe.query.filter_by(name="有機蔬菜").one().category == "青菜"
        assert KitchenRecipe.query.filter_by(name="不應匯入的舊菜").count() == 0
        original_item_count = KitchenMenuPlanItem.query.count()

    second = authed_client.post(
        "/admin/order-tool/summary/import",
        data={"menu_file": (_menu_upload_file(), "中平國小115年6月菜單.xlsx")},
        content_type="multipart/form-data",
        follow_redirects=True,
    )
    assert "略過" in second.get_data(as_text=True)
    with app.app_context():
        assert KitchenMenuPlanItem.query.count() == original_item_count


def test_summary_import_rejects_unknown_template_without_writing(app, authed_client):
    workbook = Workbook()
    workbook.active.append(["這是一個還沒支援的格式", "內容"])
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    response = authed_client.post(
        "/admin/order-tool/summary/import",
        data={"menu_file": (output, "陌生模板.xlsx")},
        content_type="multipart/form-data",
        follow_redirects=True,
    )
    assert "請提供這種格式作為新模板" in response.get_data(as_text=True)
    with app.app_context():
        assert KitchenMenuPlan.query.count() == 0


def test_full_801_person_purchase_calculation_and_snapshot(app, authed_client):
    ids = _seed_core_via_routes(app, authed_client)
    plan_id = _create_plan(app, authed_client, ids, headcount=801)

    response = authed_client.post("/admin/order-tool/purchases/generate", data={
        "start": "2026-08-13",
        "end": "2026-08-13",
    })
    assert response.status_code == 302

    with app.app_context():
        order = KitchenPurchaseOrder.query.filter_by(service_date=TEST_DAY).one()
        item = order.items[0]
        assert item.required_grams == Decimal("70488.000")
        assert item.required_qty == Decimal("70.4880")
        assert item.recommended_order_qty == Decimal("70.4880")
        assert item.actual_order_qty == Decimal("70.4880")
        assert item.unit_price_snapshot == Decimal("82.0000")
        assert item.amount == Decimal("5780.0160")
        order_id = order.id

    detail = authed_client.get(f"/admin/order-tool/purchases/{order_id}")
    assert detail.status_code == 200
    assert b"70.488" in detail.data
    assert b"5780.02" in detail.data

    assert authed_client.post(f"/admin/order-tool/purchases/{order_id}/confirm").status_code == 302

    # 修改 master data 後，已確認歷史採購單必須保持原 snapshot。
    with app.app_context():
        ingredient = db.session.get(KitchenIngredient, ids["ingredient"])
        ingredient.unit_price = Decimal("90")
        component = KitchenRecipeIngredient.query.filter_by(recipe_id=ids["recipe"], ingredient_id=ids["ingredient"]).one()
        component.grams_per_person = Decimal("90")
        db.session.commit()

        order = db.session.get(KitchenPurchaseOrder, order_id)
        item = order.items[0]
        assert order.status == "confirmed"
        assert item.required_grams == Decimal("70488.000")
        assert item.unit_price_snapshot == Decimal("82.0000")
        assert item.amount == Decimal("5780.0160")
        assert db.session.get(KitchenMenuPlan, plan_id).status == "confirmed"

    authed_client.post("/admin/order-tool/purchases/generate", data={"start": "2026-08-13", "end": "2026-08-13"})
    with app.app_context():
        item = KitchenPurchaseOrder.query.filter_by(service_date=TEST_DAY).one().items[0]
        assert item.required_grams == Decimal("70488.000")
        assert item.unit_price_snapshot == Decimal("82.0000")


def test_cancelled_purchase_is_reused_when_regenerated(app, authed_client):
    ids = _seed_core_via_routes(app, authed_client)
    _create_plan(app, authed_client, ids, headcount=801)
    authed_client.post("/admin/order-tool/purchases/generate", data={
        "start": "2026-08-13", "end": "2026-08-13",
    })

    with app.app_context():
        order = KitchenPurchaseOrder.query.one()
        order_id = order.id

    authed_client.post(f"/admin/order-tool/purchases/{order_id}/cancel")
    response = authed_client.post("/admin/order-tool/purchases/generate", data={
        "start": "2026-08-13", "end": "2026-08-13",
    })
    assert response.status_code == 302

    with app.app_context():
        orders = KitchenPurchaseOrder.query.all()
        assert len(orders) == 1
        assert orders[0].id == order_id
        assert orders[0].status == "draft"


def test_purchase_order_delete_requires_confirmation_and_removes_items(app, authed_client):
    ids = _seed_core_via_routes(app, authed_client)
    _create_plan(app, authed_client, ids, headcount=801)
    authed_client.post("/admin/order-tool/purchases/generate", data={
        "start": "2026-08-13", "end": "2026-08-13",
    })

    with app.app_context():
        order = KitchenPurchaseOrder.query.one()
        order_id = order.id
        assert KitchenPurchaseOrderItem.query.filter_by(order_id=order_id).count() == 1

    dashboard = authed_client.get("/admin/order-tool/").get_data(as_text=True)
    history = authed_client.get(
        "/admin/order-tool/purchases?start=2026-08-13&end=2026-08-13"
    ).get_data(as_text=True)
    delete_url = f"/admin/order-tool/purchases/{order_id}/delete"
    assert delete_url in dashboard
    assert delete_url in history
    assert "此動作無法復原" in dashboard
    assert "此動作無法復原" in history

    refused = authed_client.post(
        delete_url,
        data={"return_to": "history", "start": "2026-08-13", "end": "2026-08-13"},
        follow_redirects=True,
    )
    assert "未完成刪除確認，採購單已保留" in refused.get_data(as_text=True)
    with app.app_context():
        assert db.session.get(KitchenPurchaseOrder, order_id) is not None

    deleted = authed_client.post(
        delete_url,
        data={
            "confirm_delete": "1",
            "return_to": "history",
            "start": "2026-08-13",
            "end": "2026-08-13",
        },
        follow_redirects=True,
    )
    assert "已刪除 2026-08-13 的採購單與 1 筆採購品項" in deleted.get_data(as_text=True)
    with app.app_context():
        assert db.session.get(KitchenPurchaseOrder, order_id) is None
        assert KitchenPurchaseOrderItem.query.filter_by(order_id=order_id).count() == 0


def test_cross_school_aggregation(app, authed_client):
    ids = _seed_core_via_routes(app, authed_client)
    plan_id = _create_plan(app, authed_client, ids, headcount=801)

    authed_client.post("/admin/order-tool/schools", data={"name": "文山", "code": "1-21"})
    with app.app_context():
        second_school_id = KitchenSchool.query.filter_by(name="文山").one().id
    authed_client.post(f"/admin/order-tool/plans/{plan_id}/assignments", data={
        "school_id": str(second_school_id),
        "headcount": "586",
    })
    authed_client.post("/admin/order-tool/purchases/generate", data={"start": "2026-08-13", "end": "2026-08-13"})

    with app.app_context():
        item = KitchenPurchaseOrder.query.filter_by(service_date=TEST_DAY).one().items[0]
        assert item.required_grams == Decimal("122056.000")
        assert item.required_qty == Decimal("122.0560")


def test_supplier_grouping(app, authed_client):
    ids = _seed_core_via_routes(app, authed_client)
    _create_plan(app, authed_client, ids, headcount=801)

    authed_client.post("/admin/order-tool/suppliers", data={"name": "測試蔬菜"})
    with app.app_context():
        veg_supplier = KitchenSupplier.query.filter_by(name="測試蔬菜").one()
        veg_supplier_id = veg_supplier.id
    authed_client.post("/admin/order-tool/ingredients", data={
        "name": "洋芋",
        "supplier_id": str(veg_supplier_id),
        "purchase_unit": "kg",
        "grams_per_purchase_unit": "1000",
        "unit_price": "20",
        "order_increment": "0.001",
    })
    with app.app_context():
        potato_id = KitchenIngredient.query.filter_by(name="洋芋").one().id
    authed_client.post(f"/admin/order-tool/recipes/{ids['recipe']}/ingredients", data={
        "ingredient_id": str(potato_id),
        "grams_per_person": "8",
    })
    authed_client.post("/admin/order-tool/purchases/generate", data={"start": "2026-08-13", "end": "2026-08-13"})

    with app.app_context():
        orders = KitchenPurchaseOrder.query.filter_by(service_date=TEST_DAY).all()
        assert len(orders) == 1
        assert orders[0].supplier_name_snapshot == "每日採購單"
        assert {item.supplier_name_snapshot for item in orders[0].items} == {"測試肉品", "測試蔬菜"}
        potato_item = next(item for item in orders[0].items if item.ingredient_name_snapshot == "洋芋")
        assert potato_item.required_grams == Decimal("6408.000")


def test_actual_purchase_qty_and_price_are_persisted(app, authed_client):
    ids = _seed_core_via_routes(app, authed_client)
    _create_plan(app, authed_client, ids)
    authed_client.post("/admin/order-tool/purchases/generate", data={"start": "2026-08-13", "end": "2026-08-13"})

    with app.app_context():
        order = KitchenPurchaseOrder.query.one()
        item = order.items[0]
        order_id, item_id = order.id, item.id

    authed_client.post(f"/admin/order-tool/purchase-items/{item_id}/update", data={
        "actual_order_qty": "72",
        "unit_price": "83.5",
        "note": "人工調整",
    })

    with app.app_context():
        item = db.session.get(KitchenPurchaseOrderItem, item_id)
        assert item.actual_order_qty == Decimal("72.0000")
        assert item.unit_price_snapshot == Decimal("83.5000")
        assert item.amount == Decimal("6012.0000")
        assert item.note == "人工調整"
        assert db.session.get(KitchenPurchaseOrder, order_id).status == "draft"


def test_invalid_negative_values_do_not_mutate(app, authed_client):
    ids = _seed_core_via_routes(app, authed_client)
    plan_id = _create_plan(app, authed_client, ids, headcount=801)
    with app.app_context():
        assignment = KitchenMenuAssignment.query.filter_by(plan_id=plan_id, school_id=ids["school"]).one()
        assignment_id = assignment.id

    authed_client.post(f"/admin/order-tool/assignments/{assignment_id}/update", data={"headcount": "-5"})
    with app.app_context():
        assert db.session.get(KitchenMenuAssignment, assignment_id).headcount == 801

    response = authed_client.post("/admin/order-tool/ingredients", data={
        "name": "錯誤食材",
        "purchase_unit": "kg",
        "grams_per_purchase_unit": "1000",
        "unit_price": "-1",
        "order_increment": "0.001",
    })
    assert response.status_code == 302
    with app.app_context():
        assert KitchenIngredient.query.filter_by(name="錯誤食材").first() is None


def test_csrf_protects_kitchen_post(tmp_path):
    db_path = tmp_path / "csrf.db"
    app = create_app({
        "TESTING": True,
        "SECRET_KEY": "csrf-test-secret",
        "SQLALCHEMY_DATABASE_URI": f"sqlite:///{db_path}",
        "AUTO_CREATE_DB": True,
        "KITCHEN_CSRF_ENABLED": True,
        "PRODUCTION": False,
    })
    client = app.test_client()
    with client.session_transaction() as sess:
        sess["role"] = "mgr"

    client.post("/admin/order-tool/schools", data={"name": "不該成功"})
    with app.app_context():
        assert KitchenSchool.query.filter_by(name="不該成功").first() is None

    client.get("/admin/order-tool/schools")
    with client.session_transaction() as sess:
        token = sess["_kitchen_csrf"]
    client.post("/admin/order-tool/schools", data={"name": "合法學校", "_csrf_token": token})
    with app.app_context():
        assert KitchenSchool.query.filter_by(name="合法學校").one()
        db.drop_all()


def test_file_sqlite_persists_across_app_restart(tmp_path):
    db_path = tmp_path / "restart.db"
    uri = f"sqlite:///{db_path}"
    config = {
        "TESTING": True,
        "SECRET_KEY": "restart-test",
        "SQLALCHEMY_DATABASE_URI": uri,
        "AUTO_CREATE_DB": True,
        "KITCHEN_CSRF_ENABLED": False,
        "PRODUCTION": False,
    }
    app1 = create_app(config)
    with app1.app_context():
        db.session.add(KitchenSchool(name="重啟後還在"))
        db.session.commit()
        db.session.remove()

    app2 = create_app({**config, "AUTO_CREATE_DB": False})
    with app2.app_context():
        assert KitchenSchool.query.filter_by(name="重啟後還在").one()
        db.drop_all()
