from datetime import date, datetime
from decimal import Decimal
from io import BytesIO

import pytest
from openpyxl import load_workbook

from app import create_app
from extensions import db
from models import (
    KitchenIngredient,
    KitchenMenuAssignment,
    KitchenMenuPlan,
    KitchenMenuPlanItem,
    KitchenPurchaseOrder,
    KitchenPurchaseOrderItem,
    KitchenRecipe,
    KitchenRecipeIngredient,
    KitchenSchool,
    KitchenSupplier,
)

TEST_DAY = date(2026, 8, 24)


@pytest.fixture()
def app(tmp_path):
    db_path = tmp_path / "school_ingredient_export.db"
    app = create_app({
        "TESTING": True,
        "SECRET_KEY": "test-only-secret",
        "SQLALCHEMY_DATABASE_URI": f"sqlite:///{db_path}",
        "AUTO_CREATE_DB": True,
        "KITCHEN_CSRF_ENABLED": False,
        "ADMIN_HR_PASSWORD": "test-hr-password",
        "ADMIN_MGR_PASSWORD": "test-mgr-password",
        "PRODUCTION": False,
    })
    yield app
    with app.app_context():
        db.session.remove()
        db.drop_all()


@pytest.fixture()
def client(app):
    return app.test_client()


def _as_date(value):
    return value.date() if isinstance(value, datetime) else value


def test_school_ingredient_export_uses_original_template_and_current_procurement_supplier(app, client):
    with app.app_context():
        school = KitchenSchool(name="桃園市中壢區中平國小", default_headcount=529)
        supplier = KitchenSupplier(name="五五五生鮮豬肉行", active=True)
        ingredient = KitchenIngredient(
            name="絞肉",
            supplier=supplier,
            base_unit="g",
            purchase_unit="kg",
            grams_per_purchase_unit=Decimal("1000"),
            unit_price=Decimal("0"),
            order_increment=Decimal("0.001"),
        )
        recipe = KitchenRecipe(name="麻婆豆腐", category="主菜")
        db.session.add_all([school, supplier, ingredient, recipe])
        db.session.flush()
        db.session.add(KitchenRecipeIngredient(
            recipe_id=recipe.id,
            ingredient_id=ingredient.id,
            grams_per_person=Decimal("7"),
        ))
        plan = KitchenMenuPlan(service_date=TEST_DAY, meal_type="午餐", name="中平國小菜單", status="draft")
        db.session.add(plan)
        db.session.flush()
        db.session.add_all([
            KitchenMenuPlanItem(plan_id=plan.id, recipe_id=recipe.id, sort_order=0),
            KitchenMenuAssignment(
                plan_id=plan.id,
                school_id=school.id,
                headcount=529,
                service_status="serving",
            ),
        ])

        order = KitchenPurchaseOrder(
            service_date=TEST_DAY,
            supplier_key="daily",
            supplier_name_snapshot="每日採購單",
            supplier_overridden=False,
            status="draft",
        )
        db.session.add(order)
        db.session.flush()
        db.session.add(KitchenPurchaseOrderItem(
            order_id=order.id,
            ingredient_id=ingredient.id,
            supplier_id=supplier.id,
            supplier_name_snapshot=supplier.name,
            ingredient_name_snapshot=ingredient.name,
            base_unit_snapshot="g",
            required_grams=Decimal("3703"),
            required_qty=Decimal("3.703"),
            purchase_unit_snapshot="kg",
            grams_per_purchase_unit_snapshot=Decimal("1000"),
            recommended_order_qty=Decimal("3.703"),
            actual_order_qty=Decimal("3.703"),
            unit_price_snapshot=Decimal("0"),
            amount=Decimal("0"),
        ))
        db.session.commit()

    response = client.get("/admin/order-tool/summary/school-ingredient.xlsx?date=2026-08-24")
    assert response.status_code == 200
    assert "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" in response.content_type

    workbook = load_workbook(BytesIO(response.data), data_only=True)
    sheet = workbook.active
    assert sheet.title == "Sheet1"
    assert [sheet.cell(1, column).value for column in range(1, 22)] == [
        "供餐日期", "學校", "菜色名稱", "食材名稱", "進貨日期", "生產日期", "有效日期", "批號",
        "製造商", "供應商名稱", "食材驗證標章", "驗證號碼", "產品名稱", "重量(公斤)",
        "非基改玉米", "非基改黃豆", "加工品", "食材原產地", None, None, None,
    ]
    row = [sheet.cell(2, column).value for column in range(1, 22)]
    assert _as_date(row[0]) == TEST_DAY
    assert row[1:4] == ["桃園市中壢區中平國小", "麻婆豆腐", "絞肉"]
    assert _as_date(row[4]) == TEST_DAY
    assert row[5:8] == [None, None, None]
    assert row[8:10] == ["五五五生鮮豬肉行", "五五五生鮮豬肉行"]
    assert row[10:13] == ["生產追溯-豬肉", "LE300431", None]
    assert row[13] == pytest.approx(3.703)
    assert row[14:18] == ["Y", "Y", "N", "臺灣"]
    assert row[18:21] == [None, None, None]
